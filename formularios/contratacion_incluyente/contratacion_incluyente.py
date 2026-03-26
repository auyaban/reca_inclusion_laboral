import os
import json
import time
import shutil
from datetime import datetime
from difflib import SequenceMatcher
from functools import lru_cache

from formularios.evaluacion_programa import evaluacion_accesibilidad
from formularios.common import (
    _build_process_output_path,
    _get_desktop_dir,
    _next_available_file_path,
    _normalize_cedula,
    _normalize_decimal_value,
    _normalize_text,
    _parse_date_value,
    _coerce_excel_decimal_value,
    sanitize_logo_error_cells,
    autofit_rows,
    clear_written_rows,
    ws_write,
    _sanitize_filename,
    _supabase_get,
    _supabase_upsert_with_queue,
)
from logging_utils import log_excel_event
from version_info import resource_path


FORM_ID = "contratacion_incluyente"
FORM_NAME = "Contratacion Incluyente"
TEMPLATE_VARIANT_INDIVIDUAL = "individual"
TEMPLATE_VARIANT_GROUP_2_PLUS = "group_2_plus"

SHEET_NAME_BY_VARIANT = {
    TEMPLATE_VARIANT_INDIVIDUAL: "5. PROCESO CONTRATACION INCLUYE",
    TEMPLATE_VARIANT_GROUP_2_PLUS: "5. PROCESO CONTRATACION INCLUYE",
}

TEMPLATE_FILENAME_BY_VARIANT = {
    TEMPLATE_VARIANT_INDIVIDUAL: "contratacion_incluyente.xlsx",
    TEMPLATE_VARIANT_GROUP_2_PLUS: "contratacion_incluyente.xlsx",
}

FORM_CACHE = {}
SECTION_1_CACHE = {}

DISCAPACIDAD_OPTIONS = [
    "Discapacidad visual pérdida total de la visión",
    "Discapacidad visual baja visión",
    "Discapacidad auditiva",
    "Discapacidad auditiva hipoacusia",
    "Trastorno de espectro autista",
    "Discapacidad intelectual",
    "Discapacidad física",
    "Discapacidad física usuario en silla de ruedas",
    "Discapacidad psicosocial",
    "Discapacidad múltiple",
    "No aplica",
]

GENERO_OPTIONS = ["Binario", "No binario", "Otro"]

_DISCAPACIDAD_CATEGORIA_MAP = {
    "discapacidad visual perdida total de la vision": "Visual",
    "discapacidad visual baja vision": "Visual",
    "discapacidad auditiva": "Auditiva",
    "discapacidad auditiva hipoacusia": "Auditiva",
    "trastorno de espectro autista": "Intelectual",
    "discapacidad intelectual": "Intelectual",
    "discapacidad fisica": "Física",
    "discapacidad fisica usuario en silla de ruedas": "Física",
    "discapacidad psicosocial": "Psicosocial",
    "discapacidad multiple": "Múltiple",
    "no aplica": None,
}

LGTBIQ_OPTIONS = ["Si", "No", "No aplica", "Prefiere no responder"]
GRUPO_ETNICO_OPTIONS = ["Si", "No", "No aplica", "Prefiere no responder"]
GRUPO_ETNICO_CUAL_OPTIONS = [
    "Afocolombiano",
    "Afrodescendiente",
    "Rom o Gitano",
    "Indígena",
    "Palenquero de San Basilio",
    "Otro",
    "No aplica",
    "Mulato",
    "Autorreconocimiento",
    "Pueblo Indígena",
    "Negro",
    "Raizal del Archipiélago de San Andrés Y Providencia",
]
CERTIFICADO_DISCAPACIDAD_OPTIONS = ["Si", "No", "No aplica"]
TIPO_CONTRATO_FIRMADO_OPTIONS = [
    "Contrato por obra o labor",
    "Contrato de trabajo a término fijo",
    "Contrato de trabajo a término indefinido",
    "Contrato de aprendizaje",
    "Contrato temporal",
    "Contrato a término indefinido con orden clausulada",
    "Contrato a término fijo a un año",
    "Contrato a término fijo a seis meses",
    "Contrato por prestación de servicios",
]
TIPO_CONTRATO_OPTIONS = TIPO_CONTRATO_FIRMADO_OPTIONS
CONTRATO_TIPO_CONTRATO_OPTIONS = [
    "Contrato a término indefinido.",
    "Contrato a término fijo.",
    "Contrato por obra o labor.",
]
NIVEL_APOYO_OPTIONS = [
    "0. No requiere apoyo.",
    "1. Nivel de apoyo Bajo.",
    "2. Nivel de apoyo medio.",
    "3. Nivel de apoyo alto.",
    "No aplica.",
]
OBS_LECTURA_CONTRATO_OPTIONS = [
    "1. Se acompaña en la lectura del contrato.",
    "2. Se apoya en la lectura del contrato.",
    "3. Cuando requiere un apoyo adicional al del gestor (lector de pantalla, intérprete LSC u otro).",
    "No aplica.",
    "0. No requiere apoyo.",
]
OBS_LECTURA_CONTRATO_OPTIONS_GROUP = [
    "1. Se acompaña en la lectura del contrato.",
    "2. Se apoya en la lectura del contrato.",
    "3. Cuando requiere un apoyo adicional al del gestor (lector de pantalla, intérprete LSC u otro).",
    "No aplica.",
]
OBS_LECTURA_CONTRATO_OPTIONS_INDIVIDUAL = [
    "1. Se acompaña en la lectura del contrato.",
    "2. Se apoya en la lectura del contrato.",
    "3. Cuando requiere un apoyo adicional al del gestor (lector de pantalla, intérprete LSC u otro).",
    "No aplica.",
    "0. No requiere apoyo.",
]
OBS_COMPRENDE_CONTRATO_OPTIONS = [
    "1. Comprende la información, pero no se familiariza con las características del contrato.",
    "2. Explicación de algunas cláusulas del contrato.",
    "3. Explicación total del contrato.",
    "0. Comprende con claridad el contrato.",
]
OBS_TIPO_CONTRATO_OPTIONS = [
    "1. El vinculado reconoce el tipo de contrato, pero no comprende sus condiciones.",
    "2. El vinculado requiere aclaración de algunas de las condiciones del contrato.",
    "3. El vinculado no conoce ninguna de las condiciones del tipo de contrato a firmar.",
    "0. El vinculado tiene claras las condiciones del tipo de contrato a firmar.",
]
JORNADA_LABORAL_OPTIONS = ["Tiempo Completo.", "Medio Tiempo.", "Por horas."]
CLAUSULAS_CONTRATO_OPTIONS = [
    "Cláusula de confidencialidad.",
    "Cláusulas adicionales.",
]
OBS_CONDICIONES_SALARIALES_OPTIONS = [
    "1. Se aclaran las condiciones salariales asignadas al cargo.",
    "2. Se explica de manera parcial las condiciones salariales asignadas al cargo.",
    "3. Se explica de manera completa las condiciones salariales asignadas al cargo.",
    "0. Tiene claras las condiciones salariales asignadas al cargo.",
]
FRECUENCIA_PAGO_OPTIONS = ["Pago Semanal.", "Pago Quincenal.", "Pago Mensual."]
FORMA_PAGO_OPTIONS = ["Abono a cuenta bancaria.", "Nequi o Daviplata.", "Efectivo.", "Cheque."]
OBS_PRESTACIONES_OPTIONS = [
    "1. Conoce, pero es la primera vez que tiene estos beneficios.",
    "2. Requiere más información.",
    "3. Desconoce.",
    "0. Conoce los beneficios y la aplicación.",
    "No aplica.",
]
OBS_CONDUCTO_REGULAR_OPTIONS = [
    "1. Conoce el conducto por experiencias anteriores.",
    "2. Requiere más información.",
    "3. Desconoce la información.",
    "0. Conoce el conducto regular.",
]
OBS_DESCARGOS_OPTIONS = [
    "Si conoce que es una diligencia de descargos.",
    "NO conoce que es una diligencia de descargos.",
]
OBS_TRAMITES_OPTIONS = [
    "Conoce cómo es el proceso para realizar trámites administrativos (certificaciones, afiliaciones, descuentos, desprendibles de nómina).",
    "NO Conoce cómo es el proceso para realizar trámites administrativos (certificaciones, afiliaciones, descuentos, desprendibles de nómina).",
]
OBS_PERMISOS_OPTIONS = [
    "Conoce cómo es el proceso de solicitud de permisos.",
    "NO Conoce cómo es el proceso de solicitud de permisos.",
]
OBS_CAUSALES_OPTIONS = [
    "1. Tiene claro las causales de cancelación del contrato por experiencias anteriores.",
    "2. Requiere aclaración de algunas causales de cancelación del contrato.",
    "3. Desconoce las causales de cancelación del contrato.",
    "0. Tiene claro las causales de cancelación del contrato.",
]
OBS_RUTAS_OPTIONS = [
    "0. Tiene claro cuales son las rutas de atención.",
    "1. Requiere aclaración de cuales son las rutas de atención.",
    "2. Conoce las rutas de atención, pero no las usa",
    "3. Desconoce las rutas de atención.",
    "4. No aplica",
]

EVALUADOR_NOMBRES = [
    "Sandra Milena Pachon Rojas",
    "Sara Zambrano",
    "Alejandra Perez",
    "Lenny Lugo",
    "Angie Diaz",
    "Adriana Viveros",
    "Janeth Camargo",
    "Gabriela Rubiano Isaza",
    "Andres Montes",
    "Sara Sanchez",
    "Catalina Salazar",
]

EVALUADOR_CARGOS = [
    "Coordinadora de inclusion laboral",
    "Coordinacion de inclusion laboral",
    "Gestora de inclusion laboral",
    "Profesional de apoyo de inclusion laboral",
    "Gestor de inclusion laboral",
    "Lider Empleo Inclusivo",
    "Gestora de proyectos y desarrollo",
    "Profesional de inclusion laboral",
    "Directora Fundacion Reca",
]

SECTION_1 = {
    "title": "1. DATOS DE LA EMPRESA",
    "nit_lookup_field": "nit_empresa",
    "fields": [
        {"id": "fecha_visita", "label": "Fecha de la visita", "source": "input"},
        {
            "id": "modalidad",
            "label": "Modalidad",
            "source": "input",
            "options": ["Presencial", "Virtual", "Mixta", "No aplica"],
        },
        {
            "id": "nombre_empresa",
            "label": "Nombre de la empresa",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "ciudad_empresa",
            "label": "Ciudad/Municipio",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "direccion_empresa",
            "label": "Dirección de la empresa",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {"id": "nit_empresa", "label": "Número de NIT", "source": "input"},
        {
            "id": "correo_1",
            "label": "Correo electrónico",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "telefono_empresa",
            "label": "Teléfonos",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "contacto_empresa",
            "label": "Persona que atiende la visita",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "cargo",
            "label": "Cargo",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "caja_compensacion",
            "label": "Empresa afiliada a Caja de Compensación",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "sede_empresa",
            "label": "Sede Compensar",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "asesor",
            "label": "Asesor",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "profesional_asignado",
            "label": "Profesional asignado RECA",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
    ],
}

SECTION_1_SUPABASE_MAP = evaluacion_accesibilidad.SECTION_1_SUPABASE_MAP.copy()

EXCEL_MAPPING = {
    "section_1": {
        "fecha_visita": "D7",
        "modalidad": "L7",
        "nombre_empresa": "D8",
        "ciudad_empresa": "L8",
        "direccion_empresa": "D9",
        "nit_empresa": "L9",
        "correo_1": "D10",
        "telefono_empresa": "L10",
        "contacto_empresa": "D11",
        "cargo": "L11",
        "caja_compensacion": "D12",
        "sede_empresa": "L12",
        "asesor": "D13",
        "profesional_asignado": "L13",
    },
    "section_7": {
        "start_row": 74,
        "rows": 3,
        "nombre_col": "C",
        "cargo_col": "K",
    },
}

LIST_FIELD_OPTIONS_BY_ID = {
    "modalidad": ["Presencial", "Virtual", "Mixta", "No aplica"],
    "discapacidad": list(DISCAPACIDAD_OPTIONS),
    "lgtbiq": list(LGTBIQ_OPTIONS),
    "grupo_etnico": list(GRUPO_ETNICO_OPTIONS),
    "grupo_etnico_cual": list(GRUPO_ETNICO_CUAL_OPTIONS),
    "certificado_discapacidad": list(CERTIFICADO_DISCAPACIDAD_OPTIONS),
    "tipo_contrato": list(TIPO_CONTRATO_FIRMADO_OPTIONS),
    "contrato_lee_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "contrato_lee_observacion": list(OBS_LECTURA_CONTRATO_OPTIONS_INDIVIDUAL),
    "contrato_comprendido_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "contrato_comprendido_observacion": list(OBS_COMPRENDE_CONTRATO_OPTIONS),
    "contrato_tipo_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "contrato_tipo_observacion": list(OBS_TIPO_CONTRATO_OPTIONS),
    "contrato_tipo_contrato": list(CONTRATO_TIPO_CONTRATO_OPTIONS),
    "contrato_jornada": list(JORNADA_LABORAL_OPTIONS),
    "contrato_clausulas": list(CLAUSULAS_CONTRATO_OPTIONS),
    "condiciones_salariales_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "condiciones_salariales_observacion": list(OBS_CONDICIONES_SALARIALES_OPTIONS),
    "condiciones_salariales_frecuencia_pago": list(FRECUENCIA_PAGO_OPTIONS),
    "condiciones_salariales_forma_pago": list(FORMA_PAGO_OPTIONS),
    "prestaciones_cesantias_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "prestaciones_cesantias_observacion": list(OBS_PRESTACIONES_OPTIONS),
    "prestaciones_auxilio_transporte_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "prestaciones_auxilio_transporte_observacion": list(OBS_PRESTACIONES_OPTIONS),
    "prestaciones_prima_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "prestaciones_prima_observacion": list(OBS_PRESTACIONES_OPTIONS),
    "prestaciones_seguridad_social_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "prestaciones_seguridad_social_observacion": list(OBS_PRESTACIONES_OPTIONS),
    "prestaciones_vacaciones_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "prestaciones_vacaciones_observacion": list(OBS_PRESTACIONES_OPTIONS),
    "prestaciones_auxilios_beneficios_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "prestaciones_auxilios_beneficios_observacion": list(OBS_PRESTACIONES_OPTIONS),
    "conducto_regular_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "conducto_regular_observacion": list(OBS_CONDUCTO_REGULAR_OPTIONS),
    "descargos_observacion": list(OBS_DESCARGOS_OPTIONS),
    "tramites_observacion": list(OBS_TRAMITES_OPTIONS),
    "permisos_observacion": list(OBS_PERMISOS_OPTIONS),
    "causales_fin_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "causales_fin_observacion": list(OBS_CAUSALES_OPTIONS),
    "rutas_atencion_nivel_apoyo": list(NIVEL_APOYO_OPTIONS),
    "rutas_atencion_observacion": list(OBS_RUTAS_OPTIONS),
}

EXCEL_DROPDOWN_MANUAL_CANONICAL_OPTIONS = {
    "discapacidad": list(DISCAPACIDAD_OPTIONS),
}

EXCEL_DROPDOWN_EXPLICIT_ALIASES = {
    "grupo_etnico_cual": {
        "gitano (rom)": "Rom o Gitano",
        "gitano rom": "Rom o Gitano",
        "rom": "Rom o Gitano",
        "afrocolombiano": "Afocolombiano",
        "afro colombiano": "Afocolombiano",
    },
    "tipo_contrato": {
        "prestacion de servicios": "Contrato por prestación de servicios",
        "contrato de prestacion de servicios": "Contrato por prestación de servicios",
        "contrato prestacion de servicios": "Contrato por prestación de servicios",
        "termino fijo": "Contrato de trabajo a término fijo",
        "contrato a termino fijo": "Contrato de trabajo a término fijo",
        "termino indefinido": "Contrato de trabajo a término indefinido",
        "contrato a termino indefinido": "Contrato de trabajo a término indefinido",
        "obra o labor": "Contrato por obra o labor",
        "contrato obra o labor": "Contrato por obra o labor",
        "aprendizaje": "Contrato de aprendizaje",
        "contrato de aprendizaje": "Contrato de aprendizaje",
        "temporal": "Contrato temporal",
        "contrato temporal": "Contrato temporal",
    },
}

DATE_FIELD_IDS = {
    "fecha_nacimiento",
    "fecha_firma_contrato",
    "fecha_fin",
}

SECTION_2_ANCHOR = "2. DATOS DEL VINCULADO"
SECTION_6_ANCHOR = "6. AJUSTES RAZONABLES / RECOMENDACIONES AL PROCESO DE CONTRATACION"
SECTION_6_GROUP_ANCHOR = "5. AJUSTES RAZONABLES Y RECOMENDACIONES"
SECTION_7_ANCHOR = "7. ASISTENTES"
SECTION_7_GROUP_ANCHOR = "6. ASISTENTES"
SECTION_2_LAST_COLUMN = "Q"
SECTION_2_GROUP_BLOCK_HEIGHT = 52
SECTION_2_GROUP_FIRST_BLOCK_START_ROW = 19
SECTION_2_GROUP_SECOND_BLOCK_START_ROW = 71
SECTION_2_GROUP_SHARED_ACTIVITY_CELL = "A15"

SECTION_2_INDIVIDUAL_CELL_MAP = {
    "numero": ("A", 18),
    "nombre_oferente": ("C", 18),
    "cedula": ("H", 18),
    "certificado_porcentaje": ("K", 18),
    "discapacidad": ("L", 18),
    "telefono_oferente": ("O", 18),
    "genero": ("C", 19),
    "correo_oferente": ("G", 19),
    "fecha_nacimiento": ("M", 19),
    "edad": ("Q", 19),
    "lgtbiq": ("E", 20),
    "grupo_etnico": ("L", 20),
    "grupo_etnico_cual": ("O", 20),
    "cargo_oferente": ("C", 21),
    "contacto_emergencia": ("I", 21),
    "parentesco": ("M", 21),
    "telefono_emergencia": ("Q", 21),
    "certificado_discapacidad": ("F", 22),
    "lugar_firma_contrato": ("L", 22),
    "fecha_firma_contrato": ("Q", 22),
    "tipo_contrato": ("G", 24),
    "fecha_fin": ("N", 24),
    "desarrollo_actividad": ("A", 26),
    "contrato_lee_nivel_apoyo": ("G", 30),
    "contrato_lee_observacion": ("L", 30),
    "contrato_lee_nota": ("M", 31),
    "contrato_comprendido_nivel_apoyo": ("G", 32),
    "contrato_comprendido_observacion": ("L", 32),
    "contrato_comprendido_nota": ("M", 33),
    "contrato_tipo_nivel_apoyo": ("G", 34),
    "contrato_tipo_observacion": ("L", 34),
    "contrato_tipo_contrato": ("L", 35),
    "contrato_jornada": ("L", 36),
    "contrato_clausulas": ("L", 37),
    "contrato_tipo_nota": ("M", 38),
    "condiciones_salariales_nivel_apoyo": ("G", 39),
    "condiciones_salariales_observacion": ("L", 39),
    "condiciones_salariales_frecuencia_pago": ("L", 40),
    "condiciones_salariales_forma_pago": ("L", 41),
    "condiciones_salariales_nota": ("M", 42),
    "prestaciones_cesantias_nivel_apoyo": ("G", 45),
    "prestaciones_cesantias_observacion": ("L", 45),
    "prestaciones_cesantias_nota": ("M", 46),
    "prestaciones_auxilio_transporte_nivel_apoyo": ("G", 47),
    "prestaciones_auxilio_transporte_observacion": ("L", 47),
    "prestaciones_auxilio_transporte_nota": ("M", 48),
    "prestaciones_prima_nivel_apoyo": ("G", 49),
    "prestaciones_prima_observacion": ("L", 49),
    "prestaciones_prima_nota": ("M", 50),
    "prestaciones_seguridad_social_nivel_apoyo": ("G", 51),
    "prestaciones_seguridad_social_observacion": ("L", 51),
    "prestaciones_seguridad_social_nota": ("M", 52),
    "prestaciones_vacaciones_nivel_apoyo": ("G", 53),
    "prestaciones_vacaciones_observacion": ("L", 53),
    "prestaciones_vacaciones_nota": ("M", 54),
    "prestaciones_auxilios_beneficios_nivel_apoyo": ("G", 55),
    "prestaciones_auxilios_beneficios_observacion": ("L", 55),
    "prestaciones_auxilios_beneficios_nota": ("M", 56),
    "conducto_regular_nivel_apoyo": ("G", 59),
    "conducto_regular_observacion": ("L", 59),
    "descargos_observacion": ("L", 60),
    "tramites_observacion": ("L", 61),
    "permisos_observacion": ("L", 62),
    "conducto_regular_nota": ("M", 63),
    "causales_fin_nivel_apoyo": ("G", 64),
    "causales_fin_observacion": ("L", 64),
    "causales_fin_nota": ("M", 65),
    "rutas_atencion_nivel_apoyo": ("G", 66),
    "rutas_atencion_observacion": ("L", 66),
    "rutas_atencion_nota": ("M", 67),
}

SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP = dict(SECTION_2_INDIVIDUAL_CELL_MAP)
for _field_id, (_col, _row) in list(SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP.items()):
    if _field_id == "desarrollo_actividad":
        SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP.pop(_field_id, None)
    elif _row <= 24:
        SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP[_field_id] = (_col, _row + 5)
    elif _row >= 30:
        SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP[_field_id] = (_col, _row + 3)

SECTION_1_CELL_MAP_BY_TEMPLATE = {
    TEMPLATE_VARIANT_INDIVIDUAL: EXCEL_MAPPING["section_1"],
    TEMPLATE_VARIANT_GROUP_2_PLUS: dict(EXCEL_MAPPING["section_1"]),
}

SECTION_7_BASE_ROWS_BY_TEMPLATE = {
    TEMPLATE_VARIANT_INDIVIDUAL: 3,
    TEMPLATE_VARIANT_GROUP_2_PLUS: 4,
}


def register_form():
    return {"id": FORM_ID, "name": FORM_NAME, "module": __name__}


def _get_cache_dir():
    base = os.getenv("LOCALAPPDATA")
    if not base:
        userprofile = os.getenv("USERPROFILE")
        if userprofile:
            base = os.path.join(userprofile, "AppData", "Local")
    if not base:
        base = os.getcwd()
    cache_dir = os.path.join(base, "RECA", "cache")
    os.makedirs(cache_dir, exist_ok=True)
    return cache_dir


def _get_cache_path():
    return os.path.join(_get_cache_dir(), "contratacion_incluyente.json")


def cache_file_exists():
    return os.path.exists(_get_cache_path())


def save_cache_to_file():
    payload = {
        "form_id": FORM_ID,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "data": FORM_CACHE,
    }
    with open(_get_cache_path(), "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def _normalize_section_2_payload(payload):
    if not isinstance(payload, list):
        return payload
    normalized = []
    shared_desarrollo = ""
    for entry in payload:
        current = dict(entry or {})
        normalized.append(current)
        if not shared_desarrollo:
            shared_desarrollo = (current.get("desarrollo_actividad") or "").strip()
    for entry in normalized:
        entry["desarrollo_actividad"] = shared_desarrollo
    return normalized


def load_cache_from_file():
    path = _get_cache_path()
    if not os.path.exists(path):
        return False
    with open(path, "r", encoding="utf-8") as handle:
        payload = json.load(handle) or {}
    data = payload.get("data") or {}
    FORM_CACHE.clear()
    FORM_CACHE.update(data)
    section_2 = FORM_CACHE.get("section_2")
    if isinstance(section_2, list):
        FORM_CACHE["section_2"] = _normalize_section_2_payload(section_2)
    section_1 = data.get("section_1") or {}
    SECTION_1_CACHE.clear()
    SECTION_1_CACHE.update(section_1)
    return True


def clear_cache_file():
    path = _get_cache_path()
    if os.path.exists(path):
        os.remove(path)


def clear_form_cache():
    FORM_CACHE.clear()
    SECTION_1_CACHE.clear()


def set_section_cache(section_id, payload):
    if not section_id:
        raise ValueError("section_id requerido")
    if payload is None:
        payload = {}
    FORM_CACHE[section_id] = payload


def get_form_cache():
    return dict(FORM_CACHE)


def _get_section_2_entries(payload=None):
    if payload is None:
        payload = FORM_CACHE.get("section_2", [])
    if not isinstance(payload, list):
        return []
    return [dict(entry or {}) for entry in payload]


def _resolve_template_variant(section_2_payload=None):
    total_vinculados = len(_get_section_2_entries(section_2_payload))
    if total_vinculados >= 1:
        return TEMPLATE_VARIANT_GROUP_2_PLUS
    return TEMPLATE_VARIANT_INDIVIDUAL


def is_group_variant(section_2_payload=None):
    return _resolve_template_variant(section_2_payload) == TEMPLATE_VARIANT_GROUP_2_PLUS


def _section_2_group_block_start_row(entry_index):
    return SECTION_2_GROUP_FIRST_BLOCK_START_ROW + (
        SECTION_2_GROUP_BLOCK_HEIGHT * entry_index
    )


def _section_2_group_insert_row(entry_index):
    if entry_index <= 0:
        raise ValueError("entry_index debe ser mayor que 0 para bloques adicionales.")
    return SECTION_2_GROUP_SECOND_BLOCK_START_ROW + (
        SECTION_2_GROUP_BLOCK_HEIGHT * (entry_index - 1)
    )


def get_section_2_field_options(field_id, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if field_id == "contrato_lee_observacion":
        if template_variant == TEMPLATE_VARIANT_GROUP_2_PLUS:
            return list(OBS_LECTURA_CONTRATO_OPTIONS_GROUP)
        return list(OBS_LECTURA_CONTRATO_OPTIONS_INDIVIDUAL)
    return list(LIST_FIELD_OPTIONS_BY_ID.get(field_id, []))


def _find_first_row_by_texts(ws, *texts):
    last_error = None
    for text in texts:
        if not text:
            continue
        try:
            return _find_row_by_text(ws, text)
        except Exception as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    raise ValueError("No se proporcionaron textos para buscar.")


def _find_template_path(template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    def _filename_key(value):
        return "".join(ch for ch in _normalize_text(value) if ch.isalnum())

    templates_dir = resource_path("templates")
    if not templates_dir.is_dir():
        raise FileNotFoundError("No existe la carpeta templates.")
    filename = TEMPLATE_FILENAME_BY_VARIANT.get(template_variant)
    if filename:
        exact_path = templates_dir / filename
        if exact_path.exists():
            return os.fspath(exact_path)
        expected_key = _filename_key(filename)
        for name in os.listdir(templates_dir):
            if name.startswith("~$") or not name.lower().endswith(".xlsx"):
                continue
            if _filename_key(name) == expected_key:
                return os.fspath(templates_dir / name)
        if template_variant != TEMPLATE_VARIANT_INDIVIDUAL:
            raise FileNotFoundError(
                f"No se encontró el template '{filename}' para contratación incluyente."
            )
    raise FileNotFoundError("No se encontró el template de contratación incluyente.")


def _normalize_dropdown_text(value):
    normalized = _normalize_text(value or "")
    normalized = normalized.replace('"', "").replace("'", "")
    normalized = " ".join(normalized.split())
    return normalized.strip(" .")


def _iter_sqref_cells(sqref):
    from openpyxl.utils.cell import get_column_letter, range_boundaries

    for token in str(sqref or "").split():
        if ":" not in token:
            yield token
            continue
        min_col, min_row, max_col, max_row = range_boundaries(token)
        for col_idx in range(min_col, max_col + 1):
            for row_idx in range(min_row, max_row + 1):
                yield f"{get_column_letter(col_idx)}{row_idx}"


def _clean_inline_dropdown_formula(formula):
    text = str(formula or "").strip()
    if not text:
        return ""
    text = text.replace('"&"', "")
    if text.startswith("="):
        text = text[1:]
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1]
    return text


def _split_inline_dropdown_fragments(formula):
    cleaned = _clean_inline_dropdown_formula(formula)
    if not cleaned:
        return []
    return [fragment.strip() for fragment in cleaned.split(",") if fragment.strip()]


def _reconstruct_dropdown_options(fragments, expected_options):
    if not fragments or not expected_options:
        return []
    total_fragments = len(fragments)
    total_options = len(expected_options)

    @lru_cache(maxsize=None)
    def _solve(fragment_idx, option_idx):
        if option_idx == total_options:
            return (0.0, []) if fragment_idx == total_fragments else (float("inf"), [])
        remaining_options = total_options - option_idx
        remaining_fragments = total_fragments - fragment_idx
        if remaining_fragments < remaining_options:
            return float("inf"), []

        best_score = float("inf")
        best_sequence = []
        max_take = remaining_fragments - (remaining_options - 1)
        expected_norm = _normalize_dropdown_text(expected_options[option_idx])
        for take in range(1, max_take + 1):
            candidate = ", ".join(fragments[fragment_idx: fragment_idx + take]).strip()
            candidate_norm = _normalize_dropdown_text(candidate)
            distance = 1.0 - SequenceMatcher(None, candidate_norm, expected_norm).ratio()
            rest_score, rest_sequence = _solve(fragment_idx + take, option_idx + 1)
            total_score = distance + rest_score
            if total_score < best_score:
                best_score = total_score
                best_sequence = [candidate] + rest_sequence
        return best_score, best_sequence

    _score, sequence = _solve(0, 0)
    if len(sequence) != total_options:
        return []
    return sequence


def _get_list_field_cell(field_id, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    mapping = SECTION_1_CELL_MAP_BY_TEMPLATE.get(template_variant, EXCEL_MAPPING["section_1"])
    if field_id in mapping:
        return mapping[field_id]
    cell_map = (
        SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP
        if template_variant == TEMPLATE_VARIANT_GROUP_2_PLUS
        else SECTION_2_INDIVIDUAL_CELL_MAP
    )
    if field_id in cell_map:
        col, row = cell_map[field_id]
        return f"{col}{row}"
    return ""


@lru_cache(maxsize=None)
def _get_template_validation_formula_map(template_variant):
    from openpyxl import load_workbook

    path = _find_template_path(template_variant)
    workbook = load_workbook(path)
    target_sheet = SHEET_NAME_BY_VARIANT.get(template_variant) or workbook.sheetnames[0]
    worksheet = workbook[target_sheet]
    cell_map = {}
    for data_validation in getattr(worksheet.data_validations, "dataValidation", []):
        formula = getattr(data_validation, "formula1", None)
        if not formula:
            continue
        for cell in _iter_sqref_cells(getattr(data_validation, "sqref", "")):
            cell_map[cell] = formula
    workbook.close()
    return cell_map


@lru_cache(maxsize=None)
def _get_excel_canonical_options(field_id, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    manual = EXCEL_DROPDOWN_MANUAL_CANONICAL_OPTIONS.get(field_id)
    if manual:
        return tuple(manual)

    if field_id in EXCEL_MAPPING["section_1"]:
        expected_options = LIST_FIELD_OPTIONS_BY_ID.get(field_id, [])
    else:
        expected_options = get_section_2_field_options(field_id, template_variant)
    if not expected_options:
        return tuple()
    cell = _get_list_field_cell(field_id, template_variant)
    if not cell:
        return tuple(expected_options)
    formula = _get_template_validation_formula_map(template_variant).get(cell)
    fragments = _split_inline_dropdown_fragments(formula)
    reconstructed = _reconstruct_dropdown_options(fragments, tuple(expected_options))
    if len(reconstructed) == len(expected_options):
        return tuple(reconstructed)
    return tuple(expected_options)


def normalize_excel_dropdown_value(field_id, raw_value, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if raw_value in (None, ""):
        return raw_value
    current = str(raw_value).strip()
    if field_id in EXCEL_MAPPING["section_1"]:
        field_options = LIST_FIELD_OPTIONS_BY_ID.get(field_id)
    else:
        field_options = get_section_2_field_options(field_id, template_variant)
    if not field_options:
        return raw_value

    canonical_options = list(_get_excel_canonical_options(field_id, template_variant) or field_options)
    current_norm = _normalize_dropdown_text(current)

    explicit_alias = (
        EXCEL_DROPDOWN_EXPLICIT_ALIASES.get(field_id, {}).get(current_norm)
    )
    if explicit_alias:
        return explicit_alias

    for option in canonical_options:
        if _normalize_dropdown_text(option) == current_norm:
            return option

    for idx, option in enumerate(field_options):
        if _normalize_dropdown_text(option) == current_norm:
            if idx < len(canonical_options):
                return canonical_options[idx]
            return option

    _log_excel(
        f"WARN export_dropdown_unmatched field={field_id} template_variant={template_variant} "
        f"value={current!r}"
    )
    return raw_value


def _coerce_excel_date_value(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value
    raw = str(value).strip()
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return raw


def _group_export_title_for_offerentes(total_oferentes):
    total = max(0, int(total_oferentes or 0))
    if total <= 1:
        return "PROCESO DE CONTRATACIÓN INCLUYENTE INDIVIDUAL"
    if total <= 4:
        return "PROCESO CONTRATACION INCLUYENTE GRUPAL - 2 A 4 VINCULADOS"
    if total <= 7:
        return "PROCESO CONTRATACION INCLUYENTE GRUPAL - 5 A 7 VINCULADOS"
    if total <= 10:
        return "PROCESO CONTRATACION INCLUYENTE GRUPAL - 8 A 10 VINCULADOS"
    return "PROCESO CONTRATACION INCLUYENTE GRUPAL - MAS DE 10 VINCULADOS"




def _infer_discapacidad_categoria(value):
    if not value:
        return None
    normalized = _normalize_text(value)
    if "no aplica" in normalized:
        return None
    if "multiple" in normalized:
        return "Múltiple"
    if "visual" in normalized:
        return "Visual"
    if "auditiva" in normalized or "hipoacusia" in normalized:
        return "Auditiva"
    if "fisica" in normalized:
        return "Física"
    if "psicosocial" in normalized:
        return "Psicosocial"
    if "intelectual" in normalized or "autismo" in normalized or "autista" in normalized:
        return "Intelectual"
    return _DISCAPACIDAD_CATEGORIA_MAP.get(normalized)




def get_usuarios_reca_cedulas(env_path=".env"):
    params = {
        "select": "cedula_usuario",
        "cedula_usuario": "not.is.null",
        "order": "cedula_usuario.asc",
    }
    data = _supabase_get("usuarios_reca", params, env_path=env_path)
    return [row.get("cedula_usuario") for row in data if row.get("cedula_usuario")]


def get_usuario_reca_by_cedula(cedula, env_path=".env"):
    normalized = _normalize_cedula(cedula)
    if not normalized:
        return None
    select_cols = ",".join(
        [
            "cedula_usuario",
            "nombre_usuario",
            "genero_usuario",
            "discapacidad_usuario",
            "discapacidad_detalle",
            "certificado_porcentaje",
            "telefono_oferente",
            "fecha_nacimiento",
            "cargo_oferente",
            "contacto_emergencia",
            "parentesco",
            "telefono_emergencia",
            "correo_oferente",
            "lgtbiq",
            "grupo_etnico",
            "grupo_etnico_cual",
            "certificado_discapacidad",
            "lugar_firma_contrato",
            "fecha_firma_contrato",
            "tipo_contrato",
            "fecha_fin",
            "empresa_nit",
            "empresa_nombre",
        ]
    )
    params = {
        "select": select_cols,
        "cedula_usuario": f"eq.{normalized}",
        "limit": 1,
    }
    data = _supabase_get("usuarios_reca", params, env_path=env_path)
    return data[0] if data else None


def _get_log_dir():
    output_path = FORM_CACHE.get("_output_path")
    if output_path:
        base_dir = os.path.dirname(output_path)
    else:
        base_dir = os.getcwd()
    log_dir = os.path.join(base_dir, "logs")
    os.makedirs(log_dir, exist_ok=True)
    return log_dir


def _log_excel(message):
    try:
        log_excel_event(message)
    except Exception:
        return


def _ensure_output_path(template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    template_path = _find_template_path(template_variant=template_variant)
    empresa_nombre = SECTION_1_CACHE.get("nombre_empresa") or "Empresa"
    process_name = "Proceso de Contratacion Incluyente"
    output_path = _build_process_output_path(empresa_nombre, process_name)
    shutil.copy2(template_path, output_path)
    FORM_CACHE["_output_path"] = output_path
    return output_path


def _get_sheet_by_name(workbook, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    target_name = SHEET_NAME_BY_VARIANT.get(template_variant)
    target = _normalize_text(target_name).replace(" ", "")
    for ws in workbook.Worksheets:
        name_norm = _normalize_text(ws.Name).replace(" ", "")
        if name_norm == target:
            return ws
    try:
        return workbook.Worksheets(target_name)
    except Exception as exc:
        raise KeyError(f"No existe la hoja {target_name}.") from exc


def get_empresa_by_nit(nit, env_path=".env"):
    return evaluacion_accesibilidad.get_empresa_by_nit(nit, env_path=env_path)


def get_empresa_by_nombre(nombre, env_path=".env"):
    return evaluacion_accesibilidad.get_empresa_by_nombre(nombre, env_path=env_path)


def get_empresas_by_nombre_prefix(prefix, env_path=".env", limit=50):
    return evaluacion_accesibilidad.get_empresas_by_nombre_prefix(prefix, env_path=env_path, limit=limit)


def confirm_section_1(company_data, user_inputs):
    if not company_data:
        raise ValueError("No hay datos de empresa para confirmar.")
    payload = {}
    for field in SECTION_1["fields"]:
        field_id = field["id"]
        if field["source"] == "input":
            payload[field_id] = user_inputs.get(field_id)
        else:
            payload[field_id] = company_data.get(field_id)
    SECTION_1_CACHE.update(payload)
    set_section_cache("section_1", payload)
    FORM_CACHE["_last_section"] = "section_1"
    save_cache_to_file()
    return payload


def confirm_section_2(payload):
    if payload is None:
        raise ValueError("section_2 requerida")
    payload = _normalize_section_2_payload(payload)
    set_section_cache("section_2", payload)
    FORM_CACHE["_last_section"] = "section_2"
    save_cache_to_file()
    return payload


def confirm_section_6(payload):
    if payload is None:
        raise ValueError("section_6 requerida")
    set_section_cache("section_6", payload)
    FORM_CACHE["_last_section"] = "section_6"
    save_cache_to_file()
    return payload


def confirm_section_7(payload):
    if payload is None:
        raise ValueError("section_7 requerida")
    set_section_cache("section_7", payload)
    FORM_CACHE["_last_section"] = "section_7"
    save_cache_to_file()
    return payload


def sync_usuarios_reca(env_path=".env"):
    data = FORM_CACHE.get("section_2")
    if not data and cache_file_exists():
        load_cache_from_file()
        data = FORM_CACHE.get("section_2")
    if not data:
        return 0

    empresa_nit = (SECTION_1_CACHE.get("nit_empresa") or "").strip()
    empresa_nombre = (SECTION_1_CACHE.get("nombre_empresa") or "").strip()
    rows = []
    for entry in data:
        cedula = _normalize_cedula(entry.get("cedula"))
        if not cedula:
            continue
        discapacidad_detalle = (entry.get("discapacidad") or "").strip()
        discapacidad_usuario = _infer_discapacidad_categoria(discapacidad_detalle)
        row = {
            "cedula_usuario": cedula,
            "nombre_usuario": (entry.get("nombre_oferente") or "").strip(),
            "genero_usuario": (entry.get("genero") or "").strip(),
            "discapacidad_usuario": discapacidad_usuario,
            "discapacidad_detalle": discapacidad_detalle or None,
            "certificado_porcentaje": _normalize_decimal_value(
                entry.get("certificado_porcentaje"),
                decimal_separator=".",
            ),
            "telefono_oferente": (entry.get("telefono_oferente") or "").strip(),
            "fecha_nacimiento": _parse_date_value(entry.get("fecha_nacimiento")),
            "cargo_oferente": (entry.get("cargo_oferente") or "").strip(),
            "contacto_emergencia": (entry.get("contacto_emergencia") or "").strip(),
            "parentesco": (entry.get("parentesco") or "").strip(),
            "telefono_emergencia": (entry.get("telefono_emergencia") or "").strip(),
            "correo_oferente": (entry.get("correo_oferente") or "").strip(),
            "lgtbiq": (entry.get("lgtbiq") or "").strip(),
            "grupo_etnico": (entry.get("grupo_etnico") or "").strip(),
            "grupo_etnico_cual": (entry.get("grupo_etnico_cual") or "").strip(),
            "certificado_discapacidad": (entry.get("certificado_discapacidad") or "").strip(),
            "lugar_firma_contrato": (entry.get("lugar_firma_contrato") or "").strip(),
            "fecha_firma_contrato": _parse_date_value(entry.get("fecha_firma_contrato")),
            "tipo_contrato": (entry.get("tipo_contrato") or "").strip(),
            "fecha_fin": _parse_date_value(entry.get("fecha_fin")) or (entry.get("fecha_fin") or "").strip(),
            "resultado_certificado": (entry.get("resultado_certificado") or "").strip(),
            "pendiente_otros_oferentes": (entry.get("pendiente_otros_oferentes") or "").strip(),
            "cuenta_pension": (entry.get("cuenta_pension") or "").strip(),
            "tipo_pension": (entry.get("tipo_pension") or "").strip(),
            # Siempre dejamos la última empresa de contratación asociada a la cédula.
            "empresa_nit": empresa_nit,
            "empresa_nombre": empresa_nombre,
        }
        normalized_row = {
            key: (None if value == "" else value)
            for key, value in row.items()
        }
        rows.append(normalized_row)
    deduped_rows = {}
    duplicate_cedulas = []
    for row in rows:
        cedula = row.get("cedula_usuario")
        if not cedula:
            continue
        if cedula in deduped_rows:
            duplicate_cedulas.append(cedula)
            deduped_rows.pop(cedula, None)
        deduped_rows[cedula] = row
    rows = list(deduped_rows.values())

    if duplicate_cedulas:
        preview_duplicates = ", ".join(duplicate_cedulas[:10])
        extra_duplicates = "" if len(duplicate_cedulas) <= 10 else f" (+{len(duplicate_cedulas) - 10} mas)"
        _log_excel(
            f"WARN supabase_usuarios_reca_duplicate_cedulas count={len(duplicate_cedulas)} "
            f"cedulas={preview_duplicates}{extra_duplicates}"
        )

    if rows:
        sync_result = _supabase_upsert_with_queue(
            "usuarios_reca",
            rows,
            env_path=env_path,
            on_conflict="cedula_usuario",
        )
        cedulas = [row.get("cedula_usuario") for row in rows if row.get("cedula_usuario")]
        preview = ", ".join(cedulas[:10])
        extra = "" if len(cedulas) <= 10 else f" (+{len(cedulas) - 10} mas)"
        status = sync_result.get("status") or "synced"
        _log_excel(
            f"SUPABASE usuarios_reca upsert status={status} count={len(rows)} cedulas={preview}{extra}"
        )
    return len(rows)


def _find_row_by_text(ws, text):
    cell = ws.Columns("A").Find(What=text, LookAt=1)
    if cell is not None:
        return cell.Row
    cell = ws.Columns("A").Find(What=text, LookAt=2)
    if cell is not None:
        return cell.Row
    target = _normalize_text(text)
    used = ws.UsedRange
    start_row = used.Row
    end_row = used.Row + used.Rows.Count - 1
    for row in range(start_row, end_row + 1):
        value = ws.Cells(row, 1).Value
        if not value:
            continue
        value_norm = _normalize_text(str(value))
        if value_norm == target:
            return row
    for row in range(start_row, end_row + 1):
        value = ws.Cells(row, 1).Value
        if not value:
            continue
        value_norm = _normalize_text(str(value))
        if target in value_norm:
            if target.startswith("2.") or target.startswith("5.") or target.startswith("6.") or target.startswith("7."):
                if value_norm.startswith(target):
                    return row
            else:
                return row
    raise ValueError(f"No se encontró el texto '{text}' en la columna A.")


def _insert_person_block(ws, start_row, block_height, insert_at):
    start_end = start_row + block_height - 1
    dest_end = insert_at + block_height - 1
    source = ws.Range(f"A{start_row}:{SECTION_2_LAST_COLUMN}{start_end}")
    dest = ws.Range(f"A{insert_at}:{SECTION_2_LAST_COLUMN}{dest_end}")
    source.Copy()
    dest.Insert(Shift=-4121)
    for row_offset in range(block_height):
        ws.Rows(insert_at + row_offset).RowHeight = ws.Rows(start_row + row_offset).RowHeight
    ws.Application.CutCopyMode = False


def _write_section_2_entry(
    ws,
    entry,
    cell_map,
    *,
    row_offset=0,
    template_variant=TEMPLATE_VARIANT_INDIVIDUAL,
):
    for field_id, (col, row) in cell_map.items():
        value = entry.get(field_id, "")
        if field_id == "grupo_etnico_cual":
            grupo_etnico = _normalize_text(entry.get("grupo_etnico") or "")
            if grupo_etnico not in {"si", "sí"}:
                value = "No aplica"
        if value == "":
            continue
        target_row = row + row_offset
        target_ref = f"{col}{target_row}"
        if field_id == "certificado_porcentaje":
            value = _coerce_excel_decimal_value(value)
        elif field_id in DATE_FIELD_IDS:
            value = _coerce_excel_date_value(value)
            try:
                if isinstance(value, datetime):
                    ws.Range(target_ref).NumberFormat = "dd/mm/yyyy"
                else:
                    ws.Range(target_ref).NumberFormat = "@"
            except Exception:
                pass
        else:
            value = normalize_excel_dropdown_value(
                field_id,
                value,
                template_variant=template_variant,
            )
        _log_excel(
            f"WRITE section=section_2 cell={target_ref} key={field_id} value={value!r}"
        )
        ws_write(ws, target_ref, value)


def _write_section_2_individual(ws, oferentes):
    if not oferentes:
        return
    _log_excel(
        f"SECTION section=section_2 variant=individual total={len(oferentes)}"
    )
    ws_write(ws, "F1", _group_export_title_for_offerentes(1))
    _write_section_2_entry(
        ws,
        oferentes[0],
        SECTION_2_INDIVIDUAL_CELL_MAP,
        template_variant=TEMPLATE_VARIANT_INDIVIDUAL,
    )


def _write_section_2_group(ws, oferentes):
    if not oferentes:
        return
    ws_write(ws, "F1", _group_export_title_for_offerentes(len(oferentes)))
    shared_desarrollo = ""
    for entry in oferentes:
        shared_desarrollo = (entry.get("desarrollo_actividad") or "").strip()
        if shared_desarrollo:
            break
    if shared_desarrollo:
        _log_excel(
            f"WRITE section=section_2 cell={SECTION_2_GROUP_SHARED_ACTIVITY_CELL} "
            f"key=desarrollo_actividad value={shared_desarrollo!r}"
        )
        ws_write(ws, SECTION_2_GROUP_SHARED_ACTIVITY_CELL, shared_desarrollo)

    if len(oferentes) > 1:
        for idx in range(1, len(oferentes)):
            insert_at = _section_2_group_insert_row(idx)
            _insert_person_block(
                ws,
                SECTION_2_GROUP_FIRST_BLOCK_START_ROW,
                SECTION_2_GROUP_BLOCK_HEIGHT,
                insert_at,
            )
            _log_excel(
                f"INSERT section=section_2 variant=group rows={SECTION_2_GROUP_BLOCK_HEIGHT} at={insert_at}"
            )

    for idx, entry in enumerate(oferentes):
        row_offset = _section_2_group_block_start_row(idx) - SECTION_2_GROUP_FIRST_BLOCK_START_ROW
        _write_section_2_entry(
            ws,
            entry,
            SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP,
            row_offset=row_offset,
            template_variant=TEMPLATE_VARIANT_GROUP_2_PLUS,
        )


def _write_section_2(ws, oferentes, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if template_variant == TEMPLATE_VARIANT_GROUP_2_PLUS:
        return _write_section_2_group(ws, oferentes)
    return _write_section_2_individual(ws, oferentes)


def _write_section_6(ws, payload, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if not payload:
        return
    if template_variant == TEMPLATE_VARIANT_GROUP_2_PLUS:
        anchor_row = _find_first_row_by_texts(ws, SECTION_6_GROUP_ANCHOR, SECTION_6_ANCHOR)
    else:
        anchor_row = _find_first_row_by_texts(ws, SECTION_6_ANCHOR, SECTION_6_GROUP_ANCHOR)
    ajustes_row = anchor_row + 1
    ajustes_value = payload.get("ajustes_recomendaciones", "")
    _log_excel(
        f"WRITE section=section_6 cell=A{ajustes_row} key=ajustes_recomendaciones value={ajustes_value!r}"
    )
    ws_write(ws, f"A{ajustes_row}", ajustes_value)


def _write_section_7(ws, payload, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if not payload:
        return
    mapping = EXCEL_MAPPING.get("section_7", {})
    if template_variant == TEMPLATE_VARIANT_GROUP_2_PLUS:
        title_row = _find_first_row_by_texts(ws, SECTION_7_GROUP_ANCHOR, SECTION_7_ANCHOR)
    else:
        title_row = _find_first_row_by_texts(ws, SECTION_7_ANCHOR, SECTION_7_GROUP_ANCHOR)
    start_row = title_row + 1
    base_rows = SECTION_7_BASE_ROWS_BY_TEMPLATE.get(template_variant, mapping.get("rows", 3))
    nombre_col = mapping.get("nombre_col", "C")
    cargo_col = mapping.get("cargo_col", "K")
    total = len(payload)
    if total > base_rows:
        insert_at = start_row + base_rows
        template_row = start_row + base_rows - 1
        for _ in range(total - base_rows):
            ws.Rows(insert_at).Insert()
            ws.Rows(template_row).Copy(ws.Rows(insert_at))
            insert_at += 1
            _log_excel(
                f"INSERT section=section_7 rows=1 at={insert_at - 1}"
            )
    for idx, entry in enumerate(payload):
        row = start_row + idx
        nombre = entry.get("nombre", "")
        cargo = entry.get("cargo", "")
        _log_excel(
            f"WRITE section=section_7 cell={nombre_col}{row} key=nombre value={nombre!r}"
        )
        _log_excel(
            f"WRITE section=section_7 cell={cargo_col}{row} key=cargo value={cargo!r}"
        )
        ws_write(ws, f"{nombre_col}{row}", nombre)
        ws_write(ws, f"{cargo_col}{row}", cargo)


def _write_section_1(ws, payload, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if not payload:
        payload = SECTION_1_CACHE
    if not payload:
        try:
            if load_cache_from_file():
                payload = FORM_CACHE.get("section_1", {}) or SECTION_1_CACHE
        except Exception:
            payload = payload or {}
    mapping = SECTION_1_CELL_MAP_BY_TEMPLATE.get(template_variant, EXCEL_MAPPING.get("section_1", {}))
    for key, cell in mapping.items():
        if key in payload:
            value = normalize_excel_dropdown_value(
                key,
                payload.get(key),
                template_variant=template_variant,
            )
            ws_write(ws, cell, value)
            _log_excel(
                f"WRITE section=section_1 cell={cell} key={key} value={value!r}"
            )


def export_to_excel(clear_cache=True):
    clear_written_rows()
    if not FORM_CACHE.get("section_1") and cache_file_exists():
        load_cache_from_file()
    section_2_payload = FORM_CACHE.get("section_2", [])
    template_variant = _resolve_template_variant(section_2_payload)
    output_path = _ensure_output_path(template_variant=template_variant)
    _log_excel(f"START export_all output={output_path}")
    try:
        import win32com.client as win32
    except ImportError as exc:
        _log_excel("ERROR export_all error=pywin32_not_installed")
        raise RuntimeError("pywin32 no esta instalado. Instala con pip install pywin32.") from exc
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(output_path)
        ws = _get_sheet_by_name(wb, template_variant=template_variant)
        _write_section_1(ws, FORM_CACHE.get("section_1", {}), template_variant=template_variant)
        _write_section_2(ws, section_2_payload, template_variant=template_variant)
        _write_section_6(ws, FORM_CACHE.get("section_6", {}), template_variant=template_variant)
        _write_section_7(ws, FORM_CACHE.get("section_7", []), template_variant=template_variant)
        sanitize_logo_error_cells(wb)
        autofit_rows(ws, log_fn=_log_excel)
        wb.Save()
        _log_excel("SUCCESS export_all")
    except Exception as exc:
        _log_excel(f"ERROR export_all error={exc!r}")
        raise
    finally:
        if wb is not None:
            wb.Close(SaveChanges=True)
        excel.Quit()
    if clear_cache:
        clear_cache_file()
        clear_form_cache()
    return output_path

