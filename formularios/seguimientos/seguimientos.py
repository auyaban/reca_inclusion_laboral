"""
formularios/seguimientos/seguimientos.py
Formulario: "9. SEGUIMIENTO AL PROCESO DE INCLUSIÓN LABORAL"

Responsabilidades:
  - Flujo diferente al resto: NO usa el master spreadsheet directamente.
    Cada caso (persona vinculada) tiene su propio Google Sheet copiado desde
    una plantilla (SEGUIMIENTOS_TEMPLATE_ID) en una carpeta compartida de Drive
  - build_case_folder_name(): construye el nombre de carpeta por usuario/cédula
  - Gestión de casos abiertos/cerrados: un caso = una carpeta en Drive con su Sheet
  - Sincronización: lee casos existentes desde Drive (carpeta SEGUIMIENTOS_SHARED_ROOT)
  - SeguimientoEditorWindow en app.py edita un caso individual

Entry points para app.py:
  register_form()              → metadata para HubWindow
  get_empresa_by_nit/nombre/prefix() → búsqueda de empresa
  build_case_folder_name()     → nombre de carpeta del caso en Drive

Variables clave de entorno:
  SEGUIMIENTOS_SHARED_ROOT     → ID de carpeta raíz en Drive
  GOOGLE_SHEETS_SEGUIMIENTOS_TEMPLATE_ID → ID de la plantilla de seguimiento

Depende de: google_sheets_client, drive_upload, google_api_requests,
            formularios/common, openpyxl
"""
import os
import re
import shutil
import io
import tempfile
import uuid
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

from formularios.common import (
    _get_desktop_dir,
    _load_env_file,
    _merge_company_row_with_cache,
    _normalize_cedula,
    sanitize_logo_error_cells,
    _sanitize_filename,
    _supabase_get,
)
from google_api_requests import (
    execute_google_create_with_confirmation,
    execute_google_request_with_retry,
)
from google_sheets_client import (
    batch_write_sheet_updates,
    clear_protected_ranges,
    extract_spreadsheet_id,
    get_master_template_id,
    get_google_sheets_service,
    get_spreadsheet,
    read_sheet_values,
)
import drive_upload
from version_info import resource_path


FORM_ID = "seguimientos"
FORM_NAME = "Seguimientos"
DEFAULT_SHARED_ROOT = ""
SEGUIMIENTOS_FOLDER_NAME = "SEGUIMIENTOS"
DEFAULT_SEGUIMIENTOS_TEMPLATE_ID = ""
GOOGLE_SHEETS_MIME = "application/vnd.google-apps.spreadsheet"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SHEET_BASE = "9. SEGUIMIENTO AL PROCESO DE INCLUSIÓN LABORAL"
LEGACY_SHEET_BASE = "9.  SEGUIMIENTO AL PROCESO DE INCLUSION LABORAL"
LEGACY_SHEET_BASE_SHORT = "9.  SEGUIMIENTO AL PROCESO DE I"
BASE_SHEET_CANDIDATES = (
    SHEET_BASE,
    LEGACY_SHEET_BASE,
    LEGACY_SHEET_BASE_SHORT,
)
SHEET_PREFIX = "SEGUIMIENTO PROCESO IL "
SHEET_FINAL = "PONDERADO FINAL"
SHEET_META = "_META_IL"


def _load_runtime_env():
    try:
        return _load_env_file(".env") or {}
    except Exception:
        return {}


def _get_shared_root():
    runtime_env = _load_runtime_env()
    raw = str(
        os.getenv("SEGUIMIENTOS_SHARED_ROOT")
        or runtime_env.get("SEGUIMIENTOS_SHARED_ROOT")
        or drive_upload._load_config().get("seguimientos_shared_root")
        or DEFAULT_SHARED_ROOT
    ).strip()
    return raw or DEFAULT_SHARED_ROOT


def _load_workbook_safe(path, *, data_only=False):
    try:
        return load_workbook(path, data_only=data_only)
    except FileNotFoundError as exc:
        raise RuntimeError(f"No existe el archivo de seguimientos: {path}") from exc
    except Exception as exc:
        raise RuntimeError(
            f"No se pudo abrir el archivo de seguimientos. Puede estar corrupto o bloqueado por Excel: {path}"
        ) from exc

MODALIDAD_OPTIONS = ["Presencial", "Virtual", "Mixta", "No aplica"]
SI_NO_NA_OPTIONS = ["Si", "No", "No aplica"]
EVAL_OPTIONS = ["Excelente", "Bien", "Necesita mejorar", "Mal", "No aplica"]
TIPO_APOYO_OPTIONS = [
    "Requiere apoyo bajo.",
    "Requiere apoyo medio.",
    "Requiere apoyo Alto.",
    "No requiere apoyo.",
]

SECTION_1_SUPABASE_MAP = {
    "nombre_empresa": "nombre_empresa",
    "ciudad_empresa": "ciudad_empresa",
    "direccion_empresa": "direccion_empresa",
    "nit_empresa": "nit_empresa",
    "correo_1": "correo_1",
    "telefono_empresa": "telefono_empresa",
    "contacto_empresa": "contacto_empresa",
    "cargo": "cargo",
    "asesor": "asesor",
    "sede_empresa": "zona_empresa",
    "caja_compensacion": "caja_compensacion",
    "profesional_asignado": "profesional_asignado",
}

PONDERADO_COMPANY_MAP = {
    "fecha_visita": "D6",
    "modalidad": "Q6",
    "nombre_empresa": "D7",
    "ciudad_empresa": "Q7",
    "direccion_empresa": "D8",
    "nit_empresa": "Q8",
    "correo_1": "D9",
    "telefono_empresa": "Q9",
    "contacto_empresa": "D10",
    "cargo": "Q10",
    "caja_compensacion": "D11",
    "sede_empresa": "Q11",
    "asesor": "D12",
    "profesional_asignado": "Q12",
}

PONDERADO_USER_MAP = {
    "nombre_vinculado": "K15",
    "cedula": "Q15",
    "telefono_vinculado": "S15",
    "correo_vinculado": "U15",
    "cargo_vinculado": "K17",
    "certificado_discapacidad": "Q17",
    "certificado_porcentaje": "U17",
    "fecha_firma_contrato": "N18",
    "discapacidad": "U18",
}

def register_form():
    return {
        "id": FORM_ID,
        "name": FORM_NAME,
        "module": __name__,
        "supports_drafts": False,
        "hub_description": "Abre y actualiza casos con hoja base y seguimientos periódicos.",
        "singleton_window": True,
    }


def _map_company_row(row):
    if not isinstance(row, dict):
        return row
    mapped = dict(row)
    for field_id, source_key in SECTION_1_SUPABASE_MAP.items():
        if source_key in row:
            mapped[field_id] = row.get(source_key)
    return mapped


def get_linked_company_for_user(user_row, env_path=".env"):
    if not isinstance(user_row, dict):
        return {}
    nit = _get_str(user_row.get("empresa_nit"))
    nombre = _get_str(user_row.get("empresa_nombre"))
    company = None
    if nit:
        try:
            company = get_empresa_by_nit(nit, env_path=env_path)
        except Exception:
            company = None
    if not company and nombre:
        try:
            company = get_empresa_by_nombre(nombre, env_path=env_path)
        except Exception:
            company = None
    payload = _map_company_row(company or {})
    if nit and not payload.get("nit_empresa"):
        payload["nit_empresa"] = nit
    if nombre and not payload.get("nombre_empresa"):
        payload["nombre_empresa"] = nombre
    return payload


def _get_base_sheet_name_from_workbook(wb):
    for candidate in BASE_SHEET_CANDIDATES:
        if candidate in wb.sheetnames:
            return candidate
    raise ValueError("No existe la hoja base de seguimientos en el archivo.")


def _is_native_case_ref(case_ref):
    return (
        isinstance(case_ref, dict)
        and str(case_ref.get("source") or "").strip() == "drive"
        and str(case_ref.get("mime_type") or "").strip() == GOOGLE_SHEETS_MIME
        and str(case_ref.get("file_id") or "").strip()
    )


def _get_spreadsheet_id_from_case_ref(case_ref):
    if _is_native_case_ref(case_ref):
        return str(case_ref.get("file_id") or "").strip()
    return extract_spreadsheet_id(case_ref)


def _batch_read_sheet_values(spreadsheet_id_or_url, ranges):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    cleaned = [str(item or "").strip() for item in (ranges or []) if str(item or "").strip()]
    if not cleaned:
        return {}
    service = get_google_sheets_service()
    response = execute_google_request_with_retry(
        service.spreadsheets()
        .values()
        .batchGet(spreadsheetId=spreadsheet_id, ranges=cleaned),
        operation_name="seguimientos.batch_read_values",
    )
    return {item.get("range"): item.get("values", []) for item in response.get("valueRanges", [])}


def _first_batch_value(values_map, range_name):
    rows = values_map.get(range_name) or []
    if not rows or not rows[0]:
        return ""
    return str(rows[0][0]).strip()


def _column_batch_values(values_map, range_name, expected_count):
    rows = values_map.get(range_name) or []
    values = []
    for row in rows:
        if row:
            values.append(str(row[0]).strip())
        else:
            values.append("")
    if len(values) < expected_count:
        values.extend([""] * (expected_count - len(values)))
    return values[:expected_count]


def get_base_sheet_name(workbook_path):
    wb = _load_workbook_safe(workbook_path, data_only=False)
    return _get_base_sheet_name_from_workbook(wb)


def _get_drive_service():
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError(
            "Faltan dependencias para Google Drive. Instala google-api-python-client y google-auth."
        ) from exc

    creds_path = drive_upload._get_credentials_path()
    credentials = Credentials.from_service_account_file(creds_path, scopes=[drive_upload.SCOPE])
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def _get_drive_root_folder_id(service):
    configured_root = drive_upload._get_excel_folder_id()
    return drive_upload._resolve_target_root_id(service, configured_root)


def _find_named_folder(service, parent_id, folder_name):
    safe_name = str(folder_name or "").replace("'", "\\'")
    query = (
        "mimeType='application/vnd.google-apps.folder' "
        f"and name='{safe_name}' and '{parent_id}' in parents and trashed=false"
    )
    result = execute_google_request_with_retry(
        service.files().list(
            q=query,
            fields="files(id,name,webViewLink)",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            pageSize=5,
        ),
        operation_name="seguimientos.find_named_folder",
    )
    files = result.get("files", [])
    return files[0] if files else None


def _ensure_seguimientos_folder(service):
    root_id = _get_drive_root_folder_id(service)
    folder = _find_named_folder(service, root_id, SEGUIMIENTOS_FOLDER_NAME)
    if folder:
        return folder["id"]
    return drive_upload._get_or_create_folder(service, root_id, SEGUIMIENTOS_FOLDER_NAME)


def _list_drive_files(service, folder_id):
    result = execute_google_request_with_retry(
        service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="files(id,name,mimeType,webViewLink,modifiedTime,parents,appProperties)",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            pageSize=100,
            orderBy="modifiedTime desc",
        ),
        operation_name="seguimientos.list_drive_files",
    )
    return list(result.get("files", []))


def _find_drive_file_by_request_id(service, folder_id, filename, request_id):
    safe_name = str(filename or "").replace("'", "\\'")
    query = (
        f"mimeType='{GOOGLE_SHEETS_MIME}' "
        f"and name='{safe_name}' and '{folder_id}' in parents and trashed=false"
    )
    result = execute_google_request_with_retry(
        service.files().list(
            q=query,
            fields="files(id,name,mimeType,webViewLink,parents,appProperties)",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            pageSize=5,
        ),
        operation_name="seguimientos.find_drive_file_by_request_id",
    )
    for item in result.get("files", []):
        app_properties = item.get("appProperties") or {}
        if str(app_properties.get("request_id") or "").strip() == str(request_id or "").strip():
            return item
    return None


def _find_case_folder_drive(service, cedula, nombre_usuario=""):
    seguimientos_folder_id = _ensure_seguimientos_folder(service)
    preferred_name = build_case_folder_name(nombre_usuario, cedula) if nombre_usuario else ""
    candidates = []
    if preferred_name:
        direct = _find_named_folder(service, seguimientos_folder_id, preferred_name)
        if direct:
            return direct
    suffix = f"- {cedula}"
    for item in _list_drive_files(service, seguimientos_folder_id):
        if item.get("mimeType") != "application/vnd.google-apps.folder":
            continue
        name = str(item.get("name") or "").strip()
        if name.endswith(suffix):
            candidates.append(item)
    return candidates[0] if candidates else None


def _pick_case_file(files):
    preferred_order = {GOOGLE_SHEETS_MIME: 0, XLSX_MIME: 1}
    ranked = []
    for item in files:
        mime = str(item.get("mimeType") or "").strip()
        if mime not in preferred_order:
            continue
        ranked.append((preferred_order[mime], item))
    if not ranked:
        return None
    ranked.sort(key=lambda pair: pair[0])
    return ranked[0][1]


def get_seguimientos_template_id():
    try:
        return get_master_template_id()
    except Exception:
        runtime_env = _load_runtime_env()
        raw = str(
            os.getenv("GOOGLE_SHEETS_SEGUIMIENTOS_TEMPLATE_ID")
            or runtime_env.get("GOOGLE_SHEETS_SEGUIMIENTOS_TEMPLATE_ID")
            or drive_upload._load_config().get("google_sheets_seguimientos_template_id")
            or DEFAULT_SEGUIMIENTOS_TEMPLATE_ID
        ).strip()
        return extract_spreadsheet_id(raw)


def _get_drive_cache_dir():
    cache_root = Path(tempfile.gettempdir()) / "reca_seguimientos_drive"
    cache_root.mkdir(parents=True, exist_ok=True)
    return cache_root


def _build_local_case_copy_path(record):
    folder = _get_drive_cache_dir() / _sanitize_filename(record.get("folder_name") or record.get("cedula") or "caso")
    folder.mkdir(parents=True, exist_ok=True)
    base_name = _sanitize_filename(record.get("file_name") or "seguimiento")
    stem, _ = os.path.splitext(base_name)
    return str(folder / f"{stem}.xlsx")


def _download_drive_case_to_path(service, record, destination_path):
    file_id = str(record.get("file_id") or "").strip()
    mime_type = str(record.get("mime_type") or "").strip()
    if not file_id:
        raise RuntimeError("Caso sin file_id de Drive.")
    destination = Path(destination_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if mime_type == GOOGLE_SHEETS_MIME:
        request = service.files().export_media(
            fileId=file_id,
            mimeType=XLSX_MIME,
        )
    else:
        request = service.files().get_media(
            fileId=file_id,
            supportsAllDrives=True,
        )
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    destination.write_bytes(fh.getvalue())
    return str(destination)


def _ensure_dir(path):
    os.makedirs(path, exist_ok=True)
    return path


def _get_local_root():
    desktop = _get_desktop_dir()
    return _ensure_dir(os.path.join(desktop, "Formatos Inclusion Laboral", "SEGUIMIENTOS"))


def _get_roots():
    local_root = _get_local_root()
    shared_root = _get_shared_root()
    try:
        _ensure_dir(shared_root)
        shared_ok = True
    except Exception:
        shared_ok = False
    return {
        "local": local_root,
        "shared": shared_root if shared_ok else None,
    }


def _find_template_path():
    templates_dir = resource_path("templates")
    if not templates_dir.is_dir():
        raise FileNotFoundError("No existe la carpeta templates.")
    for name in os.listdir(templates_dir):
        if name.startswith("~$"):
            continue
        if name.lower().endswith(".xlsx") and "seguimiento" in name.lower():
            return os.fspath(templates_dir / name)
    raise FileNotFoundError("No se encontró el template de seguimientos.")


def _parse_first_name_lastname(full_name):
    tokens = [t for t in re.split(r"\s+", str(full_name or "").strip()) if t]
    if not tokens:
        return "Usuario", "SinApellido"
    first_name = tokens[0]
    if len(tokens) >= 4:
        first_lastname = tokens[2]
    elif len(tokens) >= 2:
        first_lastname = tokens[1]
    else:
        first_lastname = "SinApellido"
    return first_name, first_lastname


def build_case_folder_name(nombre_usuario, cedula):
    first_name, first_lastname = _parse_first_name_lastname(nombre_usuario)
    base = f"{first_name} {first_lastname} - {cedula}"
    return _sanitize_filename(base)


def _find_excel_in_folder(folder_path):
    if not os.path.isdir(folder_path):
        return None
    for name in os.listdir(folder_path):
        if name.startswith("~$"):
            continue
        if name.lower().endswith(".xlsx"):
            return os.path.join(folder_path, name)
    return None


def _scan_case_folder_by_cedula(root, cedula):
    if not root or not os.path.isdir(root):
        return None
    suffix = f"- {cedula}"
    for name in os.listdir(root):
        full = os.path.join(root, name)
        if not os.path.isdir(full):
            continue
        if name.endswith(suffix):
            found = _find_excel_in_folder(full)
            if found:
                return found
    return None


def find_case_workbook(cedula, nombre_usuario=""):
    normalized = _normalize_cedula(cedula)
    if not normalized:
        return None
    roots = _get_roots()
    folder_name = build_case_folder_name(nombre_usuario, normalized)
    ordered_roots = [roots.get("shared"), roots.get("local")]
    for root in ordered_roots:
        if not root:
            continue
        direct = _find_excel_in_folder(os.path.join(root, folder_name))
        if direct:
            return direct
        scanned = _scan_case_folder_by_cedula(root, normalized)
        if scanned:
            return scanned
    return None


def find_case_record(cedula, nombre_usuario=""):
    normalized = _normalize_cedula(cedula)
    if not normalized:
        return None
    try:
        service = _get_drive_service()
        folder = _find_case_folder_drive(service, normalized, nombre_usuario=nombre_usuario)
        if folder:
            files = _list_drive_files(service, folder["id"])
            case_file = _pick_case_file(files)
            if case_file:
                record = {
                    "source": "drive",
                    "cedula": normalized,
                    "folder_id": folder.get("id"),
                    "folder_name": folder.get("name"),
                    "file_id": case_file.get("id"),
                    "file_name": case_file.get("name"),
                    "mime_type": case_file.get("mimeType"),
                    "webViewLink": case_file.get("webViewLink"),
                    "modifiedTime": case_file.get("modifiedTime"),
                }
                if record["mime_type"] == GOOGLE_SHEETS_MIME:
                    record["local_path"] = ""
                    app_props = case_file.get("appProperties") or {}
                    try:
                        record["max_seguimientos"] = int(app_props.get("max_seguimientos") or 3)
                    except Exception:
                        record["max_seguimientos"] = 3
                else:
                    local_path = _build_local_case_copy_path(record)
                    record["local_path"] = _download_drive_case_to_path(service, record, local_path)
                    try:
                        record["max_seguimientos"] = int(
                            get_case_meta(record["local_path"]).get("max_seguimientos") or 3
                        )
                    except Exception:
                        record["max_seguimientos"] = 3
                return record
    except Exception:
        pass

    legacy_path = find_case_workbook(normalized, nombre_usuario)
    if legacy_path:
        try:
            max_seg = int(get_case_meta(legacy_path).get("max_seguimientos") or 3)
        except Exception:
            max_seg = 3
        return {
            "source": "legacy_local",
            "cedula": normalized,
            "file_name": os.path.basename(legacy_path),
            "local_path": legacy_path,
            "webViewLink": "",
            "max_seguimientos": max_seg,
        }
    return None


def _get_str(value):
    if value is None:
        return ""
    return str(value).strip()


def _set_if_empty(ws, cell, value):
    if ws[cell].value in (None, "") and value not in (None, ""):
        ws[cell].value = value


def _fill_sheet_base(wb, user_row):
    try:
        ws = wb[_get_base_sheet_name_from_workbook(wb)]
    except Exception:
        return
    company = get_linked_company_for_user(user_row)
    _set_if_empty(ws, "D9", _get_str(company.get("nombre_empresa")))
    _set_if_empty(ws, "R9", _get_str(company.get("ciudad_empresa")))
    _set_if_empty(ws, "D10", _get_str(company.get("direccion_empresa")))
    _set_if_empty(ws, "R10", _get_str(company.get("nit_empresa")))
    _set_if_empty(ws, "D11", _get_str(company.get("correo_1")))
    _set_if_empty(ws, "R11", _get_str(company.get("telefono_empresa")))
    _set_if_empty(ws, "D12", _get_str(company.get("contacto_empresa")))
    _set_if_empty(ws, "R12", _get_str(company.get("cargo")))
    _set_if_empty(ws, "D13", _get_str(company.get("asesor")))
    _set_if_empty(ws, "R13", _get_str(company.get("sede_empresa")))
    _set_if_empty(ws, "A16", _get_str(user_row.get("nombre_usuario")))
    _set_if_empty(ws, "E16", _get_str(user_row.get("cedula_usuario")))
    _set_if_empty(ws, "I16", _get_str(user_row.get("telefono_oferente")))
    _set_if_empty(ws, "K16", _get_str(user_row.get("correo_oferente")))
    _set_if_empty(ws, "P16", _get_str(user_row.get("contacto_emergencia")))
    _set_if_empty(ws, "S16", _get_str(user_row.get("parentesco")))
    _set_if_empty(ws, "U16", _get_str(user_row.get("telefono_emergencia")))
    _set_if_empty(ws, "A18", _get_str(user_row.get("cargo_oferente")))
    _set_if_empty(ws, "E18", _get_str(user_row.get("certificado_discapacidad")))
    _set_if_empty(ws, "I18", _get_str(user_row.get("certificado_porcentaje")))
    discapacidad = _get_str(user_row.get("discapacidad_detalle")) or _get_str(
        user_row.get("discapacidad_usuario")
    )
    _set_if_empty(ws, "N18", discapacidad)
    _set_if_empty(ws, "C20", _get_str(user_row.get("tipo_contrato")))


def _build_base_payload_from_user_row(user_row):
    discapacidad = _get_str(user_row.get("discapacidad_detalle")) or _get_str(
        user_row.get("discapacidad_usuario")
    )
    payload = {
        "nombre_vinculado": _get_str(user_row.get("nombre_usuario")),
        "cedula": _get_str(user_row.get("cedula_usuario")),
        "telefono_vinculado": _get_str(user_row.get("telefono_oferente")),
        "correo_vinculado": _get_str(user_row.get("correo_oferente")),
        "contacto_emergencia": _get_str(user_row.get("contacto_emergencia")),
        "parentesco": _get_str(user_row.get("parentesco")),
        "telefono_emergencia": _get_str(user_row.get("telefono_emergencia")),
        "cargo_vinculado": _get_str(user_row.get("cargo_oferente")),
        "certificado_discapacidad": _get_str(user_row.get("certificado_discapacidad")),
        "certificado_porcentaje": _get_str(user_row.get("certificado_porcentaje")),
        "discapacidad": discapacidad,
        "tipo_contrato": _get_str(user_row.get("tipo_contrato")),
    }
    payload.update(get_linked_company_for_user(user_row))
    return payload


def _build_ponderado_payload(base_payload):
    payload = dict(base_payload or {})
    fecha_firma = _get_str(payload.get("fecha_firma_contrato"))
    if not fecha_firma:
        fecha_firma = _get_str(payload.get("fecha_inicio_contrato"))
    payload["fecha_firma_contrato"] = fecha_firma
    return payload


def _sync_ponderado_from_payload(wb, base_payload, overwrite=True):
    if SHEET_FINAL not in wb.sheetnames:
        return
    ws = wb[SHEET_FINAL]
    payload = _build_ponderado_payload(base_payload)
    for field_id, cell in PONDERADO_COMPANY_MAP.items():
        if field_id in payload:
            if overwrite or ws[cell].value in (None, ""):
                ws[cell].value = payload.get(field_id, "")
    for field_id, cell in PONDERADO_USER_MAP.items():
        if field_id in payload:
            if overwrite or ws[cell].value in (None, ""):
                ws[cell].value = payload.get(field_id, "")


def _build_base_sheet_updates(payload, base_sheet_name=SHEET_BASE):
    updates = []
    mapping = {
        "fecha_visita": "D8",
        "modalidad": "R8",
        "nombre_empresa": "D9",
        "ciudad_empresa": "R9",
        "direccion_empresa": "D10",
        "nit_empresa": "R10",
        "correo_1": "D11",
        "telefono_empresa": "R11",
        "contacto_empresa": "D12",
        "cargo": "R12",
        "asesor": "D13",
        "sede_empresa": "R13",
        "nombre_vinculado": "A16",
        "cedula": "E16",
        "telefono_vinculado": "I16",
        "correo_vinculado": "K16",
        "contacto_emergencia": "P16",
        "parentesco": "S16",
        "telefono_emergencia": "U16",
        "cargo_vinculado": "A18",
        "certificado_discapacidad": "E18",
        "certificado_porcentaje": "I18",
        "discapacidad": "N18",
        "tipo_contrato": "C20",
        "fecha_inicio_contrato": "M20",
        "fecha_fin_contrato": "T20",
        "apoyos_ajustes": "E21",
    }
    for field_id, cell in mapping.items():
        updates.append({"range": f"'{base_sheet_name}'!{cell}", "value": payload.get(field_id, "")})
    for idx, row in enumerate(range(23, 28)):
        updates.append({"range": f"'{base_sheet_name}'!B{row}", "value": (payload.get('funciones_1_5') or [''] * 5)[idx] if idx < len(payload.get('funciones_1_5') or []) else ""})
        updates.append({"range": f"'{base_sheet_name}'!N{row}", "value": (payload.get('funciones_6_10') or [''] * 5)[idx] if idx < len(payload.get('funciones_6_10') or []) else ""})
    for idx, row in enumerate(range(29, 32)):
        updates.append({"range": f"'{base_sheet_name}'!C{row}", "value": (payload.get('seguimiento_fechas_1_3') or [''] * 3)[idx] if idx < len(payload.get('seguimiento_fechas_1_3') or []) else ""})
        updates.append({"range": f"'{base_sheet_name}'!P{row}", "value": (payload.get('seguimiento_fechas_4_6') or [''] * 3)[idx] if idx < len(payload.get('seguimiento_fechas_4_6') or []) else ""})
    ponderado_payload = _build_ponderado_payload(payload)
    for field_id, cell in PONDERADO_COMPANY_MAP.items():
        updates.append({"range": f"'{SHEET_FINAL}'!{cell}", "value": ponderado_payload.get(field_id, "")})
    for field_id, cell in PONDERADO_USER_MAP.items():
        updates.append({"range": f"'{SHEET_FINAL}'!{cell}", "value": ponderado_payload.get(field_id, "")})
    return updates


def _build_followup_sheet_updates(index, payload):
    sheet_name = _get_followup_sheet_name(index)
    updates = [
        {"range": f"'{sheet_name}'!E8", "value": payload.get("modalidad", "")},
        {"range": f"'{sheet_name}'!P8", "value": payload.get("seguimiento_numero", index)},
        {"range": f"'{sheet_name}'!J31", "value": payload.get("tipo_apoyo", "")},
        {"range": f"'{sheet_name}'!A43", "value": payload.get("situacion_encontrada", "")},
        {"range": f"'{sheet_name}'!A45", "value": payload.get("estrategias_ajustes", "")},
    ]
    for i, row in enumerate(range(12, 31)):
        updates.append({"range": f"'{sheet_name}'!G{row}", "value": (payload.get('item_observaciones') or [])[i] if i < len(payload.get('item_observaciones') or []) else ""})
        updates.append({"range": f"'{sheet_name}'!O{row}", "value": (payload.get('item_autoevaluacion') or [])[i] if i < len(payload.get('item_autoevaluacion') or []) else ""})
        updates.append({"range": f"'{sheet_name}'!R{row}", "value": (payload.get('item_eval_empresa') or [])[i] if i < len(payload.get('item_eval_empresa') or []) else ""})
    for i, row in enumerate(range(34, 42)):
        updates.append({"range": f"'{sheet_name}'!J{row}", "value": (payload.get('empresa_eval') or [])[i] if i < len(payload.get('empresa_eval') or []) else ""})
        updates.append({"range": f"'{sheet_name}'!L{row}", "value": (payload.get('empresa_observacion') or [])[i] if i < len(payload.get('empresa_observacion') or []) else ""})
    for i, row in enumerate(range(47, 51)):
        asistentes = payload.get("asistentes") or []
        entry = asistentes[i] if i < len(asistentes) else {}
        updates.append({"range": f"'{sheet_name}'!D{row}", "value": entry.get("nombre", "")})
        updates.append({"range": f"'{sheet_name}'!N{row}", "value": entry.get("cargo", "")})
    return updates


def _build_empty_followup_payload(index):
    return {
        "modalidad": "",
        "seguimiento_numero": "",
        "item_observaciones": ["" for _ in range(19)],
        "item_autoevaluacion": ["" for _ in range(19)],
        "item_eval_empresa": ["" for _ in range(19)],
        "tipo_apoyo": "",
        "empresa_eval": ["" for _ in range(8)],
        "empresa_observacion": ["" for _ in range(8)],
        "situacion_encontrada": "",
        "estrategias_ajustes": "",
        "asistentes": [{"nombre": "", "cargo": ""} for _ in range(4)],
    }


def _apply_empty_followup_fields(payload, index):
    payload = dict(payload or {})
    payload.update(_build_empty_followup_payload(index))
    return payload


def _clear_followup_sheet_worksheet(ws, index):
    empty_payload = _build_empty_followup_payload(index)
    ws["E8"].value = empty_payload["modalidad"]
    ws["P8"].value = empty_payload["seguimiento_numero"]
    for pos, row in enumerate(range(12, 31)):
        ws[f"G{row}"].value = empty_payload["item_observaciones"][pos]
        ws[f"O{row}"].value = empty_payload["item_autoevaluacion"][pos]
        ws[f"R{row}"].value = empty_payload["item_eval_empresa"][pos]
    ws["J31"].value = empty_payload["tipo_apoyo"]
    for pos, row in enumerate(range(34, 42)):
        ws[f"J{row}"].value = empty_payload["empresa_eval"][pos]
        ws[f"L{row}"].value = empty_payload["empresa_observacion"][pos]
    ws["A43"].value = empty_payload["situacion_encontrada"]
    ws["A45"].value = empty_payload["estrategias_ajustes"]
    for pos, row in enumerate(range(47, 51)):
        ws[f"D{row}"].value = empty_payload["asistentes"][pos]["nombre"]
        ws[f"N{row}"].value = empty_payload["asistentes"][pos]["cargo"]


@lru_cache(maxsize=6)
def _get_followup_template_defaults(index):
    workbook = _load_workbook_safe(_find_template_path(), data_only=False)
    try:
        ws = _ensure_sheet_exists(workbook, _get_followup_sheet_name(index))
        return {
            "modalidad": _cell_value(ws, "E8"),
            "item_observaciones": [_cell_value(ws, f"G{r}") for r in range(12, 31)],
            "item_autoevaluacion": [_cell_value(ws, f"O{r}") for r in range(12, 31)],
            "item_eval_empresa": [_cell_value(ws, f"R{r}") for r in range(12, 31)],
            "tipo_apoyo": _cell_value(ws, "J31"),
            "empresa_eval": [_cell_value(ws, f"J{r}") for r in range(34, 42)],
            "empresa_observacion": [_cell_value(ws, f"L{r}") for r in range(34, 42)],
            "situacion_encontrada": _cell_value(ws, "A43"),
            "estrategias_ajustes": _cell_value(ws, "A45"),
            "asistentes": [
                {"nombre": _cell_value(ws, f"D{r}"), "cargo": _cell_value(ws, f"N{r}")}
                for r in range(47, 51)
            ],
        }
    finally:
        workbook.close()


def _is_template_seeded_followup_payload(payload, index):
    defaults = _get_followup_template_defaults(index)
    payload = payload or {}
    fields = (
        "modalidad",
        "item_observaciones",
        "item_autoevaluacion",
        "item_eval_empresa",
        "tipo_apoyo",
        "empresa_eval",
        "empresa_observacion",
        "situacion_encontrada",
        "estrategias_ajustes",
        "asistentes",
    )
    for field in fields:
        if payload.get(field) != defaults.get(field):
            return False
    return True


def _ensure_meta_sheet(wb, cedula, nombre_usuario, is_compensar, max_seguimientos):
    if SHEET_META in wb.sheetnames:
        ws = wb[SHEET_META]
    else:
        ws = wb.create_sheet(SHEET_META)
    ws.sheet_state = "hidden"
    ws["A1"] = "cedula"
    ws["B1"] = _get_str(cedula)
    ws["A2"] = "nombre_usuario"
    ws["B2"] = _get_str(nombre_usuario)
    ws["A3"] = "is_compensar"
    ws["B3"] = "1" if is_compensar else "0"
    ws["A4"] = "max_seguimientos"
    ws["B4"] = int(max_seguimientos)


def _read_meta(wb):
    if SHEET_META not in wb.sheetnames:
        return {}
    ws = wb[SHEET_META]
    meta = {}
    for row in range(1, 15):
        key = _get_str(ws[f"A{row}"].value)
        val = ws[f"B{row}"].value
        if key:
            meta[key] = val
    return meta


def _apply_visibility(wb, max_seguimientos):
    relevant_titles = {
        SHEET_FINAL,
        *BASE_SHEET_CANDIDATES,
        *(f"{SHEET_PREFIX}{i}" for i in range(1, 7)),
    }
    for name in wb.sheetnames:
        if name == SHEET_META:
            wb[name].sheet_state = "hidden"
            continue
        wb[name].sheet_state = "visible" if name in relevant_titles else "hidden"


def _infer_max_seguimientos_from_workbook(wb):
    visible = 0
    total = 0
    for i in range(1, 7):
        name = f"{SHEET_PREFIX}{i}"
        if name not in wb.sheetnames:
            continue
        total += 1
        ws = wb[name]
        if str(getattr(ws, "sheet_state", "visible")).lower() != "hidden":
            visible += 1
    if visible >= 6:
        return 6
    if visible >= 3:
        return 3
    if total >= 6:
        return 6
    return 3


def _copy_to_secondary_roots(primary_path):
    roots = _get_roots()
    local = roots.get("local")
    shared = roots.get("shared")
    if not primary_path or not os.path.exists(primary_path):
        return
    folder_name = os.path.basename(os.path.dirname(primary_path))
    filename = os.path.basename(primary_path)
    targets = []
    if shared and not os.path.normcase(primary_path).startswith(os.path.normcase(shared)):
        targets.append(os.path.join(shared, folder_name, filename))
    if local and not os.path.normcase(primary_path).startswith(os.path.normcase(local)):
        targets.append(os.path.join(local, folder_name, filename))
    for target in targets:
        try:
            _ensure_dir(os.path.dirname(target))
            shutil.copy2(primary_path, target)
        except Exception:
            continue


def ensure_case_workbook(cedula, user_row, is_compensar):
    normalized = _normalize_cedula(cedula)
    if not normalized:
        raise ValueError("Cédula inválida.")
    if not user_row:
        raise ValueError("No se encontró usuario para la cédula indicada.")

    existing = find_case_workbook(normalized, user_row.get("nombre_usuario"))
    if existing:
        wb = _load_workbook_safe(existing)
        meta = _read_meta(wb)
        if meta:
            max_seguimientos = int(meta.get("max_seguimientos") or (6 if is_compensar else 3))
            is_comp = str(meta.get("is_compensar") or "0").strip() in ("1", "true", "True")
        else:
            is_comp = bool(is_compensar)
            max_seguimientos = 6 if is_comp else 3
        _fill_sheet_base(wb, user_row)
        _sync_ponderado_from_payload(wb, _build_base_payload_from_user_row(user_row), overwrite=False)
        _ensure_meta_sheet(
            wb,
            normalized,
            user_row.get("nombre_usuario"),
            is_comp,
            max_seguimientos,
        )
        _apply_visibility(wb, max_seguimientos)
        sanitize_logo_error_cells(wb)
        wb.save(existing)
        _copy_to_secondary_roots(existing)
        return {"path": existing, "created": False, "max_seguimientos": max_seguimientos}

    template_path = _find_template_path()
    roots = _get_roots()
    primary_root = roots.get("shared") or roots.get("local")
    if not primary_root:
        raise RuntimeError("No hay ruta disponible para guardar seguimientos.")

    folder_name = build_case_folder_name(user_row.get("nombre_usuario"), normalized)
    case_folder = _ensure_dir(os.path.join(primary_root, folder_name))
    output_path = os.path.join(case_folder, f"{folder_name}.xlsx")
    shutil.copy2(template_path, output_path)

    max_seguimientos = 6 if bool(is_compensar) else 3
    wb = _load_workbook_safe(output_path)
    _fill_sheet_base(wb, user_row)
    _sync_ponderado_from_payload(wb, _build_base_payload_from_user_row(user_row), overwrite=False)
    for idx in range(1, 7):
        _clear_followup_sheet_worksheet(_ensure_sheet_exists(wb, _get_followup_sheet_name(idx)), idx)
    _ensure_meta_sheet(
        wb,
        normalized,
        user_row.get("nombre_usuario"),
        bool(is_compensar),
        max_seguimientos,
    )
    _apply_visibility(wb, max_seguimientos)
    sanitize_logo_error_cells(wb)
    wb.save(output_path)
    _copy_to_secondary_roots(output_path)
    return {"path": output_path, "created": True, "max_seguimientos": max_seguimientos}


def _set_sheet_visibility(spreadsheet_id, max_seguimientos):
    try:
        service = get_google_sheets_service()
        spreadsheet = get_spreadsheet(spreadsheet_id, include_grid_data=False)
    except Exception:
        return
    requests = []
    try:
        base_sheet_name = _get_base_sheet_name_from_spreadsheet(spreadsheet)
    except Exception:
        base_sheet_name = SHEET_BASE
    relevant_titles = {
        base_sheet_name,
        SHEET_FINAL,
        *(f"{SHEET_PREFIX}{i}" for i in range(1, 7)),
    }
    for sheet in spreadsheet.get("sheets", []):
        props = sheet.get("properties", {}) or {}
        title = str(props.get("title") or "")
        try:
            hidden = bool(title not in relevant_titles)
        except Exception:
            hidden = False
        requests.append(
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": props.get("sheetId"),
                        "hidden": hidden,
                    },
                    "fields": "hidden",
                }
            }
        )
    if not requests:
        return
    execute_google_request_with_retry(
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests},
        ),
        operation_name="seguimientos.set_sheet_visibility",
    )


def _merge_fill_empty(existing_payload, new_payload):
    merged = dict(existing_payload or {})
    for key, value in (new_payload or {}).items():
        current = merged.get(key)
        if isinstance(value, list):
            current_list = list(current or [])
            merged_list = []
            for idx, item in enumerate(value):
                current_value = current_list[idx] if idx < len(current_list) else ""
                merged_list.append(current_value if str(current_value or "").strip() else item)
            merged[key] = merged_list
        else:
            merged[key] = current if str(current or "").strip() else value
    return merged


def _create_native_case_record(service, folder_id, folder_name, cedula, user_row, max_seguimientos, seed_path=None):
    template_id = get_seguimientos_template_id()
    request_id = uuid.uuid4().hex
    copied = execute_google_create_with_confirmation(
        lambda: service.files().copy(
            fileId=template_id,
            body={
                "name": folder_name,
                "parents": [folder_id],
                "appProperties": {
                    "kind": "seguimiento_il",
                    "request_id": request_id,
                    "cedula": str(cedula),
                    "max_seguimientos": str(max_seguimientos),
                },
            },
            fields="id,name,mimeType,webViewLink,parents,appProperties",
            supportsAllDrives=True,
        ),
        lambda: _find_drive_file_by_request_id(service, folder_id, folder_name, request_id),
        operation_name="seguimientos.copy_native_case",
    )
    record = {
        "source": "drive",
        "cedula": cedula,
        "folder_id": folder_id,
        "folder_name": folder_name,
        "file_id": copied.get("id"),
        "file_name": copied.get("name"),
        "mime_type": copied.get("mimeType") or GOOGLE_SHEETS_MIME,
        "webViewLink": copied.get("webViewLink") or f"https://docs.google.com/spreadsheets/d/{copied.get('id')}/edit",
        "modifiedTime": "",
        "local_path": "",
        "max_seguimientos": max_seguimientos,
        "appProperties": copied.get("appProperties") or {},
    }
    base_payload = _build_base_payload_from_user_row(user_row)
    if seed_path:
        base_payload = _merge_fill_empty(get_base_payload(seed_path), base_payload)
    clear_protected_ranges(record["file_id"])
    spreadsheet = get_spreadsheet(record["file_id"], include_grid_data=False)
    base_sheet_name = _get_base_sheet_name_from_spreadsheet(spreadsheet)
    updates = _build_base_sheet_updates(base_payload, base_sheet_name=base_sheet_name)
    for idx in range(1, 7):
        updates.extend(_build_followup_sheet_updates(idx, _build_empty_followup_payload(idx)))
    if seed_path:
        for idx in range(1, max_seguimientos + 1):
            updates.extend(_build_followup_sheet_updates(idx, get_followup_payload(seed_path, idx)))
    batch_write_sheet_updates(record["file_id"], updates)
    _set_sheet_visibility(record["file_id"], max_seguimientos)
    return record


def ensure_case_record(cedula, user_row, is_compensar):
    normalized = _normalize_cedula(cedula)
    if not normalized:
        raise ValueError("Cédula inválida.")
    existing = find_case_record(normalized, user_row.get("nombre_usuario"))
    if existing:
        if _is_native_case_ref(existing):
            save_base_payload(existing, _build_base_payload_from_user_row(user_row), overwrite=False)
            try:
                existing["max_seguimientos"] = int(
                    get_case_meta(existing).get("max_seguimientos") or 3
                )
            except Exception:
                existing["max_seguimientos"] = 6 if bool(is_compensar) else 3
        elif existing.get("local_path") and existing.get("source") != "legacy_local":
            wb = _load_workbook_safe(existing["local_path"])
            _fill_sheet_base(wb, user_row)
            _sync_ponderado_from_payload(wb, _build_base_payload_from_user_row(user_row), overwrite=False)
            sanitize_logo_error_cells(wb)
            wb.save(existing["local_path"])
            if existing.get("source") == "drive":
                sync_case_record_from_local(existing, existing["local_path"])
            try:
                existing["max_seguimientos"] = int(
                    get_case_meta(existing["local_path"]).get("max_seguimientos") or 3
                )
            except Exception:
                existing["max_seguimientos"] = 6 if bool(is_compensar) else 3
        elif existing.get("source") == "drive":
            service = _get_drive_service()
            max_seguimientos = int(existing.get("max_seguimientos") or (6 if bool(is_compensar) else 3))
            existing = _create_native_case_record(
                service,
                existing["folder_id"],
                existing["folder_name"],
                normalized,
                user_row,
                max_seguimientos,
                seed_path=existing.get("local_path"),
            )
        elif existing.get("source") == "legacy_local":
            service = _get_drive_service()
            seguimientos_folder_id = _ensure_seguimientos_folder(service)
            folder_name = build_case_folder_name(user_row.get("nombre_usuario"), normalized)
            case_folder_id = drive_upload._get_or_create_folder(service, seguimientos_folder_id, folder_name)
            max_seguimientos = int(existing.get("max_seguimientos") or (6 if bool(is_compensar) else 3))
            existing = _create_native_case_record(
                service,
                case_folder_id,
                folder_name,
                normalized,
                user_row,
                max_seguimientos,
                seed_path=existing.get("local_path"),
            )
        return {
            "record": existing,
            "created": False,
            "max_seguimientos": int(existing.get("max_seguimientos") or (6 if bool(is_compensar) else 3)),
        }

    service = _get_drive_service()
    seguimientos_folder_id = _ensure_seguimientos_folder(service)
    folder_name = build_case_folder_name(user_row.get("nombre_usuario"), normalized)
    case_folder_id = drive_upload._get_or_create_folder(service, seguimientos_folder_id, folder_name)
    max_seguimientos = 6 if bool(is_compensar) else 3
    record = _create_native_case_record(
        service,
        case_folder_id,
        folder_name,
        normalized,
        user_row,
        max_seguimientos,
    )
    return {"record": record, "created": True, "max_seguimientos": max_seguimientos}


def sync_case_record_from_local(record, local_path):
    if not record or not local_path or not os.path.exists(local_path):
        return
    source = str(record.get("source") or "").strip()
    if source == "legacy_local":
        return
    service = _get_drive_service()
    mime_type = str(record.get("mime_type") or "").strip()
    if mime_type == GOOGLE_SHEETS_MIME:
        return
    media = MediaFileUpload(local_path, mimetype=XLSX_MIME, resumable=False)
    execute_google_request_with_retry(
        service.files().update(
            fileId=record["file_id"],
            media_body=media,
            fields="id,name,modifiedTime,webViewLink",
            supportsAllDrives=True,
        ),
        operation_name="seguimientos.sync_case_record_from_local",
    )


def _get_base_sheet_name_from_spreadsheet(spreadsheet):
    for sheet in spreadsheet.get("sheets", []):
        title = str(((sheet.get("properties") or {}).get("title")) or "")
        if title in BASE_SHEET_CANDIDATES:
            return title
    raise ValueError("No existe la hoja base de seguimientos en el spreadsheet.")


def _infer_max_seguimientos_from_spreadsheet(spreadsheet):
    visible = 0
    total = 0
    for sheet in spreadsheet.get("sheets", []):
        props = sheet.get("properties") or {}
        title = str(props.get("title") or "")
        if not title.startswith(SHEET_PREFIX):
            continue
        total += 1
        if not bool(props.get("hidden")):
            visible += 1
    if visible >= 6:
        return 6
    if visible >= 3:
        return 3
    if total >= 6:
        return 6
    return 3


def describe_case(case_ref):
    if not case_ref:
        return {
            "empresa": "",
            "seguimientos": [],
            "seguimientos_count": 0,
            "ultimo_seguimiento": "",
        }
    payload = get_base_payload(case_ref)
    followup_dates = []
    for idx, value in enumerate(payload.get("seguimiento_fechas_1_3") or [], start=1):
        if str(value or "").strip():
            followup_dates.append(f"S{idx}: {str(value).strip()}")
    for offset, value in enumerate(payload.get("seguimiento_fechas_4_6") or [], start=4):
        if str(value or "").strip():
            followup_dates.append(f"S{offset}: {str(value).strip()}")
    return {
        "empresa": str(payload.get("nombre_empresa") or "").strip(),
        "seguimientos": followup_dates,
        "seguimientos_count": len(followup_dates),
        "ultimo_seguimiento": followup_dates[-1] if followup_dates else "",
    }


def _cell_has_value(ws, cell):
    value = ws[cell].value
    if value is None:
        return False
    return str(value).strip() != ""


def _is_base_completed(wb):
    try:
        ws = wb[_get_base_sheet_name_from_workbook(wb)]
    except Exception:
        return False
    required = ["A16", "E16", "A18", "N18"]
    return all(_cell_has_value(ws, c) for c in required)


def _is_followup_completed(ws):
    required = ["O12", "R12", "J31"]
    return all(_cell_has_value(ws, c) for c in required)


def _is_base_payload_completed(payload):
    payload = payload or {}
    required = ["nombre_vinculado", "cedula", "cargo_vinculado", "discapacidad"]
    return all(str(payload.get(item) or "").strip() for item in required)


def _is_followup_payload_completed(payload):
    payload = payload or {}
    auto_vals = payload.get("item_autoevaluacion") or []
    emp_vals = payload.get("item_eval_empresa") or []
    first_auto = str(auto_vals[0] or "").strip() if auto_vals else ""
    first_emp = str(emp_vals[0] or "").strip() if emp_vals else ""
    tipo_apoyo = str(payload.get("tipo_apoyo") or "").strip()
    return bool(first_auto and first_emp and tipo_apoyo)


def _get_followup_date_from_base(base_payload, index):
    try:
        idx = int(index)
    except Exception:
        return ""
    if 1 <= idx <= 3:
        values = base_payload.get("seguimiento_fechas_1_3") or []
        pos = idx - 1
    elif 4 <= idx <= 6:
        values = base_payload.get("seguimiento_fechas_4_6") or []
        pos = idx - 4
    else:
        return ""
    return str(values[pos] or "").strip() if pos < len(values) else ""


def _has_followup_meaningful_content(payload):
    payload = payload or {}
    if str(payload.get("situacion_encontrada") or "").strip():
        return True
    if str(payload.get("estrategias_ajustes") or "").strip():
        return True
    for value in (payload.get("item_observaciones") or []):
        if str(value or "").strip():
            return True
    for value in (payload.get("empresa_observacion") or []):
        if str(value or "").strip():
            return True
    return False


def _normalize_empty_followup_payload(payload, index, base_payload=None):
    payload = dict(payload or {})
    has_followup_date = bool(_get_followup_date_from_base(base_payload or {}, index))
    if has_followup_date or _has_followup_meaningful_content(payload):
        return payload
    payload["modalidad"] = ""
    payload["seguimiento_numero"] = str(index)
    payload["item_observaciones"] = ["" for _ in (payload.get("item_observaciones") or [])]
    payload["item_autoevaluacion"] = ["" for _ in (payload.get("item_autoevaluacion") or [])]
    payload["item_eval_empresa"] = ["" for _ in (payload.get("item_eval_empresa") or [])]
    payload["tipo_apoyo"] = ""
    payload["empresa_eval"] = ["" for _ in (payload.get("empresa_eval") or [])]
    payload["empresa_observacion"] = ["" for _ in (payload.get("empresa_observacion") or [])]
    payload["situacion_encontrada"] = ""
    payload["estrategias_ajustes"] = ""
    payload["asistentes"] = [{"nombre": "", "cargo": ""} for _ in (payload.get("asistentes") or [])]
    return payload


def _normalize_base_payload(payload):
    payload = dict(payload or {})
    has_started = any(
        [
            str(payload.get("fecha_visita") or "").strip(),
            str(payload.get("modalidad") or "").strip(),
            str(payload.get("apoyos_ajustes") or "").strip(),
        ]
    )
    for value in (payload.get("seguimiento_fechas_1_3") or []):
        if str(value or "").strip():
            has_started = True
            break
    if not has_started:
        for value in (payload.get("seguimiento_fechas_4_6") or []):
            if str(value or "").strip():
                has_started = True
                break
    if not has_started:
        payload["fecha_fin_contrato"] = ""
    return payload


def get_workflow_state(case_ref):
    if not case_ref:
        return {
            "base_sheet_name": SHEET_BASE,
            "max_seguimientos": 3,
            "base_completed": False,
            "completed_followups": [],
            "next_followup": 1,
            "editable_sheet": SHEET_BASE,
            "suggested_sheet": SHEET_BASE,
            "visible_sheets": [SHEET_BASE, f"{SHEET_PREFIX}1"],
            "message": "Completa primero la hoja base del proceso.",
        }

    meta = get_case_meta(case_ref)
    base_sheet_name = str(meta.get("base_sheet_name") or SHEET_BASE)
    max_seguimientos = 6 if int(meta.get("max_seguimientos") or 3) >= 6 else 3
    base_payload = get_base_payload(case_ref)
    base_completed = _is_base_payload_completed(base_payload)

    completed_followups = []
    next_followup = 1
    if base_completed:
        next_followup = None
        for idx in range(1, max_seguimientos + 1):
            has_followup_date = bool(_get_followup_date_from_base(base_payload, idx))
            followup_payload = get_followup_payload(case_ref, idx)
            followup_completed = has_followup_date and _is_followup_payload_completed(
                followup_payload
            )
            if followup_completed:
                completed_followups.append(idx)
                continue
            next_followup = idx
            break

    if not base_completed:
        editable_sheet = base_sheet_name
        suggested_sheet = base_sheet_name
        visible_followups = 1
        message = "Completa primero la hoja base del proceso."
    elif next_followup is not None and next_followup <= max_seguimientos:
        editable_sheet = f"{SHEET_PREFIX}{next_followup}"
        suggested_sheet = base_sheet_name if next_followup == 1 else editable_sheet
        visible_followups = next_followup
        if next_followup == 1:
            message = "Hoja base y seguimiento 1 habilitados hasta diligenciar el seguimiento 1."
        else:
            message = (
                f"Seguimiento {next_followup} habilitado. "
                f"En la hoja base solo puedes editar la fecha del seguimiento {next_followup}."
            )
    else:
        editable_sheet = ""
        suggested_sheet = f"{SHEET_PREFIX}{max_seguimientos}"
        visible_followups = max_seguimientos
        message = "Todos los seguimientos están diligenciados."

    visible_sheets = [base_sheet_name] + [
        f"{SHEET_PREFIX}{idx}" for idx in range(1, visible_followups + 1)
    ]

    return {
        "base_sheet_name": base_sheet_name,
        "max_seguimientos": max_seguimientos,
        "base_completed": base_completed,
        "completed_followups": completed_followups,
        "next_followup": next_followup,
        "editable_sheet": editable_sheet,
        "suggested_sheet": suggested_sheet,
        "visible_sheets": visible_sheets,
        "message": message,
    }


def suggest_next_step(case_ref):
    if not case_ref:
        return {"sheet": SHEET_BASE, "message": "Inicia con la hoja base.", "max_seguimientos": 3}
    workflow = get_workflow_state(case_ref)
    return {
        "sheet": workflow.get("suggested_sheet") or workflow.get("base_sheet_name") or SHEET_BASE,
        "message": workflow.get("message") or "",
        "max_seguimientos": int(workflow.get("max_seguimientos") or 3),
    }


def get_usuarios_reca_cedulas(env_path=".env"):
    params = {
        "select": "cedula_usuario",
        "cedula_usuario": "not.is.null",
        "order": "cedula_usuario.asc",
    }
    data = _supabase_get("usuarios_reca", params, env_path=env_path)
    return [row.get("cedula_usuario") for row in data if row.get("cedula_usuario")]


def get_usuario_reca_by_cedula(cedula, env_path=".env"):
    normalized = _normalize_cedula(cedula)
    if not normalized:
        return None
    select_cols = ",".join(
        [
            "cedula_usuario",
            "nombre_usuario",
            "discapacidad_usuario",
            "discapacidad_detalle",
            "certificado_discapacidad",
            "certificado_porcentaje",
            "telefono_oferente",
            "correo_oferente",
            "cargo_oferente",
            "contacto_emergencia",
            "parentesco",
            "telefono_emergencia",
            "fecha_firma_contrato",
            "tipo_contrato",
            "fecha_fin",
            "empresa_nit",
            "empresa_nombre",
        ]
    )
    params = {
        "select": select_cols,
        "cedula_usuario": f"eq.{normalized}",
        "limit": 1,
    }
    data = _supabase_get("usuarios_reca", params, env_path=env_path)
    return data[0] if data else None


def get_empresa_by_nit(nit, env_path=".env"):
    if not nit:
        return None
    nit = "".join(str(nit).split())
    select_cols = ",".join(sorted(set(SECTION_1_SUPABASE_MAP.values()) | {"nit_empresa"}))
    params = {
        "select": select_cols,
        "nit_empresa": f"eq.{nit}",
        "limit": 1,
    }
    data = _supabase_get("empresas", params, env_path=env_path)
    return (
        _merge_company_row_with_cache(
            _map_company_row(data[0]),
            field_map=SECTION_1_SUPABASE_MAP,
            nit=nit,
        )
        if data
        else None
    )


def get_empresa_by_nombre(nombre, env_path=".env"):
    if not nombre:
        return None
    nombre = " ".join(str(nombre).split())
    select_cols = ",".join(sorted(set(SECTION_1_SUPABASE_MAP.values()) | {"nit_empresa"}))
    params = {
        "select": select_cols,
        "nombre_empresa": f"ilike.{nombre}",
        "limit": 2,
    }
    data = _supabase_get("empresas", params, env_path=env_path)
    if not data:
        return None
    if len(data) > 1:
        raise ValueError("Hay más de una empresa con ese nombre. Usa el NIT.")
    return _merge_company_row_with_cache(
        _map_company_row(data[0]),
        field_map=SECTION_1_SUPABASE_MAP,
        nombre=nombre,
    )


def get_empresas_by_nombre_prefix(prefix, env_path=".env", limit=50):
    text = str(prefix or "").strip()
    if not text:
        return []
    select_cols = ",".join(sorted(set(SECTION_1_SUPABASE_MAP.values()) | {"nit_empresa"}))
    params = {
        "select": select_cols,
        "nombre_empresa": f"ilike.{text}%",
        "order": "nombre_empresa.asc",
        "limit": int(limit),
    }
    return _supabase_get("empresas", params, env_path=env_path)


def get_empresas_by_nit_prefix(prefix, env_path=".env", limit=10):
    text = "".join(str(prefix or "").split())
    if not text:
        return []
    select_cols = ",".join(sorted(set(SECTION_1_SUPABASE_MAP.values()) | {"nit_empresa"}))
    params = {
        "select": select_cols,
        "nit_empresa": f"like.{text}%",
        "order": "nit_empresa.asc",
        "limit": int(limit),
    }
    return _supabase_get("empresas", params, env_path=env_path)


def _ensure_sheet_exists(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_name}' en el archivo.")
    return wb[sheet_name]


def _cell_value(ws, address):
    value = ws[address].value
    if value is None:
        return ""
    return str(value).strip()


def get_case_meta(case_ref):
    if _is_native_case_ref(case_ref):
        spreadsheet = get_spreadsheet(_get_spreadsheet_id_from_case_ref(case_ref), include_grid_data=False)
        props = case_ref.get("appProperties") or {}
        try:
            max_seg = int(case_ref.get("max_seguimientos") or props.get("max_seguimientos") or _infer_max_seguimientos_from_spreadsheet(spreadsheet))
        except Exception:
            max_seg = _infer_max_seguimientos_from_spreadsheet(spreadsheet)
        max_seg = 6 if max_seg >= 6 else 3
        return {
            "cedula": _get_str((case_ref.get("cedula") or props.get("cedula"))),
            "nombre_usuario": _get_str(case_ref.get("folder_name")),
            "is_compensar": bool(max_seg >= 6),
            "max_seguimientos": max_seg,
            "base_sheet_name": _get_base_sheet_name_from_spreadsheet(spreadsheet),
        }
    wb = _load_workbook_safe(case_ref, data_only=False)
    meta = _read_meta(wb)
    try:
        max_seg = int(meta.get("max_seguimientos") or _infer_max_seguimientos_from_workbook(wb))
    except Exception:
        max_seg = _infer_max_seguimientos_from_workbook(wb)
    max_seg = 6 if max_seg >= 6 else 3
    return {
        "cedula": _get_str(meta.get("cedula")),
        "nombre_usuario": _get_str(meta.get("nombre_usuario")),
        "is_compensar": str(meta.get("is_compensar") or "0").strip() in ("1", "true", "True"),
        "max_seguimientos": max_seg,
        "base_sheet_name": _get_base_sheet_name_from_workbook(wb),
    }


def get_base_payload(case_ref):
    if _is_native_case_ref(case_ref):
        meta = get_case_meta(case_ref)
        base_sheet_name = str(meta.get("base_sheet_name") or SHEET_BASE)
        ranges = [
            f"'{base_sheet_name}'!D8",
            f"'{base_sheet_name}'!R8",
            f"'{base_sheet_name}'!D9",
            f"'{base_sheet_name}'!R9",
            f"'{base_sheet_name}'!D10",
            f"'{base_sheet_name}'!R10",
            f"'{base_sheet_name}'!D11",
            f"'{base_sheet_name}'!R11",
            f"'{base_sheet_name}'!D12",
            f"'{base_sheet_name}'!R12",
            f"'{base_sheet_name}'!D13",
            f"'{base_sheet_name}'!R13",
            f"'{base_sheet_name}'!A16",
            f"'{base_sheet_name}'!E16",
            f"'{base_sheet_name}'!I16",
            f"'{base_sheet_name}'!K16",
            f"'{base_sheet_name}'!P16",
            f"'{base_sheet_name}'!S16",
            f"'{base_sheet_name}'!U16",
            f"'{base_sheet_name}'!A18",
            f"'{base_sheet_name}'!E18",
            f"'{base_sheet_name}'!I18",
            f"'{base_sheet_name}'!N18",
            f"'{base_sheet_name}'!C20",
            f"'{base_sheet_name}'!M20",
            f"'{base_sheet_name}'!T20",
            f"'{base_sheet_name}'!E21",
            f"'{base_sheet_name}'!B23:B27",
            f"'{base_sheet_name}'!N23:N27",
            f"'{base_sheet_name}'!C29:C31",
            f"'{base_sheet_name}'!P29:P31",
            f"'{SHEET_FINAL}'!D11",
            f"'{SHEET_FINAL}'!Q12",
            f"'{SHEET_FINAL}'!N18",
        ]
        values = _batch_read_sheet_values(_get_spreadsheet_id_from_case_ref(case_ref), ranges)
        payload = {
            "fecha_visita": _first_batch_value(values, f"'{base_sheet_name}'!D8"),
            "modalidad": _first_batch_value(values, f"'{base_sheet_name}'!R8"),
            "nombre_empresa": _first_batch_value(values, f"'{base_sheet_name}'!D9"),
            "ciudad_empresa": _first_batch_value(values, f"'{base_sheet_name}'!R9"),
            "direccion_empresa": _first_batch_value(values, f"'{base_sheet_name}'!D10"),
            "nit_empresa": _first_batch_value(values, f"'{base_sheet_name}'!R10"),
            "correo_1": _first_batch_value(values, f"'{base_sheet_name}'!D11"),
            "telefono_empresa": _first_batch_value(values, f"'{base_sheet_name}'!R11"),
            "contacto_empresa": _first_batch_value(values, f"'{base_sheet_name}'!D12"),
            "cargo": _first_batch_value(values, f"'{base_sheet_name}'!R12"),
            "asesor": _first_batch_value(values, f"'{base_sheet_name}'!D13"),
            "sede_empresa": _first_batch_value(values, f"'{base_sheet_name}'!R13"),
            "caja_compensacion": _first_batch_value(values, f"'{SHEET_FINAL}'!D11"),
            "profesional_asignado": _first_batch_value(values, f"'{SHEET_FINAL}'!Q12"),
            "nombre_vinculado": _first_batch_value(values, f"'{base_sheet_name}'!A16"),
            "cedula": _first_batch_value(values, f"'{base_sheet_name}'!E16"),
            "telefono_vinculado": _first_batch_value(values, f"'{base_sheet_name}'!I16"),
            "correo_vinculado": _first_batch_value(values, f"'{base_sheet_name}'!K16"),
            "contacto_emergencia": _first_batch_value(values, f"'{base_sheet_name}'!P16"),
            "parentesco": _first_batch_value(values, f"'{base_sheet_name}'!S16"),
            "telefono_emergencia": _first_batch_value(values, f"'{base_sheet_name}'!U16"),
            "cargo_vinculado": _first_batch_value(values, f"'{base_sheet_name}'!A18"),
            "certificado_discapacidad": _first_batch_value(values, f"'{base_sheet_name}'!E18"),
            "certificado_porcentaje": _first_batch_value(values, f"'{base_sheet_name}'!I18"),
            "discapacidad": _first_batch_value(values, f"'{base_sheet_name}'!N18"),
            "tipo_contrato": _first_batch_value(values, f"'{base_sheet_name}'!C20"),
            "fecha_inicio_contrato": _first_batch_value(values, f"'{base_sheet_name}'!M20"),
            "fecha_fin_contrato": _first_batch_value(values, f"'{base_sheet_name}'!T20"),
            "fecha_firma_contrato": _first_batch_value(values, f"'{SHEET_FINAL}'!N18"),
            "apoyos_ajustes": _first_batch_value(values, f"'{base_sheet_name}'!E21"),
            "funciones_1_5": _column_batch_values(values, f"'{base_sheet_name}'!B23:B27", 5),
            "funciones_6_10": _column_batch_values(values, f"'{base_sheet_name}'!N23:N27", 5),
            "seguimiento_fechas_1_3": _column_batch_values(values, f"'{base_sheet_name}'!C29:C31", 3),
            "seguimiento_fechas_4_6": _column_batch_values(values, f"'{base_sheet_name}'!P29:P31", 3),
        }
        return _normalize_base_payload(payload)
    wb = _load_workbook_safe(case_ref, data_only=False)
    ws = _ensure_sheet_exists(wb, _get_base_sheet_name_from_workbook(wb))
    ponderado_ws = wb[SHEET_FINAL] if SHEET_FINAL in wb.sheetnames else None
    payload = {
        "fecha_visita": _cell_value(ws, "D8"),
        "modalidad": _cell_value(ws, "R8"),
        "nombre_empresa": _cell_value(ws, "D9"),
        "ciudad_empresa": _cell_value(ws, "R9"),
        "direccion_empresa": _cell_value(ws, "D10"),
        "nit_empresa": _cell_value(ws, "R10"),
        "correo_1": _cell_value(ws, "D11"),
        "telefono_empresa": _cell_value(ws, "R11"),
        "contacto_empresa": _cell_value(ws, "D12"),
        "cargo": _cell_value(ws, "R12"),
        "asesor": _cell_value(ws, "D13"),
        "sede_empresa": _cell_value(ws, "R13"),
        "caja_compensacion": _cell_value(ponderado_ws, "D11") if ponderado_ws else "",
        "profesional_asignado": _cell_value(ponderado_ws, "Q12") if ponderado_ws else "",
        "nombre_vinculado": _cell_value(ws, "A16"),
        "cedula": _cell_value(ws, "E16"),
        "telefono_vinculado": _cell_value(ws, "I16"),
        "correo_vinculado": _cell_value(ws, "K16"),
        "contacto_emergencia": _cell_value(ws, "P16"),
        "parentesco": _cell_value(ws, "S16"),
        "telefono_emergencia": _cell_value(ws, "U16"),
        "cargo_vinculado": _cell_value(ws, "A18"),
        "certificado_discapacidad": _cell_value(ws, "E18"),
        "certificado_porcentaje": _cell_value(ws, "I18"),
        "discapacidad": _cell_value(ws, "N18"),
        "tipo_contrato": _cell_value(ws, "C20"),
        "fecha_inicio_contrato": _cell_value(ws, "M20"),
        "fecha_fin_contrato": _cell_value(ws, "T20"),
        "fecha_firma_contrato": _cell_value(ponderado_ws, "N18") if ponderado_ws else "",
        "apoyos_ajustes": _cell_value(ws, "E21"),
        "funciones_1_5": [_cell_value(ws, f"B{r}") for r in range(23, 28)],
        "funciones_6_10": [_cell_value(ws, f"N{r}") for r in range(23, 28)],
        "seguimiento_fechas_1_3": [_cell_value(ws, f"C{r}") for r in range(29, 32)],
        "seguimiento_fechas_4_6": [_cell_value(ws, f"P{r}") for r in range(29, 32)],
    }
    return payload


def save_base_payload(case_ref, payload, overwrite=True):
    if _is_native_case_ref(case_ref):
        spreadsheet_id = _get_spreadsheet_id_from_case_ref(case_ref)
        base_sheet_name = str(get_case_meta(case_ref).get("base_sheet_name") or SHEET_BASE)
        if not overwrite:
            existing = get_base_payload(case_ref)
            merged = dict(existing)
            for key, value in (payload or {}).items():
                if isinstance(value, list):
                    current_list = list(existing.get(key) or [])
                    merged_list = []
                    for idx, item in enumerate(value):
                        current_value = current_list[idx] if idx < len(current_list) else ""
                        merged_list.append(current_value if str(current_value or "").strip() else item)
                    merged[key] = merged_list
                else:
                    current_value = existing.get(key)
                    merged[key] = current_value if str(current_value or "").strip() else value
            payload = merged
        updates = _build_base_sheet_updates(payload, base_sheet_name=base_sheet_name)
        batch_write_sheet_updates(spreadsheet_id, updates)
        return
    wb = _load_workbook_safe(case_ref, data_only=False)
    ws = _ensure_sheet_exists(wb, _get_base_sheet_name_from_workbook(wb))
    mapping = {
        "fecha_visita": "D8",
        "modalidad": "R8",
        "nombre_empresa": "D9",
        "ciudad_empresa": "R9",
        "direccion_empresa": "D10",
        "nit_empresa": "R10",
        "correo_1": "D11",
        "telefono_empresa": "R11",
        "contacto_empresa": "D12",
        "cargo": "R12",
        "asesor": "D13",
        "sede_empresa": "R13",
        "nombre_vinculado": "A16",
        "cedula": "E16",
        "telefono_vinculado": "I16",
        "correo_vinculado": "K16",
        "contacto_emergencia": "P16",
        "parentesco": "S16",
        "telefono_emergencia": "U16",
        "cargo_vinculado": "A18",
        "certificado_discapacidad": "E18",
        "certificado_porcentaje": "I18",
        "discapacidad": "N18",
        "tipo_contrato": "C20",
        "fecha_inicio_contrato": "M20",
        "fecha_fin_contrato": "T20",
        "apoyos_ajustes": "E21",
    }
    for key, cell in mapping.items():
        if key in payload:
            ws[cell].value = payload.get(key, "")
    f1 = payload.get("funciones_1_5") or []
    f2 = payload.get("funciones_6_10") or []
    for i, row in enumerate(range(23, 28)):
        if i < len(f1):
            ws[f"B{row}"].value = f1[i]
        if i < len(f2):
            ws[f"N{row}"].value = f2[i]
    s1 = payload.get("seguimiento_fechas_1_3") or []
    s2 = payload.get("seguimiento_fechas_4_6") or []
    for i, row in enumerate(range(29, 32)):
        if i < len(s1):
            ws[f"C{row}"].value = s1[i]
        if i < len(s2):
            ws[f"P{row}"].value = s2[i]
    _sync_ponderado_from_payload(wb, payload)
    sanitize_logo_error_cells(wb)
    wb.save(case_ref)


def _get_followup_sheet_name(index):
    idx = int(index)
    if idx < 1 or idx > 6:
        raise ValueError("El seguimiento debe estar entre 1 y 6.")
    return f"{SHEET_PREFIX}{idx}"


def get_followup_payload(case_ref, index):
    if _is_native_case_ref(case_ref):
        sheet_name = _get_followup_sheet_name(index)
        base_payload = get_base_payload(case_ref)
        ranges = [
            f"'{sheet_name}'!E8",
            f"'{sheet_name}'!P8",
            f"'{sheet_name}'!A12:A30",
            f"'{sheet_name}'!G12:G30",
            f"'{sheet_name}'!O12:O30",
            f"'{sheet_name}'!R12:R30",
            f"'{sheet_name}'!J31",
            f"'{sheet_name}'!A34:A41",
            f"'{sheet_name}'!J34:J41",
            f"'{sheet_name}'!L34:L41",
            f"'{sheet_name}'!A43",
            f"'{sheet_name}'!A45",
            f"'{sheet_name}'!D47:D50",
            f"'{sheet_name}'!N47:N50",
        ]
        values = _batch_read_sheet_values(_get_spreadsheet_id_from_case_ref(case_ref), ranges)
        payload = {
            "modalidad": _first_batch_value(values, f"'{sheet_name}'!E8"),
            "seguimiento_numero": _first_batch_value(values, f"'{sheet_name}'!P8"),
            "item_labels": _column_batch_values(values, f"'{sheet_name}'!A12:A30", 19),
            "item_observaciones": _column_batch_values(values, f"'{sheet_name}'!G12:G30", 19),
            "item_autoevaluacion": _column_batch_values(values, f"'{sheet_name}'!O12:O30", 19),
            "item_eval_empresa": _column_batch_values(values, f"'{sheet_name}'!R12:R30", 19),
            "tipo_apoyo": _first_batch_value(values, f"'{sheet_name}'!J31"),
            "empresa_item_labels": _column_batch_values(values, f"'{sheet_name}'!A34:A41", 8),
            "empresa_eval": _column_batch_values(values, f"'{sheet_name}'!J34:J41", 8),
            "empresa_observacion": _column_batch_values(values, f"'{sheet_name}'!L34:L41", 8),
            "situacion_encontrada": _first_batch_value(values, f"'{sheet_name}'!A43"),
            "estrategias_ajustes": _first_batch_value(values, f"'{sheet_name}'!A45"),
            "asistentes": [
                {"nombre": (_column_batch_values(values, f"'{sheet_name}'!D47:D50", 4)[i]),
                 "cargo": (_column_batch_values(values, f"'{sheet_name}'!N47:N50", 4)[i])}
                for i in range(4)
            ],
        }
        if _is_template_seeded_followup_payload(payload, index):
            payload = _apply_empty_followup_fields(payload, index)
        return _normalize_empty_followup_payload(payload, index, base_payload=base_payload)
    wb = _load_workbook_safe(case_ref, data_only=False)
    ws = _ensure_sheet_exists(wb, _get_followup_sheet_name(index))
    base_payload = get_base_payload(case_ref)
    item_labels = [_cell_value(ws, f"A{r}") for r in range(12, 31)]
    empresa_labels = [_cell_value(ws, f"A{r}") for r in range(34, 42)]
    payload = {
        "modalidad": _cell_value(ws, "E8"),
        "seguimiento_numero": _cell_value(ws, "P8"),
        "item_labels": item_labels,
        "item_observaciones": [_cell_value(ws, f"G{r}") for r in range(12, 31)],
        "item_autoevaluacion": [_cell_value(ws, f"O{r}") for r in range(12, 31)],
        "item_eval_empresa": [_cell_value(ws, f"R{r}") for r in range(12, 31)],
        "tipo_apoyo": _cell_value(ws, "J31"),
        "empresa_item_labels": empresa_labels,
        "empresa_eval": [_cell_value(ws, f"J{r}") for r in range(34, 42)],
        "empresa_observacion": [_cell_value(ws, f"L{r}") for r in range(34, 42)],
        "situacion_encontrada": _cell_value(ws, "A43"),
        "estrategias_ajustes": _cell_value(ws, "A45"),
        "asistentes": [
            {"nombre": _cell_value(ws, f"D{r}"), "cargo": _cell_value(ws, f"N{r}")}
            for r in range(47, 51)
        ],
    }
    if _is_template_seeded_followup_payload(payload, index):
        payload = _apply_empty_followup_fields(payload, index)
    return _normalize_empty_followup_payload(payload, index, base_payload=base_payload)


def save_followup_payload(case_ref, index, payload):
    if _is_native_case_ref(case_ref):
        updates = _build_followup_sheet_updates(index, payload)
        batch_write_sheet_updates(_get_spreadsheet_id_from_case_ref(case_ref), updates)
        return
    wb = _load_workbook_safe(case_ref, data_only=False)
    ws = _ensure_sheet_exists(wb, _get_followup_sheet_name(index))
    ws["E8"].value = payload.get("modalidad", "")
    ws["P8"].value = payload.get("seguimiento_numero", index)
    item_obs = payload.get("item_observaciones") or []
    item_auto = payload.get("item_autoevaluacion") or []
    item_emp = payload.get("item_eval_empresa") or []
    for i, row in enumerate(range(12, 31)):
        if i < len(item_obs):
            ws[f"G{row}"].value = item_obs[i]
        if i < len(item_auto):
            ws[f"O{row}"].value = item_auto[i]
        if i < len(item_emp):
            ws[f"R{row}"].value = item_emp[i]
    ws["J31"].value = payload.get("tipo_apoyo", "")
    emp_eval = payload.get("empresa_eval") or []
    emp_obs = payload.get("empresa_observacion") or []
    for i, row in enumerate(range(34, 42)):
        if i < len(emp_eval):
            ws[f"J{row}"].value = emp_eval[i]
        if i < len(emp_obs):
            ws[f"L{row}"].value = emp_obs[i]
    ws["A43"].value = payload.get("situacion_encontrada", "")
    ws["A45"].value = payload.get("estrategias_ajustes", "")
    asistentes = payload.get("asistentes") or []
    for i, row in enumerate(range(47, 51)):
        entry = asistentes[i] if i < len(asistentes) else {}
        ws[f"D{row}"].value = entry.get("nombre", "")
        ws[f"N{row}"].value = entry.get("cargo", "")
    sanitize_logo_error_cells(wb)
    wb.save(case_ref)
