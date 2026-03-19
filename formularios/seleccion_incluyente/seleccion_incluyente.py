import json
import os
import shutil
import time
from difflib import SequenceMatcher
from functools import lru_cache

from formularios.evaluacion_programa import evaluacion_accesibilidad
from formularios.common import (
    _get_desktop_dir,
    _next_available_file_path,
    _normalize_cedula,
    _normalize_decimal_value,
    _normalize_text,
    _parse_date_value,
    _coerce_excel_decimal_value,
    sanitize_logo_error_cells,
    autofit_rows,
    clear_written_rows,
    ws_write,
    _sanitize_filename,
    _supabase_get,
    _supabase_upsert_with_queue,
)
from logging_utils import log_excel_event

FORM_ID = "seleccion_incluyente"
FORM_NAME = "Proceso de Seleccion Incluyente"

SECTION_1 = {
    "title": "1. DATOS DE LA EMPRESA",
    "fields": [
        {"id": "fecha_visita", "label": "Fecha de la visita", "source": "input"},
        {
            "id": "modalidad",
            "label": "Modalidad",
            "source": "input",
            "options": ["Presencial", "Virtual", "Mixta", "No aplica"],
        },
        {
            "id": "nombre_empresa",
            "label": "Nombre de la empresa",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "ciudad_empresa",
            "label": "Ciudad/Municipio",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "direccion_empresa",
            "label": "Dirección de la empresa",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "nit_empresa",
            "label": "Número de NIT",
            "source": "input",
        },
        {
            "id": "correo_1",
            "label": "Correo electrónico",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "telefono_empresa",
            "label": "Teléfonos",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "contacto_empresa",
            "label": "Persona que atiende la visita",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "cargo",
            "label": "Cargo",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "caja_compensacion",
            "label": "Empresa afiliada a Caja de Compensación",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "sede_empresa",
            "label": "Sede Compensar",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "asesor",
            "label": "Asesor",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
        {
            "id": "profesional_asignado",
            "label": "Profesional asignado RECA",
            "source": "supabase",
            "table": "empresas",
            "readonly": True,
        },
    ],
}

SECTION_1_SUPABASE_MAP = evaluacion_accesibilidad.SECTION_1_SUPABASE_MAP.copy()

FORM_CACHE = {}
SECTION_1_CACHE = {}

SHEET_NAME = "4. PROCESO DE SELECCION INCLUYE"
SECTION_2_ANCHOR = "2. DATOS DEL OFERENTE"
SECTION_5_ANCHOR = "5. AJUSTES RAZONABLES / RECOMENDACIONES AL PROCESO DE SELECCION"
SECTION_2_TEMPLATE_ANCHOR_ROW = 14
SECTION_2_LAST_COLUMN = "U"
TEMPLATE_VARIANT_INDIVIDUAL = "individual"
TEMPLATE_VARIANT_GROUP_2_PLUS = "group_2_plus"
SECTION_2_GROUP_BLOCK_HEIGHT = 61
SECTION_2_GROUP_SHARED_ACTIVITY_CELL = "A14"
SECTION_5_GROUP_ANCHOR = "5. AJUSTES RAZONABLES"
SECTION_2_GROUP_FIRST_BLOCK_START_ROW = 16
SECTION_2_GROUP_SECOND_BLOCK_START_ROW = 77
TEMPLATE_FILENAME_BY_VARIANT = {
    TEMPLATE_VARIANT_INDIVIDUAL: "seleccion_incluyente.xlsx",
    TEMPLATE_VARIANT_GROUP_2_PLUS: "seleccion_incluyente_grupal_2_4.xlsx",
}

AJUSTES_ENTREVISTA_TEMPLATES = {
    "preparacion_proceso": """
Ajustes razonables para entrevista.

Contactar a la Agencia de Empleo para el apoyo de adaptación y organización de las pruebas psicotécnicas de la empresa.

Realizar una organización de la entrevista de forma que se pueda anticipar a los oferentes el paso a paso de lo que se realizará en el proceso, permitiendo que logren organizar su tiempo y la constancia que requiere el mismo.

Promover la aplicación de formatos para la hoja de vida con diseños sencillos, fáciles de diligenciar y a través de medios accesibles; estos pueden ser virtuales o físicos.

Para mantener un contacto directo con los oferentes que han sido preseleccionados, se recomienda evaluar alternativas hasta identificar el mejor canal de comunicación: contacto telefónico, mensajes de texto, mensajes por chat, correo electrónico y, en algunos casos, a través de familiares que hayan sido referenciados en la hoja de vida. Los mensajes deben ser precisos y concretos, incluyendo solamente la información básica y utilizando frases simples.
""".strip(),
    "trato_respetuoso": """
Evitar conductas, palabras, frases, sentimientos, preconcepciones y estigmas que impidan u obstaculicen el acceso en igualdad de condiciones de las personas con y/o en situación de discapacidad a los espacios, objetos, servicios y, en general, a las posibilidades que ofrece la sociedad.

Evitar, durante el proceso de entrevista, preguntar sobre cómo fue la adquisición de la discapacidad.

Informar a los candidatos que no fueron seleccionados sobre la finalización del proceso, basándose en sus habilidades residuales individuales. De esta manera no tendrán que esperar innecesariamente y se respeta su tiempo.

La evaluación de desempeño debe ajustarse de acuerdo con el perfil del cargo y no enfocarse en la discapacidad de la persona.

Mantener contacto visual con la persona, mirando a sus ojos y evitando enfocar su discapacidad.
""".strip(),
    "accesibilidad_entrevista": """
Asegurarse de que el lugar de la entrevista sea accesible para personas con discapacidad física. Esto incluye la disponibilidad de rampas, ascensores o espacios adecuados para sillas de ruedas, si es necesario.

Si el candidato tiene dificultades de comunicación, ofrecer alternativas como permitir el uso de comunicación por texto, proporcionar un intérprete de lengua de señas si es necesario o permitir que el candidato responda por escrito.

Considerar la posibilidad de ofrecer tiempo adicional para completar la entrevista si la discapacidad del candidato afecta su velocidad de procesamiento o comunicación.

Ser flexible en cuanto al formato de la entrevista. Algunas personas pueden necesitar entrevistas en un formato diferente, como entrevistas virtuales o en un entorno menos estimulante para quienes presentan sensibilidad sensorial.

Formular preguntas claras y directas y ser paciente al esperar la respuesta del candidato. Evitar jergas o frases complicadas que puedan ser difíciles de entender.

Al proporcionar retroalimentación al candidato, ser claro, conciso y constructivo. Destacar sus fortalezas y ofrecer sugerencias de mejora, si es necesario, de manera útil y respetuosa.

Realizar preguntas orientadoras que permitan identificar que la información está siendo recibida correctamente por el oferente durante el proceso de selección.
""".strip(),
    "accesibilidad_documentos": """
Aumentar el tamaño de letra, utilizar colores de fondo en el texto y emplear fuentes de fácil lectura como Arial o Verdana. Adicionalmente, proporcionar esos documentos de manera virtual para facilitar el uso de herramientas tiflotecnológicas.
""".strip(),
    "pruebas_seleccion": """
La aplicación de pruebas de tipo visual y gráfico, como por ejemplo Test de Percepción Temática, Wartegg, Técnica de dibujo proyectivo HTP por sus siglas en inglés (Casa, Árbol, Persona) y Test de la Figura Humana, no es recomendable para personas con discapacidad visual.

Para los procesos de selección en los que participen candidatos usuarios de lengua de señas, es clave contar con un servicio de interpretación profesional y no apoyarse en amigos o familiares de los candidatos que tengan conocimiento.

No se recomienda hacer interpretación en lengua de señas de las pruebas psicotécnicas, dado que se puede sesgar la información que se espera recoger y, por ende, los resultados. Es preferible reemplazar este tipo de pruebas por entrevistas por competencias.

En el caso de personas con discapacidad que cuentan con familias sobreprotectoras, se deben establecer límites claros con estas, restringiendo su participación durante el proceso de selección.

Se recomienda minimizar el uso de pruebas psicotécnicas y reemplazarlas por otras estrategias de selección que permitan alcanzar los mismos objetivos. Solo en caso de que lo anterior no sea posible, se sugiere priorizar la aplicación de pruebas gráficas proyectivas que buscan identificar rasgos de personalidad.
""".strip(),
}

AJUSTES_ENTREVISTA_TEMPLATE_BUTTONS = [
    ("preparacion_proceso", "Preparación del proceso"),
    ("trato_respetuoso", "Trato respetuoso"),
    ("accesibilidad_entrevista", "Accesibilidad entrevista"),
    ("accesibilidad_documentos", "Accesibilidad documentos"),
    ("pruebas_seleccion", "Pruebas de selección"),
]

SECTION_2 = {
    "title": "2. DATOS DEL OFERENTE",
    "fields": [
        {"id": "numero", "label": "No", "type": "texto"},
        {"id": "nombre_oferente", "label": "Nombre oferente", "type": "texto"},
        {"id": "cedula", "label": "Cédula", "type": "texto"},
        {"id": "certificado_porcentaje", "label": "Certificado %", "type": "texto"},
        {
            "id": "discapacidad",
            "label": "Discapacidad",
            "type": "lista",
            "options": [
                "Discapacidad visual pérdida total de la visión",
                "Discapacidad visual baja visión",
                "Discapacidad auditiva",
                "Discapacidad auditiva hipoacusia",
                "Trastorno de espectro autista",
                "Discapacidad intelectual",
                "Discapacidad física",
                "Discapacidad física usuario en silla de ruedas",
                "Discapacidad psicosocial",
                "Discapacidad múltiple",
                "No aplica",
            ],
        },
        {"id": "telefono_oferente", "label": "Teléfono oferente", "type": "texto"},
        {
            "id": "resultado_certificado",
            "label": "Resultado certificado",
            "type": "lista",
            "options": ["Aprobado", "No aprobado", "Pendiente"],
        },
        {"id": "cargo_oferente", "label": "Cargo oferente", "type": "texto"},
        {
            "id": "nombre_contacto_emergencia",
            "label": "Nombre contacto emergencia",
            "type": "texto",
        },
        {"id": "parentesco", "label": "Parentesco", "type": "texto"},
        {"id": "telefono_emergencia", "label": "Teléfono", "type": "texto"},
        {"id": "fecha_nacimiento", "label": "Fecha de nacimiento", "type": "texto"},
        {"id": "edad", "label": "Edad", "type": "texto"},
        {
            "id": "pendiente_otros_oferentes",
            "label": "Pendiente otros oferentes",
            "type": "lista",
            "options": ["Si", "No", "Por Confirmar"],
        },
        {
            "id": "lugar_firma_contrato",
            "label": "Lugar firma de contrato",
            "type": "texto",
        },
        {
            "id": "fecha_firma_contrato",
            "label": "Fecha firma de contrato",
            "type": "texto",
        },
        {
            "id": "cuenta_pension",
            "label": "Cuenta con pension",
            "type": "lista",
            "options": ["Si", "No", "Por Confirmar"],
        },
        {
            "id": "tipo_pension",
            "label": "Tipo de pension",
            "type": "lista",
            "options": [
                "Pension Invalidez",
                "Subsidiada",
                "Especial de vejez",
                "Victimas conflicto",
                "Familiar",
                "Regimen especial",
                "No aplica",
            ],
        },
        {
            "id": "desarrollo_actividad",
            "label": "Desarrollo de la actividad",
            "type": "texto_largo",
        },
        {
            "id": "medicamentos_nivel_apoyo",
            "label": "Toma medicamentos - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "medicamentos_conocimiento",
            "label": "Toma medicamentos - Conocimiento de medicamentos",
            "type": "lista",
            "options": [
                "1. Conoce los medicamentos que consume.",
                "2. Un tercero es quien conoce los medicamentos que consume.",
                "3. No conoce los medicamentos que consume.",
                "No aplica.",
                "0. No requiere apoyo.",
            ],
        },
        {
            "id": "medicamentos_horarios",
            "label": "Toma medicamentos - Conocimiento de horarios",
            "type": "lista",
            "options": [
                "1. Conoce los horarios de toma de medicamentos que consume.",
                "2. Es un tercero quien conoce los horarios de la toma de medicamentos.",
                "3. No conoce los horarios de toma de medicamentos que consume.",
                "0. No requiere apoyo.",
                "No aplica.",
            ],
        },
        {
            "id": "medicamentos_nota",
            "label": "Toma medicamentos - Nota",
            "type": "texto",
        },
        {
            "id": "alergias_nivel_apoyo",
            "label": "Presenta alergia - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "alergias_tipo",
            "label": "Presenta alergia - Tipo de alergia",
            "type": "lista",
            "options": [
                "0. No presenta alergias.",
                "1. Presenta alergias y sabe darle manejo.",
                "2. No conoce si presenta alguna alergia.",
                "3. Presenta alergias a: medicamentos, sustancias y productos quimicos, alimentos, animales, entre otros.",
                "No aplica.",
            ],
        },
        {
            "id": "alergias_nota",
            "label": "Presenta alergia - Nota",
            "type": "texto",
        },
        {
            "id": "restriccion_nivel_apoyo",
            "label": "Tiene restriccion medica - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "restriccion_conocimiento",
            "label": "Tiene restriccion medica - Conocimiento",
            "type": "lista",
            "options": [
                "0. No tiene restricciones medicas.",
                "1. Tiene restricciones medicas y conoce su manejo.",
                "2. No conoce si tiene restricciones medicas.",
                "3. Si tiene restricciones medicas y desconoce su manejo.",
                "No aplica.",
            ],
        },
        {
            "id": "restriccion_nota",
            "label": "Tiene restriccion medica - Nota",
            "type": "texto",
        },
        {
            "id": "controles_nivel_apoyo",
            "label": "Asiste a controles medicos - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "controles_asistencia",
            "label": "Asiste a controles medicos - Asistencia a controles",
            "type": "lista",
            "options": [
                "No aplica.",
                "2. Si asiste a controles medicos con especialista.",
                "3. No sabe si tiene controles medicos con especialista.",
                "1. Asiste a controles medicos con especialista y conoce el manejo.",
                "0. No requiere apoyo.",
            ],
        },
        {
            "id": "controles_frecuencia",
            "label": "Asiste a controles medicos - Frecuencia",
            "type": "lista",
            "options": [
                "Mensual",
                "Trimestral",
                "Semestral",
                "Otra frecuencia",
                "No aplica",
            ],
        },
        {
            "id": "controles_nota",
            "label": "Asiste a controles medicos - Nota",
            "type": "texto",
        },
        {
            "id": "desplazamiento_nivel_apoyo",
            "label": "Desplazamiento independiente - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "desplazamiento_modo",
            "label": "Desplazamiento independiente - Modo de desplazamiento",
            "type": "lista",
            "options": [
                "0. Se desplaza de manera independiente sin necesidad de apoyos (ortesis, baston, silla de ruedas entre otros).",
                "1. Se desplaza de forma independiente con un apoyo temporal (ortesis, baston, silla de ruedas entre otros).",
                "2. Se desplaza de manera independiente con un apoyo permanente (ortesis, baston, silla de ruedas entre otros).",
                "3. No se desplaza de manera independiente. Requiere el acompanamiento de un tercero y un apoyo (ortesis, baston, silla de ruedas entre otros).",
                "No aplica.",
            ],
        },
        {
            "id": "desplazamiento_transporte",
            "label": "Desplazamiento independiente - Medio de transporte",
            "type": "lista",
            "options": [
                "Caminando.",
                "Bicicleta.",
                "Transmilenio, Sitp.",
                "Vehiculo propio.",
                "Vehiculo especial.",
                "No aplica.",
            ],
        },
        {
            "id": "desplazamiento_nota",
            "label": "Desplazamiento independiente - Nota",
            "type": "texto",
        },
        {
            "id": "ubicacion_nivel_apoyo",
            "label": "Ubicacion en la ciudad - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "ubicacion_ciudad",
            "label": "Ubicacion en la ciudad",
            "type": "lista",
            "options": [
                "0. Sabe ubicarse en la ciudad de manera autonoma.",
                "1. Sabe ubicarse en la ciudad pero haciendo uso de aplicaciones (Maps, Waze, entre otros).",
                "2. Requiere de acompanamiento para ubicarse.",
                "3. No sabe ubicarse en la ciudad.",
            ],
        },
        {
            "id": "ubicacion_aplicaciones",
            "label": "Manejo de aplicaciones",
            "type": "lista",
            "options": [
                "Se ubica por puntos de referencia y direcciones.",
                "No se ubica por puntos de referencia.",
                "Se ubica por puntos cardinales.",
                "No aplica",
            ],
        },
        {
            "id": "ubicacion_nota",
            "label": "Ubicacion en la ciudad - Nota",
            "type": "texto",
        },
        {
            "id": "dinero_nivel_apoyo",
            "label": "Reconoce y maneja el dinero - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "dinero_reconocimiento",
            "label": "Reconocimiento del dinero",
            "type": "lista",
            "options": ["Autonomo.", "Con apoyo familiar."],
        },
        {
            "id": "dinero_manejo",
            "label": "Manejo del dinero",
            "type": "lista",
            "options": [
                "0. Reconoce y maneja el dinero de manera autonoma.",
                "1. Reconoce y maneja el dinero pero en ocasiones requiere apoyo.",
                "2. Solo reconoce el dinero pero no lo sabe manejar.",
                "3. No reconoce el dinero y no lo sabe manejar.",
                "No aplica.",
            ],
        },
        {
            "id": "dinero_medios",
            "label": "Uso de medios electronicos",
            "type": "lista",
            "options": [
                "Dinero fisico, plastico y digital.",
                "Dinero fisico y plastico.",
                "Dinero fisico.",
                "Dinero plastico y digital.",
                "Dinero plastico.",
                "Dinero digital.",
                "Dinero digital y fisico.",
            ],
        },
        {
            "id": "dinero_nota",
            "label": "Reconoce y maneja el dinero - Nota",
            "type": "texto",
        },
        {
            "id": "presentacion_nivel_apoyo",
            "label": "Presentacion personal - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "presentacion_personal",
            "label": "Presentacion personal",
            "type": "lista",
            "options": [
                "0. Su codigo de vestuario es acorde al contexto.",
                "1. Su codigo de vestuario es acorde al contexto, pero presenta oportunidades de mejora.",
                "2. Su codigo de vestuario es medianamente acorde al contexto.",
                "3. Su codigo de vestuario no es acorde al contexto.",
                "No aplica.",
            ],
        },
        {
            "id": "presentacion_nota",
            "label": "Presentacion personal - Nota",
            "type": "texto",
        },
        {
            "id": "comunicacion_escrita_nivel_apoyo",
            "label": "Apoyo comunicacion escrita - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "comunicacion_escrita_apoyo",
            "label": "Apoyo comunicacion escrita",
            "type": "lista",
            "options": [
                "0. Si conoce y maneja los apoyos (Jaws, Magic, el lector de pantalla de Windows/IOS).",
                "1. Maneja algunos apoyos de comunicacion escrita, pero no todos en general.",
                "2. Conoce pero no maneja apoyos.",
                "3. No conoce, ni maneja los apoyos.",
                "No aplica.",
            ],
        },
        {
            "id": "comunicacion_escrita_nota",
            "label": "Apoyo comunicacion escrita - Nota",
            "type": "texto",
        },
        {
            "id": "comunicacion_verbal_nivel_apoyo",
            "label": "Apoyo comunicacion verbal - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "comunicacion_verbal_apoyo",
            "label": "Apoyo comunicacion verbal",
            "type": "lista",
            "options": [
                "0. Si conoce y maneja los apoyos (Centro de relevo, entre otros).",
                "1. Maneja algunos apoyos, pero no los conoce todos en general (Centro de relevo, entre otros).",
                "2. Conoce pero no maneja apoyos.",
                "3. No conoce, ni maneja los apoyos.",
                "No aplica.",
            ],
        },
        {
            "id": "comunicacion_verbal_nota",
            "label": "Apoyo comunicacion verbal - Nota",
            "type": "texto",
        },
        {
            "id": "decisiones_nivel_apoyo",
            "label": "Toma de decisiones - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "toma_decisiones",
            "label": "Toma de decisiones",
            "type": "lista",
            "options": [
                "0. Toma las decisiones de manera autonoma.",
                "1. Toma decisiones pero en ocasiones requiere el apoyo de un tercero.",
                "2. Debe consultar con un tercero para la toma de decisiones.",
                "3. Requiere el apoyo de un tercero para tomar decisiones.",
                "No aplica.",
            ],
        },
        {
            "id": "toma_decisiones_nota",
            "label": "Toma de decisiones - Nota",
            "type": "texto",
        },
        {
            "id": "aseo_nivel_apoyo",
            "label": "Apoyo en aseo personal - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "alimentacion",
            "label": "Alimentacion",
            "type": "lista",
            "options": [
                "0. No requiere apoyo en sus actividades de la vida diaria.",
                "1. Requiere apoyo en algunas actividades de la vida diaria.",
                "2. Requiere apoyo en la mayoria de actividades de la vida diaria.",
                "3. Requiere apoyo en todas las actividades de la vida diaria.",
                "No aplica.",
            ],
        },
        {
            "id": "aseo_criar_apoyo",
            "label": "Criar y cuidado de ninos - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "aseo_comunicacion_apoyo",
            "label": "Uso de sistemas de comunicacion - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "aseo_ayudas_apoyo",
            "label": "Cuidado de ayudas tecnicas - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "aseo_alimentacion",
            "label": "Alimentacion",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "aseo_movilidad_funcional",
            "label": "Movilidad funcional",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "aseo_higiene_aseo",
            "label": "Higiene personal y aseo (Control de esfinter)",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "aseo_nota",
            "label": "Apoyo en aseo personal - Nota",
            "type": "texto",
        },
        {
            "id": "instrumentales_nivel_apoyo",
            "label": "Apoyo en actividades instrumentales - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "instrumentales_actividades",
            "label": "Actividades instrumentales",
            "type": "lista",
            "options": [
                "0. No requiere apoyo en actividades instrumentales de la vida diaria.",
                "1. Requiere apoyo en algunas actividades instrumentales de la vida diaria.",
                "2. Requiere apoyo en la mayoria de actividades instrumentales de la vida diaria.",
                "3. Requiere apoyo en todas las actividades instrumentales de la vida diaria.",
                "No aplica.",
            ],
        },
        {
            "id": "instrumentales_criar_apoyo",
            "label": "Criar y cuidado de ninos - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "instrumentales_comunicacion_apoyo",
            "label": "Uso de sistemas de comunicacion - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "instrumentales_movilidad_apoyo",
            "label": "Movilidad en la comunidad - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "instrumentales_finanzas",
            "label": "Manejo de tematicas financieras",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "instrumentales_cocina_limpieza",
            "label": "Cocina y limpieza",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "instrumentales_crear_hogar",
            "label": "Crear y mantener un hogar",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "instrumentales_salud_cuenta_apoyo",
            "label": "Cuidado de salud y manutencion - Cuenta con apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "instrumentales_nota",
            "label": "Apoyo en actividades instrumentales - Nota",
            "type": "texto",
        },
        {
            "id": "actividades_nivel_apoyo",
            "label": "Apoyo durante actividades - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "actividades_apoyo",
            "label": "Apoyo durante actividades",
            "type": "lista",
            "options": [
                "0. No requiere apoyo en sus actividades laborales.",
                "1. Requiere apoyo en algunas actividades laborales.",
                "2. Requiere apoyo en la mayoria de actividades laborales.",
                "3. Requiere apoyo en todas las actividades laborales.",
                "No aplica",
            ],
        },
        {
            "id": "actividades_esparcimiento_apoyo",
            "label": "Actividades de esparcimiento con familia - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "actividades_esparcimiento_cuenta_apoyo",
            "label": "Actividades de esparcimiento con familia - Cuenta con apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "actividades_complementarios_apoyo",
            "label": "Complementarios medicos - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "actividades_complementarios_cuenta_apoyo",
            "label": "Complementarios medicos - Cuenta con apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "actividades_subsidios_cuenta_apoyo",
            "label": "Subsidios economicos para estudio de hijos - Cuenta con apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "actividades_nota",
            "label": "Apoyo durante actividades - Nota",
            "type": "texto",
        },
        {
            "id": "discriminacion_nivel_apoyo",
            "label": "Discriminacion - Nivel de apoyo",
            "type": "lista",
            "options": [
                "0. No requiere apoyo.",
                "1. Nivel de apoyo Bajo.",
                "2. Nivel de apoyo medio.",
                "3. Nivel de apoyo alto.",
                "No aplica.",
            ],
        },
        {
            "id": "discriminacion",
            "label": "Discriminacion",
            "type": "lista",
            "options": [
                "0. No ha sufrido de discriminacion.",
                "1. Ha sufrido de discriminacion en algunos contextos.",
                "2. Ha sufrido de discriminacion en repetidas ocasiones.",
                "3. Ha sufrido de discriminacion a los largo del ciclo vital.",
                "No aplica.",
            ],
        },
        {
            "id": "discriminacion_violencia_apoyo",
            "label": "Violencia fisica - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "discriminacion_violencia_cuenta_apoyo",
            "label": "Violencia fisica - Cuenta con apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "discriminacion_vulneracion_apoyo",
            "label": "Vulneracion de derechos - Requiere apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "discriminacion_vulneracion_cuenta_apoyo",
            "label": "Vulneracion de derechos - Cuenta con apoyo",
            "type": "lista",
            "options": ["Si", "No", "No aplica"],
        },
        {
            "id": "discriminacion_nota",
            "label": "Discriminacion - Nota",
            "type": "texto",
        },
    ],
}

SECTION_5 = {
    "title": "5. AJUSTES RAZONABLES / RECOMENDACIONES AL PROCESO DE SELECCION",
    "fields": [
        {"id": "ajustes_recomendaciones", "label": "Ajustes razonables", "type": "texto_largo"},
        {"id": "nota", "label": "Nota", "type": "texto"},
    ],
}

SECTION_6 = {
    "title": "6. ASISTENTES",
    "rows": 4,
}

SECTION_1_FIELD_MAP = {field["id"]: field for field in SECTION_1["fields"]}
SECTION_2_FIELD_MAP = {field["id"]: field for field in SECTION_2["fields"]}
LIST_FIELD_OPTIONS_BY_ID = {
    field["id"]: list(field.get("options", []))
    for field in SECTION_1["fields"] + SECTION_2["fields"]
    if field.get("type") == "lista"
}

SECTION_2_CELL_MAP = {
    "numero": ("A", 17),
    "nombre_oferente": ("C", 17),
    "cedula": ("H", 17),
    "certificado_porcentaje": ("K", 17),
    "discapacidad": ("L", 17),
    "telefono_oferente": ("O", 17),
    "resultado_certificado": ("R", 17),
    "cargo_oferente": ("A", 19),
    "nombre_contacto_emergencia": ("F", 19),
    "parentesco": ("I", 19),
    "telefono_emergencia": ("K", 19),
    "fecha_nacimiento": ("N", 19),
    "edad": ("S", 19),
    "pendiente_otros_oferentes": ("G", 20),
    "lugar_firma_contrato": ("L", 20),
    "fecha_firma_contrato": ("R", 20),
    "cuenta_pension": ("I", 21),
    "tipo_pension": ("Q", 21),
    "desarrollo_actividad": ("A", 23),
    "medicamentos_nivel_apoyo": ("I", 27),
    "medicamentos_conocimiento": ("N", 27),
    "medicamentos_horarios": ("N", 28),
    "medicamentos_nota": ("O", 29),
    "alergias_nivel_apoyo": ("I", 30),
    "alergias_tipo": ("N", 30),
    "alergias_nota": ("O", 31),
    "restriccion_nivel_apoyo": ("I", 32),
    "restriccion_conocimiento": ("N", 32),
    "restriccion_nota": ("O", 33),
    "controles_nivel_apoyo": ("I", 34),
    "controles_asistencia": ("N", 34),
    "controles_frecuencia": ("N", 35),
    "controles_nota": ("O", 36),
    "desplazamiento_nivel_apoyo": ("I", 40),
    "desplazamiento_modo": ("N", 40),
    "desplazamiento_transporte": ("N", 41),
    "desplazamiento_nota": ("O", 42),
    "ubicacion_nivel_apoyo": ("I", 43),
    "ubicacion_ciudad": ("N", 43),
    "ubicacion_aplicaciones": ("N", 44),
    "ubicacion_nota": ("O", 45),
    "dinero_nivel_apoyo": ("I", 46),
    "dinero_reconocimiento": ("N", 46),
    "dinero_manejo": ("N", 47),
    "dinero_medios": ("N", 48),
    "dinero_nota": ("O", 49),
    "presentacion_nivel_apoyo": ("I", 50),
    "presentacion_personal": ("N", 50),
    "presentacion_nota": ("O", 51),
    "comunicacion_escrita_nivel_apoyo": ("I", 52),
    "comunicacion_escrita_apoyo": ("N", 52),
    "comunicacion_escrita_nota": ("N", 53),
    "comunicacion_verbal_nivel_apoyo": ("I", 54),
    "comunicacion_verbal_apoyo": ("N", 54),
    "comunicacion_verbal_nota": ("O", 55),
    "decisiones_nivel_apoyo": ("I", 56),
    "toma_decisiones": ("N", 56),
    "toma_decisiones_nota": ("O", 57),
    "aseo_nivel_apoyo": ("I", 58),
    "alimentacion": ("N", 58),
    "aseo_criar_apoyo": ("Q", 59),
    "aseo_comunicacion_apoyo": ("Q", 60),
    "aseo_ayudas_apoyo": ("Q", 61),
    "aseo_alimentacion": ("U", 59),
    "aseo_movilidad_funcional": ("U", 60),
    "aseo_higiene_aseo": ("U", 61),
    "aseo_nota": ("O", 62),
    "instrumentales_nivel_apoyo": ("I", 63),
    "instrumentales_actividades": ("N", 63),
    "instrumentales_criar_apoyo": ("Q", 64),
    "instrumentales_finanzas": ("U", 64),
    "instrumentales_comunicacion_apoyo": ("Q", 65),
    "instrumentales_cocina_limpieza": ("U", 65),
    "instrumentales_movilidad_apoyo": ("Q", 66),
    "instrumentales_crear_hogar": ("U", 66),
    "instrumentales_salud_cuenta_apoyo": ("U", 67),
    "instrumentales_nota": ("O", 68),
    "actividades_nivel_apoyo": ("I", 69),
    "actividades_apoyo": ("N", 69),
    "actividades_esparcimiento_apoyo": ("Q", 70),
    "actividades_esparcimiento_cuenta_apoyo": ("U", 70),
    "actividades_complementarios_apoyo": ("Q", 71),
    "actividades_complementarios_cuenta_apoyo": ("U", 71),
    "actividades_subsidios_cuenta_apoyo": ("U", 72),
    "actividades_nota": ("O", 73),
    "discriminacion_nivel_apoyo": ("I", 74),
    "discriminacion": ("N", 74),
    "discriminacion_violencia_apoyo": ("Q", 75),
    "discriminacion_violencia_cuenta_apoyo": ("U", 75),
    "discriminacion_vulneracion_apoyo": ("Q", 76),
    "discriminacion_vulneracion_cuenta_apoyo": ("U", 76),
    "discriminacion_nota": ("O", 77),
}

_DISCAPACIDAD_CATEGORIA_MAP = {
    "discapacidad visual perdida total de la vision": "Visual",
    "discapacidad visual baja vision": "Visual",
    "discapacidad auditiva": "Auditiva",
    "discapacidad auditiva hipoacusia": "Auditiva",
    "trastorno de espectro autista": "Intelectual",
    "discapacidad intelectual": "Intelectual",
    "discapacidad fisica": "Física",
    "discapacidad fisica usuario en silla de ruedas": "Física",
    "discapacidad psicosocial": "Psicosocial",
    "discapacidad multiple": "Múltiple",
    "no aplica": None,
}

EXCEL_MAPPING = {
    "section_1": {
        "fecha_visita": "F7",
        "modalidad": "N7",
        "nombre_empresa": "F8",
        "ciudad_empresa": "N8",
        "direccion_empresa": "F9",
        "nit_empresa": "N9",
        "correo_1": "F10",
        "telefono_empresa": "N10",
        "contacto_empresa": "F11",
        "cargo": "N11",
        "caja_compensacion": "F12",
        "sede_empresa": "N12",
        "asesor": "F13",
        "profesional_asignado": "N13",
    },
    "section_6": {
        "start_row": 85,
        "rows": 4,
        "nombre_col": "E",
        "cargo_col": "M",
    },
}

SECTION_1_CELL_MAP_BY_TEMPLATE = {
    TEMPLATE_VARIANT_INDIVIDUAL: EXCEL_MAPPING["section_1"],
    TEMPLATE_VARIANT_GROUP_2_PLUS: {
        "fecha_visita": "F7",
        "modalidad": "N7",
        "nombre_empresa": "F8",
        "ciudad_empresa": "N8",
        "direccion_empresa": "F9",
        "nit_empresa": "N9",
        "correo_1": "F10",
        "telefono_empresa": "N10",
        "contacto_empresa": "F11",
        "cargo": "N11",
        "asesor": "F12",
        "sede_empresa": "N12",
    },
}

SECTION_2_INDIVIDUAL_CELL_MAP = dict(SECTION_2_CELL_MAP)

SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP = dict(SECTION_2_CELL_MAP)
SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP.update(
    {
        "numero": ("A", 19),
        "nombre_oferente": ("C", 19),
        "cedula": ("H", 19),
        "certificado_porcentaje": ("K", 19),
        "discapacidad": ("L", 19),
        "telefono_oferente": ("O", 19),
        "resultado_certificado": ("R", 19),
        "cargo_oferente": ("A", 21),
        "nombre_contacto_emergencia": ("F", 21),
        "parentesco": ("I", 21),
        "telefono_emergencia": ("K", 21),
        "fecha_nacimiento": ("N", 21),
        "edad": ("S", 21),
        "pendiente_otros_oferentes": ("G", 22),
        "lugar_firma_contrato": ("L", 22),
        "fecha_firma_contrato": ("R", 22),
        "cuenta_pension": ("I", 23),
        "tipo_pension": ("Q", 23),
    }
)
SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP.pop("desarrollo_actividad", None)
for _field_id, (_col, _row) in list(SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP.items()):
    if _row >= 40:
        SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP[_field_id] = (_col, _row - 1)

SECTION_6_BASE_ROWS_BY_TEMPLATE = {
    TEMPLATE_VARIANT_INDIVIDUAL: 4,
    TEMPLATE_VARIANT_GROUP_2_PLUS: 2,
}

EXCEL_DROPDOWN_MANUAL_CANONICAL_OPTIONS = {
    "tipo_pension": [
        "Pensión Invalidez",
        "Pensión Subsidiada",
        "Pensión especial de vejez anticipada",
        "Pensión para víctimas de conflicto armado",
        "Pensión familiar",
        "Pensión régimen especial (fuerzas militares)",
        "No aplica",
    ],
    # El dropdown del template tiene una comilla residual en una opción;
    # preservamos el texto limpio de la UI para no exportar ese artefacto.
    "discapacidad": list(LIST_FIELD_OPTIONS_BY_ID.get("discapacidad", [])),
}


def register_form():
    return {
        "id": FORM_ID,
        "name": FORM_NAME,
    }


def _get_cache_dir():
    base = os.getenv("LOCALAPPDATA")
    if not base:
        userprofile = os.getenv("USERPROFILE")
        if userprofile:
            base = os.path.join(userprofile, "AppData", "Local")
    if not base:
        base = os.getcwd()
    cache_dir = os.path.join(base, "RECA", "cache")
    os.makedirs(cache_dir, exist_ok=True)
    return cache_dir




def _infer_discapacidad_categoria(value):
    if not value:
        return None
    normalized = _normalize_text(value)
    if "no aplica" in normalized:
        return None
    if "multiple" in normalized:
        return "Múltiple"
    if "visual" in normalized:
        return "Visual"
    if "auditiva" in normalized or "hipoacusia" in normalized:
        return "Auditiva"
    if "fisica" in normalized:
        return "Física"
    if "psicosocial" in normalized:
        return "Psicosocial"
    if "intelectual" in normalized or "autismo" in normalized or "autista" in normalized:
        return "Intelectual"
    return _DISCAPACIDAD_CATEGORIA_MAP.get(normalized)




def _get_cache_path():
    return os.path.join(_get_cache_dir(), "seleccion_incluyente.json")


def cache_file_exists():
    return os.path.exists(_get_cache_path())


def save_cache_to_file():
    payload = {
        "form_id": FORM_ID,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "data": FORM_CACHE,
    }
    with open(_get_cache_path(), "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def _normalize_section_2_payload(payload):
    if not isinstance(payload, list):
        return payload
    normalized = []
    shared_desarrollo = ""
    for entry in payload:
        current = dict(entry or {})
        normalized.append(current)
        if not shared_desarrollo:
            shared_desarrollo = (current.get("desarrollo_actividad") or "").strip()
    for entry in normalized:
        entry["desarrollo_actividad"] = shared_desarrollo
    return normalized


def load_cache_from_file():
    path = _get_cache_path()
    if not os.path.exists(path):
        return False
    with open(path, "r", encoding="utf-8") as handle:
        payload = json.load(handle) or {}
    data = payload.get("data") or {}
    FORM_CACHE.clear()
    FORM_CACHE.update(data)
    section_2 = FORM_CACHE.get("section_2")
    if isinstance(section_2, list):
        FORM_CACHE["section_2"] = _normalize_section_2_payload(section_2)
    section_1 = data.get("section_1") or {}
    SECTION_1_CACHE.clear()
    SECTION_1_CACHE.update(section_1)
    return True


def clear_cache_file():
    path = _get_cache_path()
    if os.path.exists(path):
        os.remove(path)


def clear_form_cache():
    FORM_CACHE.clear()
    SECTION_1_CACHE.clear()


def set_section_cache(section_id, payload):
    if not section_id:
        raise ValueError("section_id requerido")
    FORM_CACHE[section_id] = payload


def get_form_cache():
    return dict(FORM_CACHE)


def _get_section_2_entries(payload=None):
    if payload is None:
        payload = FORM_CACHE.get("section_2", [])
    if not isinstance(payload, list):
        return []
    return [dict(entry or {}) for entry in payload]


def _resolve_template_variant(section_2_payload=None):
    total_oferentes = len(_get_section_2_entries(section_2_payload))
    if total_oferentes >= 2:
        return TEMPLATE_VARIANT_GROUP_2_PLUS
    return TEMPLATE_VARIANT_INDIVIDUAL


def _find_first_row_by_texts(ws, *texts):
    last_error = None
    for text in texts:
        if not text:
            continue
        try:
            return _find_row_by_text(ws, text)
        except Exception as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    raise ValueError("No se proporcionaron textos para buscar.")




def _find_template_path(template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    templates_dir = os.path.join(base_dir, "templates")
    if not os.path.isdir(templates_dir):
        raise FileNotFoundError("No existe la carpeta templates.")
    filename = TEMPLATE_FILENAME_BY_VARIANT.get(template_variant)
    if filename:
        exact_path = os.path.join(templates_dir, filename)
        if os.path.exists(exact_path):
            return exact_path
        if template_variant != TEMPLATE_VARIANT_INDIVIDUAL:
            raise FileNotFoundError(
                f"No se encontró el template '{filename}' para seleccion incluyente."
            )
    for name in os.listdir(templates_dir):
        if name.startswith("~$"):
            continue
        normalized = _normalize_text(name).replace("_", "")
        if "seleccion" in normalized and "incluyente" in normalized and normalized.endswith(".xlsx"):
            return os.path.join(templates_dir, name)
    raise FileNotFoundError("No se encontró el template de seleccion incluyente.")


def _normalize_dropdown_text(value):
    normalized = _normalize_text(value or "")
    normalized = normalized.replace('"', "").replace("'", "")
    normalized = " ".join(normalized.split())
    return normalized.strip(" .")


def _iter_sqref_cells(sqref):
    from openpyxl.utils.cell import get_column_letter, range_boundaries

    for token in str(sqref or "").split():
        if ":" not in token:
            yield token
            continue
        min_col, min_row, max_col, max_row = range_boundaries(token)
        for col_idx in range(min_col, max_col + 1):
            for row_idx in range(min_row, max_row + 1):
                yield f"{get_column_letter(col_idx)}{row_idx}"


def _clean_inline_dropdown_formula(formula):
    text = str(formula or "").strip()
    if not text:
        return ""
    text = text.replace('"&"', "")
    if text.startswith("="):
        text = text[1:]
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1]
    return text


def _split_inline_dropdown_fragments(formula):
    cleaned = _clean_inline_dropdown_formula(formula)
    if not cleaned:
        return []
    return [fragment.strip() for fragment in cleaned.split(",") if fragment.strip()]


def _reconstruct_dropdown_options(fragments, expected_options):
    if not fragments or not expected_options:
        return []
    total_fragments = len(fragments)
    total_options = len(expected_options)

    @lru_cache(maxsize=None)
    def _solve(fragment_idx, option_idx):
        if option_idx == total_options:
            return (0.0, []) if fragment_idx == total_fragments else (float("inf"), [])
        remaining_options = total_options - option_idx
        remaining_fragments = total_fragments - fragment_idx
        if remaining_fragments < remaining_options:
            return float("inf"), []

        best_score = float("inf")
        best_sequence = []
        max_take = remaining_fragments - (remaining_options - 1)
        expected_norm = _normalize_dropdown_text(expected_options[option_idx])
        for take in range(1, max_take + 1):
            candidate = ", ".join(fragments[fragment_idx: fragment_idx + take]).strip()
            candidate_norm = _normalize_dropdown_text(candidate)
            distance = 1.0 - SequenceMatcher(None, candidate_norm, expected_norm).ratio()
            rest_score, rest_sequence = _solve(fragment_idx + take, option_idx + 1)
            total_score = distance + rest_score
            if total_score < best_score:
                best_score = total_score
                best_sequence = [candidate] + rest_sequence
        return best_score, best_sequence

    _score, sequence = _solve(0, 0)
    if len(sequence) != total_options:
        return []
    return sequence


@lru_cache(maxsize=None)
def _get_template_validation_formula_map():
    from openpyxl import load_workbook

    path = _find_template_path(TEMPLATE_VARIANT_INDIVIDUAL)
    workbook = load_workbook(path)
    worksheet = workbook[workbook.sheetnames[0]]
    cell_map = {}
    for data_validation in getattr(worksheet.data_validations, "dataValidation", []):
        formula = getattr(data_validation, "formula1", None)
        if not formula:
            continue
        for cell in _iter_sqref_cells(getattr(data_validation, "sqref", "")):
            cell_map[cell] = formula
    workbook.close()
    return cell_map


def _get_list_field_cell(field_id):
    if field_id in EXCEL_MAPPING.get("section_1", {}):
        return EXCEL_MAPPING["section_1"][field_id]
    if field_id in SECTION_2_CELL_MAP:
        col, row = SECTION_2_CELL_MAP[field_id]
        return f"{col}{row}"
    return ""


@lru_cache(maxsize=None)
def _get_excel_canonical_options(field_id):
    manual = EXCEL_DROPDOWN_MANUAL_CANONICAL_OPTIONS.get(field_id)
    if manual:
        return tuple(manual)

    expected_options = LIST_FIELD_OPTIONS_BY_ID.get(field_id, [])
    if not expected_options:
        return tuple()
    cell = _get_list_field_cell(field_id)
    if not cell:
        return tuple(expected_options)
    formula = _get_template_validation_formula_map().get(cell)
    fragments = _split_inline_dropdown_fragments(formula)
    reconstructed = _reconstruct_dropdown_options(fragments, tuple(expected_options))
    if len(reconstructed) == len(expected_options):
        return tuple(reconstructed)
    return tuple(expected_options)


def normalize_excel_dropdown_value(field_id, raw_value, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if raw_value in (None, ""):
        return raw_value
    current = str(raw_value).strip()
    field_options = LIST_FIELD_OPTIONS_BY_ID.get(field_id)
    if not field_options:
        return raw_value

    canonical_options = list(_get_excel_canonical_options(field_id) or field_options)
    current_norm = _normalize_dropdown_text(current)

    for option in canonical_options:
        if _normalize_dropdown_text(option) == current_norm:
            return option

    for idx, option in enumerate(field_options):
        if _normalize_dropdown_text(option) == current_norm:
            if idx < len(canonical_options):
                return canonical_options[idx]
            return option

    _log_excel(
        f"WARN export_dropdown_unmatched field={field_id} template_variant={template_variant} "
        f"value={current!r}"
    )
    return raw_value


def _get_log_dir():
    output_path = FORM_CACHE.get("_output_path")
    if output_path:
        base_dir = os.path.dirname(output_path)
    else:
        base_dir = os.getcwd()
    log_dir = os.path.join(base_dir, "logs")
    os.makedirs(log_dir, exist_ok=True)
    return log_dir


def _log_excel(message):
    try:
        log_excel_event(message)
    except Exception:
        return


def _ensure_output_path(template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    template_path = _find_template_path(template_variant=template_variant)
    desktop = _get_desktop_dir()
    empresa_nombre = SECTION_1_CACHE.get("nombre_empresa") or "Empresa"
    safe_company = _sanitize_filename(empresa_nombre)
    if not safe_company:
        safe_company = "Empresa"
    output_dir = os.path.join(desktop, "Formatos Inclusion Laboral", safe_company)
    os.makedirs(output_dir, exist_ok=True)
    process_name = "Proceso de Seleccion Incluyente"
    output_name = f"{process_name} - {safe_company}.xlsx"
    output_path = _next_available_file_path(os.path.join(output_dir, output_name))
    shutil.copy2(template_path, output_path)
    FORM_CACHE["_output_path"] = output_path
    return output_path


def _get_sheet_by_name(workbook):
    target = _normalize_text(SHEET_NAME).replace(" ", "")
    for ws in workbook.Worksheets:
        name_norm = _normalize_text(ws.Name).replace(" ", "")
        if name_norm == target:
            return ws
    try:
        return workbook.Worksheets(SHEET_NAME)
    except Exception as exc:
        raise KeyError(f"No existe la hoja {SHEET_NAME}.") from exc


def _find_row_by_text(ws, text):
    cell = ws.Columns("A").Find(What=text, LookAt=1)
    if cell is not None:
        return cell.Row
    cell = ws.Columns("A").Find(What=text, LookAt=2)
    if cell is not None:
        return cell.Row
    target = _normalize_text(text)
    used = ws.UsedRange
    start_row = used.Row
    end_row = used.Row + used.Rows.Count - 1
    for row in range(start_row, end_row + 1):
        value = ws.Cells(row, 1).Value
        if not value:
            continue
        value_norm = _normalize_text(str(value))
        if value_norm == target:
            return row
    for row in range(start_row, end_row + 1):
        value = ws.Cells(row, 1).Value
        if not value:
            continue
        value_norm = _normalize_text(str(value))
        if target in value_norm:
            if target.startswith("2.") or target.startswith("5."):
                if value_norm.startswith(target):
                    return row
            else:
                return row
    raise ValueError(f"No se encontró el texto '{text}' en la columna A.")


def get_usuarios_reca_cedulas(env_path=".env"):
    params = {
        "select": "cedula_usuario",
        "cedula_usuario": "not.is.null",
        "order": "cedula_usuario.asc",
    }
    data = _supabase_get("usuarios_reca", params, env_path=env_path)
    return [row.get("cedula_usuario") for row in data if row.get("cedula_usuario")]


def get_usuario_reca_by_cedula(cedula, env_path=".env"):
    normalized = _normalize_cedula(cedula)
    if not normalized:
        return None
    select_cols = ",".join(
        [
            "cedula_usuario",
            "nombre_usuario",
            "discapacidad_usuario",
            "discapacidad_detalle",
            "certificado_porcentaje",
            "telefono_oferente",
            "fecha_nacimiento",
            "cargo_oferente",
            "contacto_emergencia",
            "parentesco",
            "telefono_emergencia",
            "resultado_certificado",
            "pendiente_otros_oferentes",
            "cuenta_pension",
            "tipo_pension",
        ]
    )
    params = {
        "select": select_cols,
        "cedula_usuario": f"eq.{normalized}",
        "limit": 1,
    }
    data = _supabase_get("usuarios_reca", params, env_path=env_path)
    return data[0] if data else None


def get_empresa_by_nit(nit, env_path=".env"):
    return evaluacion_accesibilidad.get_empresa_by_nit(nit, env_path=env_path)


def get_empresa_by_nombre(nombre, env_path=".env"):
    return evaluacion_accesibilidad.get_empresa_by_nombre(nombre, env_path=env_path)


def get_empresas_by_nombre_prefix(prefix, env_path=".env", limit=50):
    return evaluacion_accesibilidad.get_empresas_by_nombre_prefix(prefix, env_path=env_path, limit=limit)


def confirm_section_1(company_data, user_inputs):
    if not company_data:
        raise ValueError("No hay datos de empresa para confirmar.")
    payload = {}
    for field in SECTION_1["fields"]:
        field_id = field["id"]
        if field["source"] == "input":
            payload[field_id] = user_inputs.get(field_id)
        else:
            payload[field_id] = company_data.get(field_id)
    SECTION_1_CACHE.update(payload)
    set_section_cache("section_1", payload)
    FORM_CACHE["_last_section"] = "section_1"
    save_cache_to_file()
    return payload


def confirm_section_2(payload):
    if payload is None:
        raise ValueError("section_2 requerida")
    payload = _normalize_section_2_payload(payload)
    set_section_cache("section_2", payload)
    FORM_CACHE["_last_section"] = "section_2"
    save_cache_to_file()
    return payload


def confirm_section_5(payload):
    if payload is None:
        raise ValueError("section_5 requerida")
    set_section_cache("section_5", payload)
    FORM_CACHE["_last_section"] = "section_5"
    save_cache_to_file()
    return payload


def confirm_section_6(payload):
    if payload is None:
        raise ValueError("section_6 requerida")
    set_section_cache("section_6", payload)
    FORM_CACHE["_last_section"] = "section_6"
    save_cache_to_file()
    return payload


def sync_usuarios_reca(env_path=".env"):
    data = FORM_CACHE.get("section_2")
    if not data and cache_file_exists():
        load_cache_from_file()
        data = FORM_CACHE.get("section_2")
    if not data:
        return 0

    rows = []
    for entry in data:
        cedula = _normalize_cedula(entry.get("cedula"))
        if not cedula:
            continue
        discapacidad_detalle = (entry.get("discapacidad") or "").strip()
        discapacidad_usuario = _infer_discapacidad_categoria(discapacidad_detalle)
        row = {
            "cedula_usuario": cedula,
            "nombre_usuario": (entry.get("nombre_oferente") or "").strip(),
            "discapacidad_usuario": discapacidad_usuario,
            "discapacidad_detalle": discapacidad_detalle or None,
            "certificado_porcentaje": _normalize_decimal_value(
                entry.get("certificado_porcentaje"),
                decimal_separator=".",
            ),
            "telefono_oferente": (entry.get("telefono_oferente") or "").strip(),
            "fecha_nacimiento": _parse_date_value(entry.get("fecha_nacimiento")),
            "cargo_oferente": (entry.get("cargo_oferente") or "").strip(),
            "contacto_emergencia": (entry.get("nombre_contacto_emergencia") or "").strip(),
            "parentesco": (entry.get("parentesco") or "").strip(),
            "telefono_emergencia": (entry.get("telefono_emergencia") or "").strip(),
            "resultado_certificado": (entry.get("resultado_certificado") or "").strip(),
            "pendiente_otros_oferentes": (entry.get("pendiente_otros_oferentes") or "").strip(),
            "cuenta_pension": (entry.get("cuenta_pension") or "").strip(),
            "tipo_pension": (entry.get("tipo_pension") or "").strip(),
        }
        normalized_row = {
            key: (None if value == "" else value)
            for key, value in row.items()
        }
        rows.append(normalized_row)
    if rows:
        sync_result = _supabase_upsert_with_queue(
            "usuarios_reca",
            rows,
            env_path=env_path,
            on_conflict="cedula_usuario",
        )
        cedulas = [row.get("cedula_usuario") for row in rows if row.get("cedula_usuario")]
        preview = ", ".join(cedulas[:10])
        extra = "" if len(cedulas) <= 10 else f" (+{len(cedulas) - 10} mas)"
        status = sync_result.get("status") or "synced"
        _log_excel(
            f"SUPABASE usuarios_reca upsert status={status} count={len(rows)} cedulas={preview}{extra}"
        )
    return len(rows)


def _write_section_1(ws, payload, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if not payload:
        payload = SECTION_1_CACHE
    if not payload:
        try:
            if load_cache_from_file():
                payload = FORM_CACHE.get("section_1", {}) or SECTION_1_CACHE
        except Exception:
            payload = payload or {}
    mapping = SECTION_1_CELL_MAP_BY_TEMPLATE.get(
        template_variant,
        EXCEL_MAPPING.get("section_1", {}),
    )
    for key, cell in mapping.items():
        if key in payload:
            value = payload.get(key)
            value = normalize_excel_dropdown_value(
                key,
                value,
                template_variant=template_variant,
            )
            ws_write(ws, cell, value)
            _log_excel(
                f"WRITE section=section_1 cell={cell} key={key} value={value!r}"
            )


def _insert_person_block(ws, start_row, block_height, insert_at):
    start_end = start_row + block_height - 1
    dest_end = insert_at + block_height - 1
    source = ws.Range(f"A{start_row}:{SECTION_2_LAST_COLUMN}{start_end}")
    dest = ws.Range(f"A{insert_at}:{SECTION_2_LAST_COLUMN}{dest_end}")
    source.Copy()
    dest.Insert(Shift=-4121)
    for row_offset in range(block_height):
        ws.Rows(insert_at + row_offset).RowHeight = ws.Rows(start_row + row_offset).RowHeight
    ws.Application.CutCopyMode = False


def _write_section_2_entry(ws, entry, cell_map, *, row_offset=0, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    for field_id, (col, row) in cell_map.items():
        value = entry.get(field_id, "")
        if value == "":
            continue
        if field_id == "certificado_porcentaje":
            value = _coerce_excel_decimal_value(value)
        else:
            value = normalize_excel_dropdown_value(
                field_id,
                value,
                template_variant=template_variant,
            )
        target_row = row + row_offset
        _log_excel(
            f"WRITE section=section_2 cell={col}{target_row} key={field_id} value={value!r}"
        )
        ws_write(ws, f"{col}{target_row}", value)


def _write_section_2_individual(ws, oferentes):
    if not oferentes:
        return
    _log_excel(f"SECTION section=section_2 variant=individual total={len(oferentes)}")
    _write_section_2_entry(
        ws,
        oferentes[0],
        SECTION_2_INDIVIDUAL_CELL_MAP,
        template_variant=TEMPLATE_VARIANT_INDIVIDUAL,
    )


def _group_export_title_for_offerentes(total_oferentes):
    total = max(0, int(total_oferentes or 0))
    if total <= 1:
        return "PROCESO DE SELECCION INCLUYENTE INDIVIDUAL"
    if total <= 4:
        return "PROCESO DE SELECCION INCLUYENTE GRUPAL - 2 A 4 OFERENTES"
    if total <= 7:
        return "PROCESO DE SELECCION INCLUYENTE GRUPAL - 5 A 7 OFERENTES"
    if total <= 10:
        return "PROCESO DE SELECCION INCLUYENTE GRUPAL - 8 A 10 OFERENTES"
    return "PROCESO DE SELECCION INCLUYENTE GRUPAL - MAS DE 10 OFERENTES"


def _write_section_2_group(ws, oferentes):
    if not oferentes:
        return
    ws_write(ws, "G1", _group_export_title_for_offerentes(len(oferentes)))
    shared_desarrollo = ""
    for entry in oferentes:
        shared_desarrollo = (entry.get("desarrollo_actividad") or "").strip()
        if shared_desarrollo:
            break
    if shared_desarrollo:
        _log_excel(
            f"WRITE section=section_2 cell={SECTION_2_GROUP_SHARED_ACTIVITY_CELL} "
            f"key=desarrollo_actividad value={shared_desarrollo!r}"
        )
        ws_write(ws, SECTION_2_GROUP_SHARED_ACTIVITY_CELL, shared_desarrollo)

    if len(oferentes) > 2:
        for idx in range(2, len(oferentes)):
            insert_at = SECTION_2_GROUP_FIRST_BLOCK_START_ROW + (SECTION_2_GROUP_BLOCK_HEIGHT * idx)
            _insert_person_block(
                ws,
                SECTION_2_GROUP_SECOND_BLOCK_START_ROW,
                SECTION_2_GROUP_BLOCK_HEIGHT,
                insert_at,
            )
            _log_excel(
                f"INSERT section=section_2 variant=group rows={SECTION_2_GROUP_BLOCK_HEIGHT} at={insert_at}"
            )

    for idx, entry in enumerate(oferentes):
        row_offset = SECTION_2_GROUP_BLOCK_HEIGHT * idx
        title_row = SECTION_2_GROUP_FIRST_BLOCK_START_ROW + row_offset
        ws_write(ws, f"A{title_row}", f"OFERENTE {idx + 1}")
        _log_excel(
            f"WRITE section=section_2 cell=A{title_row} key=oferente_titulo value={'OFERENTE ' + str(idx + 1)!r}"
        )
        _write_section_2_entry(
            ws,
            entry,
            SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP,
            row_offset=row_offset,
            template_variant=TEMPLATE_VARIANT_GROUP_2_PLUS,
        )


def _write_section_2(ws, oferentes, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if template_variant == TEMPLATE_VARIANT_GROUP_2_PLUS:
        return _write_section_2_group(ws, oferentes)
    return _write_section_2_individual(ws, oferentes)


def _write_section_5(ws, payload, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if not payload:
        return
    if template_variant == TEMPLATE_VARIANT_GROUP_2_PLUS:
        anchor_row = _find_first_row_by_texts(ws, SECTION_5_GROUP_ANCHOR, SECTION_5_ANCHOR)
    else:
        anchor_row = _find_first_row_by_texts(ws, SECTION_5_ANCHOR, SECTION_5_GROUP_ANCHOR)
    ajustes_row = anchor_row + 1
    nota_row = anchor_row + 2
    ajustes_value = payload.get("ajustes_recomendaciones", "")
    nota_value = payload.get("nota", "")
    nota_value = f"Nota: {nota_value}" if nota_value else "Nota:"
    _log_excel(
        f"WRITE section=section_5 cell=A{ajustes_row} key=ajustes_recomendaciones value={ajustes_value!r}"
    )
    _log_excel(
        f"WRITE section=section_5 cell=A{nota_row} key=nota value={nota_value!r}"
    )
    ws_write(ws, f"A{ajustes_row}", ajustes_value)
    ws_write(ws, f"A{nota_row}", nota_value)


def _write_section_6(ws, payload, template_variant=TEMPLATE_VARIANT_INDIVIDUAL):
    if not payload:
        return
    mapping = EXCEL_MAPPING.get("section_6", {})
    title_row = _find_row_by_text(ws, "6. ASISTENTES")
    start_row = title_row + 1
    base_rows = SECTION_6_BASE_ROWS_BY_TEMPLATE.get(template_variant, mapping.get("rows", 4))
    nombre_col = mapping.get("nombre_col", "E")
    cargo_col = mapping.get("cargo_col", "M")
    total = len(payload)
    if total > base_rows:
        insert_at = start_row + base_rows
        template_row = start_row + base_rows - 1
        for _ in range(total - base_rows):
            ws.Rows(insert_at).Insert()
            ws.Rows(template_row).Copy(ws.Rows(insert_at))
            insert_at += 1
    for idx, entry in enumerate(payload):
        row = start_row + idx
        nombre = entry.get("nombre", "")
        cargo = entry.get("cargo", "")
        _log_excel(
            f"WRITE section=section_6 cell={nombre_col}{row} key=nombre value={nombre!r}"
        )
        _log_excel(
            f"WRITE section=section_6 cell={cargo_col}{row} key=cargo value={cargo!r}"
        )
        ws_write(ws, f"{nombre_col}{row}", nombre)
        ws_write(ws, f"{cargo_col}{row}", cargo)


def export_to_excel(clear_cache=True):
    clear_written_rows()
    section_2_payload = FORM_CACHE.get("section_2", [])
    template_variant = _resolve_template_variant(section_2_payload)
    output_path = _ensure_output_path(template_variant=template_variant)
    _log_excel(f"START export_all output={output_path}")
    try:
        import win32com.client as win32
    except ImportError as exc:
        _log_excel("ERROR export_all error=pywin32_not_installed")
        raise RuntimeError("pywin32 no esta instalado. Instala con pip install pywin32.") from exc
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(output_path)
        ws = _get_sheet_by_name(wb)
        _write_section_1(ws, FORM_CACHE.get("section_1", {}), template_variant=template_variant)
        _write_section_2(ws, section_2_payload, template_variant=template_variant)
        _write_section_5(ws, FORM_CACHE.get("section_5", {}), template_variant=template_variant)
        _write_section_6(
            ws,
            FORM_CACHE.get("section_6", []),
            template_variant=template_variant,
        )
        sanitize_logo_error_cells(wb)
        autofit_rows(ws, log_fn=_log_excel)
        wb.Save()
        _log_excel("SUCCESS export_all")
    except Exception as exc:
        _log_excel(f"ERROR export_all error={exc!r}")
        raise
    finally:
        if wb is not None:
            wb.Close(SaveChanges=True)
        excel.Quit()
    if clear_cache:
        clear_cache_file()
        clear_form_cache()
    return output_path

