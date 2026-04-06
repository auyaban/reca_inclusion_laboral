"""
google_sheets_client.py — Capa de acceso a Google Sheets API.

Responsabilidades:
  - Autenticación con service account (busca service-account.json en varias rutas)
  - Caché singleton del servicio Google Sheets (_SHEETS_SERVICE_CACHE) — thread-safe
  - Lectura y escritura de rangos de celdas en hojas de cálculo
  - Sanitización de fórmulas (FORMULA_PREFIXES) para evitar inyección en Sheets

Depende de: formularios/common (_load_env_file, _load_json_config, etc.),
            google_api_requests (execute_google_request_with_retry)
Usado por: módulos de formularios, app.py

NO contiene: lógica de negocio, mapeo de campos, construcción de payloads.
"""
import json
import os
import re
import sys

from openpyxl.utils.cell import range_boundaries

from formularios.common import (
    DEFAULT_SERVICE_ACCOUNT_FILE_NAME,
    FORBIDDEN_CONFIG_KEYS,
    _get_roaming_app_dir,
    _load_env_file,
    _load_json_config,
    _resolve_existing_path,
)
from google_api_requests import execute_google_request_with_retry


DEFAULT_CONFIG_PATH = "config.json"
SCOPE = "https://www.googleapis.com/auth/spreadsheets"
SPREADSHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
_SHEETS_SERVICE_CACHE = {"signature": None, "service": None}


def _get_bundle_dir():
    if getattr(sys, "frozen", False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def _load_config():
    return _load_json_config(
        DEFAULT_CONFIG_PATH,
        forbidden_keys=FORBIDDEN_CONFIG_KEYS,
    )


def _load_runtime_env():
    try:
        return _load_env_file(".env") or {}
    except Exception:
        return {}


def _service_account_search_dirs(env_source=None):
    dirs = []
    if env_source:
        dirs.append(os.path.dirname(os.path.abspath(env_source)))
    roaming_dir = _get_roaming_app_dir(create=False)
    if roaming_dir:
        dirs.append(roaming_dir)
    dirs.append(_get_bundle_dir())
    dirs.append(os.getcwd())
    unique = []
    seen = set()
    for path in dirs:
        key = os.path.normcase(str(path or "").strip())
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def _get_credentials_path():
    runtime_env, env_source = _load_env_file(".env", include_source=True)
    env_path = str(
        os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE")
        or os.getenv("GOOGLE_SHEETS_SA_JSON")
        or runtime_env.get("GOOGLE_SERVICE_ACCOUNT_FILE")
        or runtime_env.get("GOOGLE_SHEETS_SA_JSON")
        or ""
    ).strip()
    search_dirs = _service_account_search_dirs(env_source)
    if env_path:
        path = _resolve_existing_path(env_path, base_dirs=search_dirs)
        if not path:
            raise RuntimeError("El archivo de credenciales de Google Sheets configurado no es válido.")
        return path
    path = _resolve_existing_path(DEFAULT_SERVICE_ACCOUNT_FILE_NAME, base_dirs=search_dirs)
    if not path:
        raise RuntimeError(
            "Faltan credenciales de Google Sheets. Configure GOOGLE_SERVICE_ACCOUNT_FILE "
            "en el .env de la aplicación o coloque service-account.json en %APPDATA%\\RECA Inclusion Laboral."
        )
    return path


def extract_spreadsheet_id(value):
    text = str(value or "").strip()
    if not text:
        raise RuntimeError("Debe indicar spreadsheet_id o URL de Google Sheets.")

    match = SPREADSHEET_ID_RE.search(text)
    if match:
        return match.group(1)
    return text


def get_default_spreadsheet_id():
    runtime_env = _load_runtime_env()
    env_value = str(
        os.getenv("GOOGLE_SHEETS_DEFAULT_SPREADSHEET_ID")
        or runtime_env.get("GOOGLE_SHEETS_DEFAULT_SPREADSHEET_ID")
        or ""
    ).strip()
    if env_value:
        return extract_spreadsheet_id(env_value)

    config = _load_config()
    config_value = str(config.get("google_sheets_default_spreadsheet_id") or "").strip()
    if config_value:
        return extract_spreadsheet_id(config_value)

    raise RuntimeError(
        "Falta GOOGLE_SHEETS_DEFAULT_SPREADSHEET_ID o config.json con "
        "google_sheets_default_spreadsheet_id."
    )


def get_template_id(config_key, env_key=None):
    """Read a Google Sheets template ID from env var or config.json.

    Parameters
    ----------
    config_key : str
        Key inside ``config.json`` (e.g. ``"google_sheets_master_template_id"``).
    env_key : str | None
        Optional environment-variable name.  When *None* the *config_key* is
        upper-cased and used as env-var name (e.g. ``GOOGLE_SHEETS_MASTER_TEMPLATE_ID``).
    """
    if env_key is None:
        env_key = config_key.upper()

    runtime_env = _load_runtime_env()
    env_value = str(os.getenv(env_key) or runtime_env.get(env_key) or "").strip()
    if env_value:
        return extract_spreadsheet_id(env_value)

    config = _load_config()
    config_value = str(config.get(config_key) or "").strip()
    if config_value:
        return extract_spreadsheet_id(config_value)

    raise RuntimeError(
        f"Falta {env_key} o config.json con {config_key}."
    )


def get_master_template_id():
    """Shortcut – returns the master Google Sheets template spreadsheet ID."""
    return get_template_id("google_sheets_master_template_id")


def get_evaluacion_accesibilidad_template_id():
    return get_template_id(
        "google_sheets_evaluacion_accesibilidad_template_id",
        "GOOGLE_SHEETS_EVALUACION_ACCESIBILIDAD_TEMPLATE_ID",
    )


def get_google_sheets_service():
    creds_path = _get_credentials_path()
    try:
        stat_result = os.stat(creds_path)
        cache_signature = (
            os.path.normcase(os.path.abspath(creds_path)),
            int(getattr(stat_result, "st_mtime_ns", 0) or 0),
            int(getattr(stat_result, "st_size", 0) or 0),
        )
    except OSError:
        cache_signature = os.path.normcase(os.path.abspath(creds_path))
    cached_service = _SHEETS_SERVICE_CACHE.get("service")
    if cached_service is not None and _SHEETS_SERVICE_CACHE.get("signature") == cache_signature:
        return cached_service
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError(
            "Faltan dependencias de Google Sheets. "
            "Instala google-api-python-client y google-auth."
        ) from exc

    credentials = Credentials.from_service_account_file(creds_path, scopes=[SCOPE])
    service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    _SHEETS_SERVICE_CACHE["signature"] = cache_signature
    _SHEETS_SERVICE_CACHE["service"] = service
    return service


def clear_google_sheets_service_cache():
    _SHEETS_SERVICE_CACHE["signature"] = None
    _SHEETS_SERVICE_CACHE["service"] = None


def _escape_sheet_value(value, *, value_input_option="RAW"):
    if value is None:
        return ""
    mode = str(value_input_option or "RAW").strip().upper()
    if (
        mode == "USER_ENTERED"
        and isinstance(value, str)
        and value
        and value[0] in FORMULA_PREFIXES
    ):
        return "'" + value
    return value


def _sanitize_sheet_values(values, *, value_input_option="RAW"):
    if isinstance(values, (list, tuple)):
        return [
            _sanitize_sheet_values(item, value_input_option=value_input_option)
            for item in values
        ]
    return _escape_sheet_value(values, value_input_option=value_input_option)


def get_sheet_titles(spreadsheet_id_or_url):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()
    meta = execute_google_request_with_retry(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets.properties.title",
        ),
        operation_name="sheets.get_titles",
    )
    titles = []
    for sheet in meta.get("sheets", []):
        title = str((sheet.get("properties") or {}).get("title") or "").strip()
        if title:
            titles.append(title)
    return titles


def get_spreadsheet(spreadsheet_id_or_url, include_grid_data=False):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()
    return execute_google_request_with_retry(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            includeGridData=include_grid_data,
        ),
        operation_name="sheets.get_spreadsheet",
    )


def read_sheet_values(spreadsheet_id_or_url, range_name):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()
    response = execute_google_request_with_retry(
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=range_name),
        operation_name="sheets.read_values",
    )
    return list(response.get("values", []))


def batch_read_sheet_values(spreadsheet_id_or_url, ranges):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    cleaned = [str(item or "").strip() for item in (ranges or []) if str(item or "").strip()]
    if not cleaned:
        return {}
    service = get_google_sheets_service()
    response = execute_google_request_with_retry(
        service.spreadsheets()
        .values()
        .batchGet(spreadsheetId=spreadsheet_id, ranges=cleaned),
        operation_name="sheets.batch_get_values",
    )
    result = {}
    for item in response.get("valueRanges", []):
        range_name = str(item.get("range") or "").strip()
        if not range_name:
            continue
        result[range_name] = list(item.get("values", []))
    return result


def write_sheet_values(
    spreadsheet_id_or_url,
    range_name,
    values,
    value_input_option="RAW",
):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()
    body = {
        "values": _sanitize_sheet_values(
            values,
            value_input_option=value_input_option,
        )
    }
    return execute_google_request_with_retry(
        service.spreadsheets()
        .values()
        .update(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption=value_input_option,
            body=body,
        ),
        operation_name="sheets.write_values",
    )


def clear_sheet_ranges(spreadsheet_id_or_url, ranges):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    cleaned_ranges = [str(item or "").strip() for item in (ranges or []) if str(item or "").strip()]
    if not cleaned_ranges:
        return {"clearedRanges": []}
    service = get_google_sheets_service()
    return execute_google_request_with_retry(
        service.spreadsheets()
        .values()
        .batchClear(
            spreadsheetId=spreadsheet_id,
            body={"ranges": cleaned_ranges},
        ),
        operation_name="sheets.clear_ranges",
    )


def unmerge_cells_in_area(spreadsheet_id_or_url, sheet_name, start_row, end_row, start_col=0, end_col=21):
    """Unmerge all merged cells overlapping the given area (0-indexed rows/cols)."""
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()
    meta = execute_google_request_with_retry(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets(properties(sheetId,title),merges)",
        ),
        operation_name="sheets.get_merges",
    )
    sheet_id = None
    for s in meta.get("sheets", []):
        if s.get("properties", {}).get("title") == sheet_name:
            sheet_id = s["properties"]["sheetId"]
            break
    if sheet_id is None:
        return
    requests = []
    for s in meta.get("sheets", []):
        if s.get("properties", {}).get("sheetId") != sheet_id:
            continue
        for m in s.get("merges", []):
            sr, er = m.get("startRowIndex", 0), m.get("endRowIndex", 0)
            sc, ec = m.get("startColumnIndex", 0), m.get("endColumnIndex", 0)
            if sr < end_row and er > start_row and sc < end_col and ec > start_col:
                requests.append({"unmergeCells": {"range": {
                    "sheetId": sheet_id,
                    "startRowIndex": sr, "endRowIndex": er,
                    "startColumnIndex": sc, "endColumnIndex": ec,
                }}})
    if requests:
        execute_google_request_with_retry(
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": requests},
            ),
            operation_name="sheets.unmerge_cells",
        )


def list_protected_ranges(spreadsheet_id_or_url):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    spreadsheet = get_spreadsheet(spreadsheet_id, include_grid_data=False)
    rows = []
    for sheet in spreadsheet.get("sheets", []):
        props = sheet.get("properties", {}) or {}
        sheet_id = props.get("sheetId")
        sheet_title = str(props.get("title") or "")
        for protected_range in sheet.get("protectedRanges", []) or []:
            if not isinstance(protected_range, dict):
                continue
            protected_range_id = protected_range.get("protectedRangeId")
            if protected_range_id is None:
                continue
            rows.append(
                {
                    "protectedRangeId": int(protected_range_id),
                    "sheetId": sheet_id,
                    "sheetTitle": sheet_title,
                    "warningOnly": bool(protected_range.get("warningOnly")),
                    "description": str(protected_range.get("description") or ""),
                }
            )
    return rows


def clear_protected_ranges(spreadsheet_id_or_url, protected_range_ids=None):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    if protected_range_ids is None:
        target_ids = [item["protectedRangeId"] for item in list_protected_ranges(spreadsheet_id)]
    else:
        target_ids = []
        for protected_range_id in protected_range_ids or []:
            if protected_range_id is None:
                continue
            target_ids.append(int(protected_range_id))
    if not target_ids:
        return {"deletedProtectedRangeIds": [], "deletedProtectedRangeCount": 0}

    unique_target_ids = list(dict.fromkeys(target_ids))
    service = get_google_sheets_service()
    execute_google_request_with_retry(
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                "requests": [
                    {"deleteProtectedRange": {"protectedRangeId": protected_range_id}}
                    for protected_range_id in unique_target_ids
                ]
            },
        ),
        operation_name="sheets.clear_protected_ranges",
    )
    return {
        "deletedProtectedRangeIds": unique_target_ids,
        "deletedProtectedRangeCount": len(unique_target_ids),
    }


def remove_sheet_protection(spreadsheet_id_or_url):
    """Remove all protectedRanges from every sheet in the spreadsheet.

    When a template is copied via files().copy(), any sheet protection from
    the template is inherited by the copy.  Professionals who receive the copy
    are not the owner, so the inherited protection blocks cells that should be
    editable.  Calling this right after the copy is created strips all
    protected ranges so the copy is fully editable.
    """
    return clear_protected_ranges(spreadsheet_id_or_url)


def copy_sheet_to_spreadsheet(
    source_spreadsheet_id_or_url,
    source_sheet_name,
    destination_spreadsheet_id_or_url,
    *,
    new_sheet_name=None,
):
    source_spreadsheet_id = extract_spreadsheet_id(source_spreadsheet_id_or_url)
    destination_spreadsheet_id = extract_spreadsheet_id(destination_spreadsheet_id_or_url)
    service = get_google_sheets_service()

    meta = execute_google_request_with_retry(
        service.spreadsheets().get(
            spreadsheetId=source_spreadsheet_id,
            fields="sheets.properties(sheetId,title)",
        ),
        operation_name="sheets.get_source_sheet",
    )
    source_sheet_id = None
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {}) or {}
        if str(props.get("title") or "").strip() == str(source_sheet_name or "").strip():
            source_sheet_id = props.get("sheetId")
            break
    if source_sheet_id is None:
        raise RuntimeError(f"No existe la hoja '{source_sheet_name}' en el archivo maestro.")

    copied = execute_google_request_with_retry(
        service.spreadsheets()
        .sheets()
        .copyTo(
            spreadsheetId=source_spreadsheet_id,
            sheetId=source_sheet_id,
            body={"destinationSpreadsheetId": destination_spreadsheet_id},
        ),
        operation_name="sheets.copy_sheet_to_spreadsheet",
    )
    copied_sheet_id = copied.get("sheetId")
    copied_title = str(copied.get("title") or "").strip()
    target_title = str(new_sheet_name or copied_title or source_sheet_name or "").strip()

    if copied_sheet_id is not None and target_title and target_title != copied_title:
        execute_google_request_with_retry(
            service.spreadsheets().batchUpdate(
                spreadsheetId=destination_spreadsheet_id,
                body={
                    "requests": [
                        {
                            "updateSheetProperties": {
                                "properties": {
                                    "sheetId": copied_sheet_id,
                                    "title": target_title,
                                },
                                "fields": "title",
                            }
                        }
                    ]
                },
            ),
            operation_name="sheets.rename_copied_sheet",
        )
        copied_title = target_title

    return {
        "sheetId": copied_sheet_id,
        "title": copied_title or target_title,
    }


def insert_template_rows(
    spreadsheet_id_or_url,
    sheet_name,
    *,
    insert_at_row,
    template_row,
    count,
    paste_type="PASTE_NORMAL",
):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    total_rows = max(0, int(count or 0))
    if total_rows <= 0:
        return {"insertedRows": 0}

    service = get_google_sheets_service()
    meta = execute_google_request_with_retry(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets.properties(sheetId,title)",
        ),
        operation_name="sheets.get_insert_rows_target",
    )

    sheet_id = None
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {}) or {}
        if str(props.get("title") or "").strip() == str(sheet_name or "").strip():
            sheet_id = props.get("sheetId")
            break
    if sheet_id is None:
        raise RuntimeError(f"No existe la hoja '{sheet_name}' en la spreadsheet destino.")

    insert_index = max(0, int(insert_at_row or 1) - 1)
    template_index = max(0, int(template_row or 1) - 1)

    requests = [
        {
            "insertDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": insert_index,
                    "endIndex": insert_index + total_rows,
                },
                "inheritFromBefore": insert_index > 0,
            }
        }
    ]
    if paste_type:
        requests.append(
            {
                "copyPaste": {
                    "source": {
                        "sheetId": sheet_id,
                        "startRowIndex": template_index,
                        "endRowIndex": template_index + 1,
                    },
                    "destination": {
                        "sheetId": sheet_id,
                        "startRowIndex": insert_index,
                        "endRowIndex": insert_index + total_rows,
                    },
                    "pasteType": str(paste_type),
                    "pasteOrientation": "NORMAL",
                }
            }
        )

    execute_google_request_with_retry(
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests},
        ),
        operation_name="sheets.insert_template_rows",
    )
    return {"sheetId": sheet_id, "insertedRows": total_rows}


def insert_template_block_rows(
    spreadsheet_id_or_url,
    sheet_name,
    *,
    insert_at_row,
    template_start_row,
    template_end_row,
    repeat_count=1,
    paste_type="PASTE_NORMAL",
    copy_row_heights=False,
):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    total_blocks = max(0, int(repeat_count or 0))
    if total_blocks <= 0:
        return {"insertedRows": 0, "insertedBlocks": 0}

    start_row = int(template_start_row or 0)
    end_row = int(template_end_row or 0)
    if start_row <= 0 or end_row < start_row:
        raise RuntimeError("template_start_row/template_end_row invalidos para insertar bloques.")

    block_height = (end_row - start_row) + 1
    total_rows = block_height * total_blocks

    service = get_google_sheets_service()
    meta = execute_google_request_with_retry(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets.properties(sheetId,title)",
        ),
        operation_name="sheets.get_insert_block_target",
    )

    sheet_id = None
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {}) or {}
        if str(props.get("title") or "").strip() == str(sheet_name or "").strip():
            sheet_id = props.get("sheetId")
            break
    if sheet_id is None:
        raise RuntimeError(f"No existe la hoja '{sheet_name}' en la spreadsheet destino.")

    insert_index = max(0, int(insert_at_row or 1) - 1)
    source_start_index = start_row - 1
    source_end_index = end_row

    execute_google_request_with_retry(
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                "requests": [
                    {
                        "insertDimension": {
                            "range": {
                                "sheetId": sheet_id,
                                "dimension": "ROWS",
                                "startIndex": insert_index,
                                "endIndex": insert_index + total_rows,
                            },
                            "inheritFromBefore": insert_index > 0,
                        }
                    }
                ]
            },
        ),
        operation_name="sheets.insert_template_block_rows",
    )

    if paste_type:
        # The master sheet contains merged cells around the vinculado block. After
        # inserting rows, Google may temporarily carry overlapping merge metadata
        # into the new range, which makes copyPaste fail with
        # "can't perform a paste that partially intersects a merge".
        unmerge_cells_in_area(
            spreadsheet_id,
            sheet_name,
            insert_index,
            insert_index + total_rows,
        )

        copy_requests = []
        for block_index in range(total_blocks):
            destination_start = insert_index + (block_index * block_height)
            copy_requests.append(
                {
                    "copyPaste": {
                        "source": {
                            "sheetId": sheet_id,
                            "startRowIndex": source_start_index,
                            "endRowIndex": source_end_index,
                        },
                        "destination": {
                            "sheetId": sheet_id,
                            "startRowIndex": destination_start,
                            "endRowIndex": destination_start + block_height,
                        },
                        "pasteType": str(paste_type),
                        "pasteOrientation": "NORMAL",
                    }
                }
            )

        execute_google_request_with_retry(
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": copy_requests},
            ),
            operation_name="sheets.copy_template_block_rows",
        )

    if copy_row_heights:
        # copyPaste PASTE_NORMAL does not copy row dimension properties (pixelSize).
        # The inserted rows inherit their height from inheritFromBefore, which is the
        # last row before the insertion point.  When that row is a tall "nota" cell
        # (excluded from auto-resize), every structure row in the new blocks ends up
        # at the wrong height.  We fix this by reading the template row heights and
        # applying them explicitly to every copied block.
        height_meta = execute_google_request_with_retry(
            service.spreadsheets().get(
                spreadsheetId=spreadsheet_id,
                ranges=[f"'{sheet_name}'!A{start_row}:A{end_row}"],
                includeGridData=True,
                fields="sheets.data.rowMetadata",
            ),
            operation_name="sheets.get_template_row_heights",
        )
        row_heights: list[int] = []
        for _sheet in height_meta.get("sheets", []):
            for grid in _sheet.get("data", []):
                for rm in grid.get("rowMetadata", []):
                    pix = int(rm.get("pixelSize") or 0)
                    row_heights.append(pix if pix > 0 else 21)
            if row_heights:
                break

        # Pad / truncate to exactly block_height entries
        if len(row_heights) < block_height:
            row_heights.extend([21] * (block_height - len(row_heights)))
        else:
            row_heights = row_heights[:block_height]

        height_requests = []
        for block_idx in range(total_blocks):
            dest_start = insert_index + (block_idx * block_height)
            # Group consecutive rows that share the same height into one request
            run_start = dest_start
            run_h = row_heights[0]
            for i in range(1, block_height):
                h = row_heights[i]
                if h == run_h:
                    continue
                height_requests.append(
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id,
                                "dimension": "ROWS",
                                "startIndex": run_start,
                                "endIndex": dest_start + i,
                            },
                            "properties": {"pixelSize": run_h},
                            "fields": "pixelSize",
                        }
                    }
                )
                run_start = dest_start + i
                run_h = h
            height_requests.append(
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "ROWS",
                            "startIndex": run_start,
                            "endIndex": dest_start + block_height,
                        },
                        "properties": {"pixelSize": run_h},
                        "fields": "pixelSize",
                    }
                }
            )

        if height_requests:
            execute_google_request_with_retry(
                service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={"requests": height_requests},
                ),
                operation_name="sheets.copy_template_row_heights",
            )

    return {
        "sheetId": sheet_id,
        "insertedRows": total_rows,
        "insertedBlocks": total_blocks,
        "blockHeight": block_height,
    }


def hide_sheets(spreadsheet_id_or_url, sheet_names_to_keep):
    """Hide all sheets in the spreadsheet except those in *sheet_names_to_keep*.

    At least one sheet must remain visible (Google Sheets requirement).
    Sheets whose title matches any entry in *sheet_names_to_keep* stay visible;
    every other sheet is hidden.

    Parameters
    ----------
    spreadsheet_id_or_url : str
        Spreadsheet ID or full URL.
    sheet_names_to_keep : list[str]
        Sheet titles that should remain visible.
    """
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    keep = {s.strip() for s in (sheet_names_to_keep or []) if s and s.strip()}
    if not keep:
        return

    service = get_google_sheets_service()
    meta = execute_google_request_with_retry(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets.properties",
        ),
        operation_name="sheets.get_hide_targets",
    )

    resolved_keep_ids = set()
    sheets = list(meta.get("sheets", []))
    for sheet in sheets:
        props = sheet.get("properties", {}) or {}
        title = str(props.get("title") or "").strip()
        sheet_id = props.get("sheetId")
        if title in keep and sheet_id is not None:
            resolved_keep_ids.add(sheet_id)
    if not resolved_keep_ids:
        missing = ", ".join(sorted(keep))
        raise RuntimeError(
            f"No existe ninguna hoja con el nombre solicitado en la spreadsheet destino: {missing}"
        )

    visible_after_update = 0
    unhide_requests = []
    hide_requests = []
    for sheet in sheets:
        props = sheet.get("properties", {}) or {}
        sheet_id = props.get("sheetId")
        is_hidden = bool(props.get("hidden", False))
        should_keep = sheet_id in resolved_keep_ids
        will_be_hidden = not should_keep
        if not will_be_hidden:
            visible_after_update += 1
        if sheet_id is None:
            continue
        if should_keep and is_hidden:
            unhide_requests.append(
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": sheet_id,
                            "hidden": False,
                        },
                        "fields": "hidden",
                    }
                }
            )
        elif not should_keep and not is_hidden:
            hide_requests.append(
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": sheet_id,
                            "hidden": True,
                        },
                        "fields": "hidden",
                    }
                }
            )

    if visible_after_update <= 0:
        raise RuntimeError("Google Sheets requiere al menos una hoja visible.")

    requests = [*unhide_requests, *hide_requests]
    if not requests:
        return

    execute_google_request_with_retry(
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests},
        ),
        operation_name="sheets.hide_sheets",
    )


def _col_letter_to_index(col_str):
    """Convert column letters (A, B, ..., Z, AA, ...) to 0-indexed column number."""
    result = 0
    for ch in col_str.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def _parse_a1_cell(a1_range):
    """Parse "'SheetName'!B5" into (sheet_name, row_0idx, col_0idx)."""
    import re
    m = re.match(r"^'([^']+)'!([A-Z]+)(\d+)$", a1_range)
    if not m:
        raise ValueError(f"Cannot parse A1 range: {a1_range}")
    sheet_name = m.group(1)
    col_idx = _col_letter_to_index(m.group(2))
    row_idx = int(m.group(3)) - 1
    return sheet_name, row_idx, col_idx


def _parse_a1_rect(a1_range):
    match = re.match(r"^'([^']+)'!([A-Z]+\d+(?::[A-Z]+\d+)?)$", str(a1_range or "").strip())
    if not match:
        raise ValueError(f"Cannot parse A1 range: {a1_range}")
    sheet_name = match.group(1)
    cell_range = match.group(2)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return {
        "sheet_name": sheet_name,
        "start_row_index": max(0, min_row - 1),
        "end_row_index": max_row,
        "start_col_index": max(0, min_col - 1),
        "end_col_index": max_col,
    }


def auto_resize_rows_for_ranges(spreadsheet_id_or_url, ranges):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    cleaned_ranges = [str(item or "").strip() for item in (ranges or []) if str(item or "").strip()]
    if not cleaned_ranges:
        return {"updatedRanges": []}

    spreadsheet = get_spreadsheet(spreadsheet_id, include_grid_data=False)
    name_to_id = {}
    for sheet in spreadsheet.get("sheets", []):
        props = sheet.get("properties", {}) or {}
        name_to_id[str(props.get("title") or "")] = props.get("sheetId")

    sheet_segments = {}
    updated_ranges = []
    for range_name in cleaned_ranges:
        try:
            rect = _parse_a1_rect(range_name)
        except ValueError:
            continue
        sheet_id = name_to_id.get(rect["sheet_name"])
        if sheet_id is None:
            continue
        updated_ranges.append(range_name)
        segments = sheet_segments.setdefault(sheet_id, [])
        segments.append((rect["start_row_index"], rect["end_row_index"]))

    requests = []
    for sheet_id, segments in sheet_segments.items():
        merged_segments = []
        for start_index, end_index in sorted(segments):
            if not merged_segments:
                merged_segments.append([start_index, end_index])
                continue
            last_start, last_end = merged_segments[-1]
            if start_index <= last_end:
                merged_segments[-1][1] = max(last_end, end_index)
                continue
            merged_segments.append([start_index, end_index])
        for start_index, end_index in merged_segments:
            requests.append(
                {
                    "autoResizeDimensions": {
                        "dimensions": {
                            "sheetId": sheet_id,
                            "dimension": "ROWS",
                            "startIndex": start_index,
                            "endIndex": end_index,
                        }
                    }
                }
            )
    if requests:
        service = get_google_sheets_service()
        execute_google_request_with_retry(
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": requests},
            ),
            operation_name="sheets.auto_resize_rows",
        )
    return {"updatedRanges": updated_ranges}


def _filter_auto_resize_ranges(ranges, excluded_rows_by_sheet=None):
    cleaned_ranges = [str(item or "").strip() for item in (ranges or []) if str(item or "").strip()]
    if not cleaned_ranges:
        return []
    excluded = {}
    for sheet_name, rows in (excluded_rows_by_sheet or {}).items():
        key = str(sheet_name or "").strip()
        if not key:
            continue
        normalized_rows = set()
        for row in rows or []:
            try:
                row_number = int(row or 0)
            except Exception:
                continue
            if row_number > 0:
                normalized_rows.add(row_number)
        if normalized_rows:
            excluded[key] = normalized_rows
    if not excluded:
        return cleaned_ranges

    filtered = []
    for range_name in cleaned_ranges:
        try:
            sheet_name, row_idx, _col_idx = _parse_a1_cell(range_name)
        except ValueError:
            filtered.append(range_name)
            continue
        if (row_idx + 1) in excluded.get(sheet_name, set()):
            continue
        filtered.append(range_name)
    return filtered


def set_native_checkboxes(spreadsheet_id_or_url, checkbox_cells):
    """Set native Google Sheets checkboxes on the given cells.

    Parameters
    ----------
    spreadsheet_id_or_url : str
        Spreadsheet ID or full URL.
    checkbox_cells : list[dict]
        Each dict: {"range": "'Sheet'!A1", "value": True/False}
    """
    if not checkbox_cells:
        return
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    service = get_google_sheets_service()

    meta = execute_google_request_with_retry(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets.properties",
        ),
        operation_name="sheets.get_checkbox_targets",
    )
    name_to_id = {}
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {})
        name_to_id[props.get("title", "")] = props.get("sheetId")

    requests = []
    for cell in checkbox_cells:
        sheet_name, row_idx, col_idx = _parse_a1_cell(cell["range"])
        sheet_id = name_to_id.get(sheet_name)
        if sheet_id is None:
            continue
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_idx,
                    "endRowIndex": row_idx + 1,
                    "startColumnIndex": col_idx,
                    "endColumnIndex": col_idx + 1,
                },
                "cell": {
                    "dataValidation": {
                        "condition": {"type": "BOOLEAN"},
                    },
                    "userEnteredValue": {
                        "boolValue": bool(cell.get("value", False)),
                    },
                },
                "fields": "dataValidation,userEnteredValue",
            }
        })

    if requests:
        execute_google_request_with_retry(
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": requests},
            ),
            operation_name="sheets.set_native_checkboxes",
        )


def set_sheet_ranges_bold(spreadsheet_id_or_url, ranges, *, bold):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    cleaned_ranges = [str(item or "").strip() for item in (ranges or []) if str(item or "").strip()]
    if not cleaned_ranges:
        return {"updatedRanges": []}

    spreadsheet = get_spreadsheet(spreadsheet_id, include_grid_data=False)
    name_to_id = {}
    for sheet in spreadsheet.get("sheets", []):
        props = sheet.get("properties", {}) or {}
        name_to_id[str(props.get("title") or "")] = props.get("sheetId")

    requests = []
    updated_ranges = []
    for range_name in cleaned_ranges:
        rect = _parse_a1_rect(range_name)
        sheet_id = name_to_id.get(rect["sheet_name"])
        if sheet_id is None:
            continue
        updated_ranges.append(range_name)
        requests.append(
            {
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": rect["start_row_index"],
                        "endRowIndex": rect["end_row_index"],
                        "startColumnIndex": rect["start_col_index"],
                        "endColumnIndex": rect["end_col_index"],
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "textFormat": {
                                "bold": bool(bold),
                            }
                        }
                    },
                    "fields": "userEnteredFormat.textFormat.bold",
                }
            }
        )
    if requests:
        service = get_google_sheets_service()
        execute_google_request_with_retry(
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": requests},
            ),
            operation_name="sheets.set_ranges_bold",
        )
    return {"updatedRanges": updated_ranges}


def batch_write_sheet_updates(
    spreadsheet_id_or_url,
    updates,
    value_input_option="RAW",
    auto_resize_rows=True,
    auto_resize_excluded_rows=None,
):
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    rows = []
    for update in updates or []:
        if not isinstance(update, dict):
            continue
        range_name = str(update.get("range") or "").strip()
        if not range_name:
            continue
        value = update.get("value", "")
        value = _escape_sheet_value(value, value_input_option=value_input_option)
        rows.append(
            {
                "range": range_name,
                "majorDimension": "ROWS",
                "values": [[value]],
            }
        )
    if not rows:
        return {"totalUpdatedCells": 0, "totalUpdatedRows": 0, "responses": []}

    service = get_google_sheets_service()
    response = execute_google_request_with_retry(
        service.spreadsheets()
        .values()
        .batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                "valueInputOption": value_input_option,
                "data": rows,
            },
        ),
        operation_name="sheets.batch_write_updates",
    )
    if auto_resize_rows:
        resize_ranges = _filter_auto_resize_ranges(
            [row.get("range") for row in rows],
            excluded_rows_by_sheet=auto_resize_excluded_rows,
        )
        if resize_ranges:
            auto_resize_rows_for_ranges(
                spreadsheet_id,
                resize_ranges,
            )
    return response
