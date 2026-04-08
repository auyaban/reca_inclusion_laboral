"""
Backfill one-shot para reconstruir el mapeo historico de empresa/NIT por cÃ©dula
en contrataciÃ³n incluyente.

Se conserva por trazabilidad operativa; no forma parte del runtime de la app.
"""

import json
import re
import sys
import os
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

def _iso_date(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if value is None:
        return ""
    return str(value).strip()


def build_mapping(workbook_path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - utilidad operativa opcional
        raise SystemExit(
            "Este script requiere openpyxl de forma opcional para leer el archivo histórico."
        ) from exc

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    header_idx = {header: index for index, header in enumerate(headers)}
    required = {"REFERENCIA", "CEDULA", "EMPRESA", "NIT", "FECHA"}
    missing = sorted(required - set(header_idx))
    if missing:
        raise SystemExit(f"Faltan columnas requeridas: {', '.join(missing)}")

    records = {}
    for row in rows:
        referencia = str(row[header_idx["REFERENCIA"]] or "").strip().upper()
        if not referencia.startswith("IL6.PC"):
            continue
        raw_cedulas = str(row[header_idx["CEDULA"]] or "").strip()
        if not raw_cedulas:
            continue
        empresa_nombre = str(row[header_idx["EMPRESA"]] or "").strip()
        empresa_nit = str(row[header_idx["NIT"]] or "").strip()
        fecha = _iso_date(row[header_idx["FECHA"]])
        for cedula in re.split(r"[^0-9]+", raw_cedulas):
            cedula = cedula.strip()
            if not cedula:
                continue
            candidate = {
                "cedula_usuario": cedula,
                "empresa_nit": empresa_nit,
                "empresa_nombre": empresa_nombre,
                "fecha_referencia": fecha,
                "referencia_servicio": referencia,
            }
            previous = records.get(cedula)
            if previous is None or candidate["fecha_referencia"] >= previous["fecha_referencia"]:
                records[cedula] = candidate
    return list(records.values())


def _request_json(url, headers, method="GET", body=None):
    request = urllib.request.Request(url, data=body, headers=headers, method=method)
    with urllib.request.urlopen(request, timeout=30) as response:
        payload = response.read()
    if not payload:
        return None
    return json.loads(payload.decode("utf-8"))


def apply_mapping(mapping, supabase_url, service_role, missing_only=True):
    base_url = supabase_url.rstrip("/") + "/rest/v1/usuarios_reca"
    headers = {
        "apikey": service_role,
        "Authorization": f"Bearer {service_role}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    updated = 0
    skipped_existing = 0
    skipped_missing = 0
    for record in mapping:
        cedula = record["cedula_usuario"]
        query = urllib.parse.urlencode(
            {
                "select": "cedula_usuario,empresa_nit,empresa_nombre",
                "cedula_usuario": f"eq.{cedula}",
                "limit": "1",
            }
        )
        rows = _request_json(f"{base_url}?{query}", headers) or []
        if not rows:
            skipped_missing += 1
            continue
        existing = rows[0]
        if missing_only and (
            (existing.get("empresa_nit") or "").strip()
            or (existing.get("empresa_nombre") or "").strip()
        ):
            skipped_existing += 1
            continue
        patch_headers = dict(headers)
        patch_headers["Prefer"] = "return=minimal"
        body = json.dumps(
            {
                "empresa_nit": record.get("empresa_nit") or None,
                "empresa_nombre": record.get("empresa_nombre") or None,
            }
        ).encode("utf-8")
        _request_json(
            f"{base_url}?cedula_usuario=eq.{cedula}",
            patch_headers,
            method="PATCH",
            body=body,
        )
        updated += 1
    return {
        "mapped": len(mapping),
        "updated": updated,
        "skipped_existing": skipped_existing,
        "skipped_missing": skipped_missing,
    }


def main():
    args = sys.argv[1:]
    apply_changes = False
    if "--apply" in args:
        apply_changes = True
        args.remove("--apply")
    if len(args) != 1:
        raise SystemExit(
            "Uso: python scripts/build_contratacion_company_backfill.py <ruta_ods_xlsx> [--apply]"
        )
    workbook_path = Path(args[0])
    if not workbook_path.exists():
        raise SystemExit(f"No existe el archivo: {workbook_path}")
    mapping = build_mapping(workbook_path)
    if not apply_changes:
        json.dump(mapping, sys.stdout, ensure_ascii=False, indent=2)
        return
    supabase_url = os.environ.get("SUPABASE_URL")
    service_role = os.environ.get("SUPABASE_SERVICE_ROLE")
    if not supabase_url or not service_role:
        raise SystemExit(
            "Para --apply define SUPABASE_URL y SUPABASE_SERVICE_ROLE en el entorno."
        )
    result = apply_mapping(mapping, supabase_url, service_role, missing_only=True)
    json.dump(result, sys.stdout, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
