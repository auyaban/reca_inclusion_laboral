from __future__ import annotations

import os
import unittest

from openpyxl import load_workbook

from formularios.contratacion_incluyente import contratacion_incluyente as contratacion


class ContratacionIncluyenteTemplateTests(unittest.TestCase):
    def test_resolve_template_variant_uses_group_layout_for_one_or_more_vinculados(self) -> None:
        self.assertEqual(
            contratacion._resolve_template_variant([]),
            contratacion.TEMPLATE_VARIANT_INDIVIDUAL,
        )
        self.assertEqual(
            contratacion._resolve_template_variant([{"numero": "1"}]),
            contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(
            contratacion._resolve_template_variant([{"numero": "1"}, {"numero": "2"}]),
            contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )

    def test_group_and_individual_template_files_exist(self) -> None:
        individual = contratacion._find_template_path(contratacion.TEMPLATE_VARIANT_INDIVIDUAL)
        group = contratacion._find_template_path(contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS)
        self.assertTrue(os.path.exists(individual))
        self.assertTrue(os.path.exists(group))
        self.assertTrue(individual.lower().endswith("contratacion_incluyente.xlsx"))
        self.assertTrue(group.lower().endswith("contratacion_incluyente.xlsx"))

    def test_certificado_porcentaje_cell_keeps_general_format_in_template(self) -> None:
        path = contratacion._find_template_path(contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS)
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb[contratacion.SHEET_NAME_BY_VARIANT[contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS]]
            self.assertNotIn("%", ws["K23"].number_format)
        finally:
            wb.close()

    def test_sheet_name_by_variant_matches_real_workbooks(self) -> None:
        for variant in (
            contratacion.TEMPLATE_VARIANT_INDIVIDUAL,
            contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
        ):
            path = contratacion._find_template_path(variant)
            wb = load_workbook(path, read_only=True)
            try:
                self.assertIn(contratacion.SHEET_NAME_BY_VARIANT[variant], wb.sheetnames)
            finally:
                wb.close()

    def test_individual_template_mapping_keeps_expected_rows(self) -> None:
        mapping = contratacion.SECTION_2_INDIVIDUAL_CELL_MAP
        self.assertEqual(mapping["numero"], ("A", 23))
        self.assertEqual(mapping["nombre_oferente"], ("C", 23))
        self.assertEqual(mapping["cedula"], ("H", 23))
        self.assertEqual(mapping["certificado_porcentaje"], ("K", 23))
        self.assertEqual(mapping["discapacidad"], ("L", 23))
        self.assertEqual(mapping["telefono_oferente"], ("O", 23))
        self.assertEqual(mapping["contacto_emergencia"], ("I", 26))
        self.assertEqual(mapping["parentesco"], ("M", 26))
        self.assertEqual(mapping["telefono_emergencia"], ("Q", 26))
        self.assertEqual(mapping["tipo_contrato"], ("G", 29))
        self.assertEqual(mapping["desarrollo_actividad"], ("A", 15))
        self.assertEqual(mapping["contrato_lee_nivel_apoyo"], ("G", 33))
        self.assertEqual(mapping["rutas_atencion_nivel_apoyo"], ("G", 69))

    def test_group_template_mapping_matches_audited_rows(self) -> None:
        mapping = contratacion.SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP
        self.assertEqual(mapping["numero"], ("A", 23))
        self.assertEqual(mapping["nombre_oferente"], ("C", 23))
        self.assertEqual(mapping["cedula"], ("H", 23))
        self.assertEqual(mapping["certificado_porcentaje"], ("K", 23))
        self.assertEqual(mapping["discapacidad"], ("L", 23))
        self.assertEqual(mapping["telefono_oferente"], ("O", 23))
        self.assertEqual(mapping["contacto_emergencia"], ("I", 26))
        self.assertEqual(mapping["parentesco"], ("M", 26))
        self.assertEqual(mapping["telefono_emergencia"], ("Q", 26))
        self.assertEqual(mapping["tipo_contrato"], ("G", 29))
        self.assertEqual(mapping["contrato_lee_nivel_apoyo"], ("G", 33))
        self.assertEqual(mapping["condiciones_salariales_nivel_apoyo"], ("G", 42))
        self.assertEqual(mapping["rutas_atencion_nivel_apoyo"], ("G", 69))
        self.assertNotIn("desarrollo_actividad", mapping)
        self.assertEqual(mapping["numero"][1] + contratacion.SECTION_2_GROUP_BLOCK_HEIGHT, 75)
        self.assertEqual(contratacion.SECTION_2_GROUP_SHARED_ACTIVITY_CELL, "A15")

    def test_group_template_offsets_keep_second_and_third_blocks_aligned(self) -> None:
        mapping = contratacion.SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP
        second_offset = contratacion.SECTION_2_GROUP_BLOCK_HEIGHT
        third_offset = contratacion.SECTION_2_GROUP_BLOCK_HEIGHT * 2

        self.assertEqual(mapping["numero"][1] + second_offset, 75)
        self.assertEqual(mapping["telefono_oferente"][1] + second_offset, 75)
        self.assertEqual(mapping["contacto_emergencia"][1] + second_offset, 78)
        self.assertEqual(mapping["numero"][1] + third_offset, 127)
        self.assertEqual(mapping["telefono_oferente"][1] + third_offset, 127)
        self.assertEqual(mapping["contacto_emergencia"][1] + third_offset, 130)
        self.assertEqual(mapping["rutas_atencion_nivel_apoyo"][1] + third_offset, 173)

    def test_section_2_template_mappings_target_empty_cells(self) -> None:
        path = contratacion._find_template_path(contratacion.TEMPLATE_VARIANT_INDIVIDUAL)
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb[contratacion.SHEET_NAME_BY_VARIANT[contratacion.TEMPLATE_VARIANT_INDIVIDUAL]]
            for field_id, (col, row) in contratacion.SECTION_2_INDIVIDUAL_CELL_MAP.items():
                value = ws[f"{col}{row}"].value
                self.assertIn(
                    value,
                    (None, ""),
                    msg=f"individual field {field_id} points to non-empty cell {col}{row}: {value!r}",
                )
            for field_id, (col, row) in contratacion.SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP.items():
                value = ws[f"{col}{row}"].value
                self.assertIn(
                    value,
                    (None, ""),
                    msg=f"group field {field_id} points to non-empty cell {col}{row}: {value!r}",
                )
        finally:
            wb.close()

    def test_group_block_rows_expand_from_single_template_block(self) -> None:
        self.assertEqual(contratacion._section_2_group_block_start_row(0), 19)
        self.assertEqual(contratacion._section_2_group_block_start_row(1), 71)
        self.assertEqual(contratacion._section_2_group_block_start_row(2), 123)
        self.assertEqual(contratacion._section_2_group_insert_row(1), 71)
        self.assertEqual(contratacion._section_2_group_insert_row(2), 123)

    def test_section_6_and_section_7_rows_shift_only_by_inserted_group_blocks(self) -> None:
        self.assertEqual(
            contratacion._section_row_after_section_2(
                72,
                1,
                contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
            ),
            72,
        )
        self.assertEqual(
            contratacion._section_row_after_section_2(
                72,
                2,
                contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
            ),
            124,
        )
        self.assertEqual(
            contratacion._section_row_after_section_2(
                78,
                2,
                contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
            ),
            130,
        )

    def test_section_7_base_rows_change_by_template_variant(self) -> None:
        self.assertEqual(
            contratacion.SECTION_7_BASE_ROWS_BY_TEMPLATE[contratacion.TEMPLATE_VARIANT_INDIVIDUAL],
            3,
        )
        self.assertEqual(
            contratacion.SECTION_7_BASE_ROWS_BY_TEMPLATE[contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS],
            4,
        )

    def test_group_export_title_for_offerentes_uses_expected_ranges(self) -> None:
        self.assertEqual(
            contratacion._group_export_title_for_offerentes(1),
            "PROCESO DE CONTRATACIÓN INCLUYENTE INDIVIDUAL",
        )
        self.assertEqual(
            contratacion._group_export_title_for_offerentes(2),
            "PROCESO CONTRATACION INCLUYENTE GRUPAL - 2 A 4 VINCULADOS",
        )
        self.assertEqual(
            contratacion._group_export_title_for_offerentes(5),
            "PROCESO CONTRATACION INCLUYENTE GRUPAL - 5 A 7 VINCULADOS",
        )
        self.assertEqual(
            contratacion._group_export_title_for_offerentes(8),
            "PROCESO CONTRATACION INCLUYENTE GRUPAL - 8 A 10 VINCULADOS",
        )
        self.assertEqual(
            contratacion._group_export_title_for_offerentes(11),
            "PROCESO CONTRATACION INCLUYENTE GRUPAL - MAS DE 10 VINCULADOS",
        )

    def test_contrato_lectura_options_are_variant_specific(self) -> None:
        individual = contratacion.get_section_2_field_options(
            "contrato_lee_observacion",
            contratacion.TEMPLATE_VARIANT_INDIVIDUAL,
        )
        group = contratacion.get_section_2_field_options(
            "contrato_lee_observacion",
            contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertIn("0. No requiere apoyo.", individual)
        self.assertNotIn("0. No requiere apoyo.", group)

    def test_normalize_excel_dropdown_value_maps_grupo_etnico_cual(self) -> None:
        value = contratacion.normalize_excel_dropdown_value(
            "grupo_etnico_cual",
            "Indigena",
            template_variant=contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(value, "Indígena")

    def test_normalize_excel_dropdown_value_maps_grupo_etnico_alias(self) -> None:
        value = contratacion.normalize_excel_dropdown_value(
            "grupo_etnico_cual",
            "Gitano (ROM)",
            template_variant=contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(value, "Rom o Gitano")

    def test_normalize_excel_dropdown_value_maps_tipo_contrato_amplio(self) -> None:
        value = contratacion.normalize_excel_dropdown_value(
            "tipo_contrato",
            "Contrato de trabajo a término fijo",
            template_variant=contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(value, "Contrato de trabajo a término fijo")

    def test_normalize_excel_dropdown_value_maps_tipo_contrato_alias(self) -> None:
        value = contratacion.normalize_excel_dropdown_value(
            "tipo_contrato",
            "Prestacion de servicios",
            template_variant=contratacion.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(value, "Contrato por prestación de servicios")


if __name__ == "__main__":
    unittest.main()
