import inspect
from pathlib import Path
import unittest
import warnings

from openpyxl import load_workbook

from formularios.common import _normalize_text
from formularios.condiciones_vacante import condiciones_vacante as condiciones
from formularios.contratacion_incluyente import contratacion_incluyente as contratacion
from formularios.evaluacion_programa import evaluacion_accesibilidad
from formularios.induccion_operativa import induccion_operativa
from formularios.induccion_organizacional import induccion_organizacional
from formularios.presentacion_programa import presentacion_programa
from formularios.seleccion_incluyente import seleccion_incluyente as seleccion
from formularios.seguimientos import seguimientos
from formularios.sensibilizacion import sensibilizacion


REPO_ROOT = Path(__file__).resolve().parents[1]
TEMPLATES_DIR = REPO_ROOT / "templates"

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl",
)


def _worksheet(path_name: str, sheet_name: str):
    workbook = load_workbook(TEMPLATES_DIR / path_name, data_only=False)
    target = _normalize_text(sheet_name).replace(" ", "")
    for worksheet in workbook.worksheets:
        if _normalize_text(worksheet.title).replace(" ", "") == target:
            return worksheet
    raise KeyError(f"Worksheet {sheet_name} does not exist in {path_name}.")


def _cell_text(ws, row: int, column: str = "A") -> str:
    return _normalize_text(str(ws[f"{column}{row}"].value or ""))


class DeterministicTemplateLayoutTests(unittest.TestCase):
    def assertRowStartsWith(self, ws, row: int, expected_prefix: str, column: str = "A") -> None:
        actual = _cell_text(ws, row, column=column)
        self.assertTrue(
            actual.startswith(_normalize_text(expected_prefix)),
            msg=f"row={row} column={column} actual={actual!r} expected_prefix={expected_prefix!r}",
        )

    def assertSourceDoesNotUseAnchorLookup(self, func) -> None:
        self.assertNotIn("_find_row_by_text(", inspect.getsource(func))

    def test_seleccion_incluyente_template_contract_rows(self) -> None:
        ws = _worksheet("seleccion_incluyente.xlsx", "4. PROCESO DE SELECCION INCLUYE")
        self.assertRowStartsWith(
            ws,
            seleccion.SECTION_5_TITLE_ROW_BY_TEMPLATE[seleccion.TEMPLATE_VARIANT_INDIVIDUAL],
            "5. AJUSTES RAZONABLES Y RECOMENDACIONES",
        )
        self.assertRowStartsWith(
            ws,
            seleccion.SECTION_6_TITLE_ROW_BY_TEMPLATE[seleccion.TEMPLATE_VARIANT_INDIVIDUAL],
            "6. ASISTENTES",
        )

    def test_presentacion_programa_template_contract(self) -> None:
        workbook = load_workbook(TEMPLATES_DIR / "presentación_programa.xlsx", data_only=False)
        try:
            ws = workbook.worksheets[0]
            for cell in (
                presentacion_programa.EXCEL_MAPPING["section_1"]["fecha_visita"],
                presentacion_programa.EXCEL_MAPPING["section_1"]["correo_asesor"],
                presentacion_programa.EXCEL_MAPPING["section_4"]["acuerdos_observaciones"],
            ):
                with self.subTest(cell=cell):
                    self.assertIsNotNone(ws[cell])
        finally:
            workbook.close()

    def test_evaluacion_accesibilidad_template_contract(self) -> None:
        ws = _worksheet("evaluacion_accesibilidad.xlsx", evaluacion_accesibilidad.SHEET_NAME)
        for cell in (
            evaluacion_accesibilidad.EXCEL_MAPPING["section_1"]["fecha_visita"],
            evaluacion_accesibilidad.EXCEL_MAPPING["section_2_1"]["transporte_publico_accesible"],
            evaluacion_accesibilidad.EXCEL_MAPPING["section_6"]["observaciones_generales"],
        ):
            with self.subTest(cell=cell):
                self.assertIsNotNone(ws[cell])

    def test_seguimientos_template_contract(self) -> None:
        workbook = load_workbook(TEMPLATES_DIR / "seguimientos.xlsx", data_only=False)
        try:
            base_sheet_name = seguimientos._get_base_sheet_name_from_workbook(workbook)
            self.assertIn(base_sheet_name, {seguimientos.SHEET_BASE, seguimientos.LEGACY_SHEET_BASE})
            self.assertIn(seguimientos.SHEET_FINAL, workbook.sheetnames)
            ws = workbook[base_sheet_name]
            for cell in ("D8", "R8", "B23", "P29"):
                with self.subTest(cell=cell):
                    self.assertIsNotNone(ws[cell])
        finally:
            workbook.close()

    def test_contratacion_incluyente_template_contract_rows(self) -> None:
        ws = _worksheet("contratacion_incluyente.xlsx", "5. PROCESO CONTRATACION INCLUYE")
        self.assertRowStartsWith(
            ws,
            contratacion.SECTION_6_TITLE_ROW_BY_TEMPLATE[contratacion.TEMPLATE_VARIANT_INDIVIDUAL],
            "5. AJUSTES RAZONABLES Y RECOMENDACIONES",
        )
        self.assertRowStartsWith(
            ws,
            contratacion.SECTION_7_TITLE_ROW_BY_TEMPLATE[contratacion.TEMPLATE_VARIANT_INDIVIDUAL],
            "6. ASISTENTES",
        )

    def test_induccion_organizacional_template_rows_and_offsets(self) -> None:
        ws = _worksheet("induccion_organizacional.xlsx", "6. INDUCCION ORGANIZACIONAL")
        self.assertRowStartsWith(ws, induccion_organizacional.SECTION_3_TITLE_ROW, "3. DESARROLLO DEL PROCESO")
        self.assertRowStartsWith(
            ws,
            induccion_organizacional.SECTION_4_TITLE_ROW,
            "4. AJUSTES RAZONABLES AL PROCESO DE INDUCCION",
        )
        self.assertRowStartsWith(ws, induccion_organizacional.SECTION_5_TITLE_ROW, "5. OBSERVACIONES")
        self.assertRowStartsWith(ws, induccion_organizacional.SECTION_6_TITLE_ROW, "6. ASISTENTES")
        self.assertEqual(
            induccion_organizacional._row_after_section_2(
                induccion_organizacional.SECTION_5_TITLE_ROW, 1
            ),
            67,
        )
        self.assertEqual(
            induccion_organizacional._row_after_section_2(
                induccion_organizacional.SECTION_6_TITLE_ROW, 4
            ),
            73,
        )

    def test_induccion_operativa_template_rows_and_offsets(self) -> None:
        ws = _worksheet("induccion_operativa.xlsx", "7. INDUCCION OPERATIVA")
        self.assertRowStartsWith(
            ws,
            induccion_operativa.SECTION_3_TITLE_ROW,
            "3. DESARROLLO DEL PROCESO DE INDUCCION OPERATIVA",
        )
        self.assertRowStartsWith(ws, induccion_operativa.SECTION_4_TITLE_ROW, "4. HABILIDADES SOCIOEMOCIONALES")
        self.assertRowStartsWith(ws, induccion_operativa.SECTION_5_TITLE_ROW, "5. NIVEL DE APOYO REQUERIDO")
        self.assertRowStartsWith(ws, induccion_operativa.SECTION_6_TITLE_ROW, "6. AJUSTES RAZONABLES REQUERIDOS")
        self.assertRowStartsWith(
            ws,
            induccion_operativa.SECTION_7_TITLE_ROW,
            "7. PRIMER SEGUIMIENTO ESTABLECIDO PARA EL VINCULADO",
        )
        self.assertRowStartsWith(
            ws,
            induccion_operativa.SECTION_8_TITLE_ROW,
            "8. OBSERVACIONES /RECOMENDACIONES",
        )
        self.assertRowStartsWith(ws, induccion_operativa.SECTION_9_TITLE_ROW, "9.ASISTENTES")
        self.assertEqual(
            induccion_operativa._row_after_section_2(induccion_operativa.SECTION_8_TITLE_ROW, 1),
            66,
        )
        self.assertEqual(
            induccion_operativa._row_after_section_2(induccion_operativa.SECTION_9_TITLE_ROW, 3),
            72,
        )

    def test_condiciones_vacante_template_rows(self) -> None:
        ws = _worksheet("revision_condicion.xlsx", "3. REVISION DE LAS CONDICIONES")
        self.assertRowStartsWith(ws, condiciones.SECTION_7_TITLE_ROW, "7. OBSERVACIONES / RECOMENDACIONES")
        self.assertRowStartsWith(ws, condiciones.SECTION_8_TITLE_ROW, "8.ASISTENTES")
        self.assertEqual(condiciones._section_7_content_row_for_payload([{}] * 4), 159)
        self.assertEqual(condiciones._section_7_content_row_for_payload([{}] * 8), 163)
        self.assertEqual(condiciones._section_8_start_row_for_payload([{}] * 4), 161)
        self.assertEqual(condiciones._section_8_start_row_for_payload([{}] * 8), 165)

    def test_sensibilizacion_template_rows(self) -> None:
        ws = _worksheet("sensibilizacion.xlsx", "8. SENSIBILIZACION")
        self.assertRowStartsWith(ws, sensibilizacion.SECTION_3_TITLE_ROW, "3. OBSERVACIONES")
        self.assertRowStartsWith(ws, sensibilizacion.SECTION_5_TITLE_ROW, "5. ASISTENTES")

    def test_hot_paths_are_deterministic(self) -> None:
        anchor_free_functions = [
            condiciones._write_section_with_ws,
            induccion_organizacional._write_section_2,
            induccion_organizacional._write_section_3,
            induccion_organizacional._write_section_4,
            induccion_organizacional._write_section_5,
            induccion_organizacional._write_section_6,
            induccion_operativa._write_section_2,
            induccion_operativa._write_section_3,
            induccion_operativa._write_section_4,
            induccion_operativa._write_section_5,
            induccion_operativa._write_section_6,
            induccion_operativa._write_section_7,
            induccion_operativa._write_section_8,
            induccion_operativa._write_section_9,
            sensibilizacion._write_section_3,
            sensibilizacion._write_section_5,
            condiciones._write_section_with_ws,
            seleccion._write_section_5,
            seleccion._write_section_6,
            contratacion._write_section_6,
            contratacion._write_section_7,
        ]
        for func in anchor_free_functions:
            with self.subTest(func=func.__module__ + "." + func.__name__):
                self.assertSourceDoesNotUseAnchorLookup(func)

    def test_selection_and_contratacion_no_template_validation_parsing_in_hot_path(self) -> None:
        for module in (seleccion, contratacion):
            source = inspect.getsource(module)
            with self.subTest(module=module.__name__):
                self.assertNotIn("_get_template_validation_formula_map", source)
                self.assertNotIn("_find_first_row_by_texts", source)
                self.assertNotIn("_iter_sqref_cells", source)


if __name__ == "__main__":
    unittest.main()
