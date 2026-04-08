from __future__ import annotations

import shutil
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from formularios.seguimientos import seguimientos


class SeguimientosWorkflowTests(unittest.TestCase):
    def test_workflow_keeps_sheet_9_suggested_when_only_prefilled_data_exists(self) -> None:
        base_payload = {
            "nombre_vinculado": "Persona Demo",
            "cedula": "123456",
            "cargo_vinculado": "Auxiliar",
            "discapacidad": "Fisica",
            "seguimiento_fechas_1_3": ["2026-04-01", "", ""],
            "seguimiento_fechas_4_6": ["", "", ""],
        }

        with patch.object(
            seguimientos,
            "get_case_meta",
            return_value={"base_sheet_name": seguimientos.SHEET_BASE, "max_seguimientos": 3},
        ):
            with patch.object(seguimientos, "get_base_payload", return_value=base_payload):
                with patch.object(
                    seguimientos,
                    "get_followup_payload",
                    return_value={
                        "item_autoevaluacion": ["", ""],
                        "item_eval_empresa": ["", ""],
                        "tipo_apoyo": "",
                    },
                ):
                    workflow = seguimientos.get_workflow_state({"source": "drive", "file_id": "demo"})

        self.assertEqual(workflow["next_followup"], 1)
        self.assertEqual(workflow["completed_followups"], [])
        self.assertEqual(workflow["editable_sheet"], seguimientos.SHEET_BASE)
        self.assertEqual(workflow["suggested_sheet"], seguimientos.SHEET_BASE)
        self.assertFalse(workflow["base_completed"])
        self.assertEqual(workflow["base_coverage_percent"], 0)
        self.assertEqual(workflow["message"], "Empieza por la ficha inicial del proceso.")
        self.assertEqual(workflow["stage_model"][0]["stage_id"], "base_process")
        self.assertEqual(workflow["stage_model"][0]["title"], "Ficha inicial del proceso")
        self.assertEqual(workflow["stage_model"][-1]["title"], "Resultado final")
        self.assertEqual(workflow["stage_model"][-1]["status"], "review_only")

    def test_workflow_advances_to_followup_2_after_base_and_followup_1_reach_90_percent(self) -> None:
        base_payload = {
            "fecha_visita": "2026-04-01",
            "modalidad": "Presencial",
            "contacto_emergencia": "Contacto Demo",
            "parentesco": "Hermana",
            "telefono_emergencia": "3111111111",
            "certificado_discapacidad": "Si",
            "certificado_porcentaje": "25%",
            "tipo_contrato": "Indefinido",
            "fecha_firma_contrato": "2026-03-01",
            "fecha_inicio_contrato": "2026-03-05",
            "fecha_fin_contrato": "",
            "apoyos_ajustes": "Ajuste razonable",
            "funciones_1_5": ["F1", "F2", "F3", "F4", "F5"],
            "funciones_6_10": ["F6", "F7", "F8", "F9", "F10"],
            "seguimiento_fechas_1_3": ["2026-04-01", "", ""],
            "seguimiento_fechas_4_6": ["", "", ""],
        }

        def _payload_for_index(_case_ref, index):
            if index == 1:
                return {
                    "modalidad": "Mixta",
                    "fecha_seguimiento": "2026-04-01",
                    "item_autoevaluacion": ["Excelente"] * 19,
                    "item_eval_empresa": ["Bien"] * 19,
                    "tipo_apoyo": "Requiere apoyo bajo.",
                    "empresa_eval": ["Bien"] * 8,
                    "situacion_encontrada": "Sin novedades",
                    "estrategias_ajustes": "Acompañamiento mensual",
                }
            return {
                "modalidad": "",
                "fecha_seguimiento": "",
                "item_autoevaluacion": [""],
                "item_eval_empresa": [""],
                "tipo_apoyo": "",
                "empresa_eval": [""],
                "situacion_encontrada": "",
                "estrategias_ajustes": "",
            }

        with patch.object(
            seguimientos,
            "get_case_meta",
            return_value={"base_sheet_name": seguimientos.SHEET_BASE, "max_seguimientos": 3},
        ):
            with patch.object(seguimientos, "get_base_payload", return_value=base_payload):
                with patch.object(seguimientos, "get_followup_payload", side_effect=_payload_for_index):
                    workflow = seguimientos.get_workflow_state({"source": "drive", "file_id": "demo"})

        self.assertEqual(workflow["next_followup"], 2)
        self.assertEqual(workflow["completed_followups"], [1])
        self.assertEqual(workflow["editable_sheet"], f"{seguimientos.SHEET_PREFIX}2")
        self.assertEqual(workflow["suggested_sheet"], f"{seguimientos.SHEET_PREFIX}2")
        self.assertEqual(
            workflow["completed_sheets"],
            ["Ficha inicial del proceso", "Seguimiento 1"],
        )
        self.assertEqual(workflow["sheet_progress"][0]["coverage_percent"], 95)
        followup_2 = next(
            entry
            for entry in workflow["stage_model"]
            if entry["sheet_name"] == f"{seguimientos.SHEET_PREFIX}2"
        )
        self.assertEqual(followup_2["stage_id"], "followup_2")
        self.assertEqual(followup_2["title"], "Seguimiento 2")
        self.assertTrue(followup_2["is_suggested"])
        self.assertTrue(followup_2["is_editable"])


class SeguimientosBaseSheetMappingTests(unittest.TestCase):
    def test_get_base_sheet_name_from_spreadsheet_accepts_master_title(self) -> None:
        spreadsheet = {
            "sheets": [
                {
                    "properties": {
                        "title": "9. SEGUIMIENTO AL PROCESO DE INCLUSIÓN LABORAL",
                    }
                }
            ]
        }

        self.assertEqual(
            seguimientos._get_base_sheet_name_from_spreadsheet(spreadsheet),
            seguimientos.SHEET_BASE,
        )

    def test_build_base_payload_from_user_row_does_not_prefill_fecha_firma_contrato(self) -> None:
        payload = seguimientos._build_base_payload_from_user_row(
            {
                "nombre_usuario": "Persona Demo",
                "cedula_usuario": "123456",
                "telefono_oferente": "3000000000",
                "correo_oferente": "persona@example.com",
                "cargo_oferente": "Analista",
                "certificado_discapacidad": "Si",
                "certificado_porcentaje": "25%",
                "discapacidad_detalle": "Física",
                "tipo_contrato": "Indefinido",
                "fecha_firma_contrato": "2026-04-01",
            }
        )

        self.assertNotIn("fecha_firma_contrato", payload)

    def test_build_base_sheet_updates_writes_apoyos_ajustes_to_e21(self) -> None:
        updates = seguimientos._build_base_sheet_updates(
            {"apoyos_ajustes": "Texto de ajuste"},
            base_sheet_name=seguimientos.SHEET_BASE,
        )

        target = next(
            item for item in updates if item["value"] == "Texto de ajuste"
        )

        self.assertEqual(target["range"], f"'{seguimientos.SHEET_BASE}'!E21")

    def test_local_roundtrip_keeps_label_in_a21_and_value_in_e21(self) -> None:
        repo_root = Path(__file__).resolve().parents[1]
        template_path = repo_root / "templates" / "seguimientos.xlsx"
        with tempfile.TemporaryDirectory() as tmp_dir:
            case_path = Path(tmp_dir) / "seguimiento_demo.xlsx"
            shutil.copy2(template_path, case_path)

            seguimientos.save_base_payload(
                str(case_path),
                {"apoyos_ajustes": "Ajuste razonable de prueba"},
            )

            workbook = load_workbook(case_path, data_only=False)
            try:
                sheet_name = seguimientos._get_base_sheet_name_from_workbook(workbook)
                ws = workbook[sheet_name]
                self.assertEqual(
                    str(ws["A21"].value or "").strip(),
                    "Apoyos y/o ajustes razonables requeridos:",
                )
                self.assertEqual(ws["E21"].value, "Ajuste razonable de prueba")
            finally:
                workbook.close()

            payload = seguimientos.get_base_payload(str(case_path))
            self.assertEqual(payload["apoyos_ajustes"], "Ajuste razonable de prueba")

    def test_get_followup_payload_ignores_template_seeded_defaults_until_user_edits(self) -> None:
        repo_root = Path(__file__).resolve().parents[1]
        template_path = repo_root / "templates" / "seguimientos.xlsx"
        with tempfile.TemporaryDirectory() as tmp_dir:
            case_path = Path(tmp_dir) / "seguimiento_demo.xlsx"
            shutil.copy2(template_path, case_path)

            seguimientos.save_base_payload(
                str(case_path),
                {
                    "nombre_vinculado": "Persona Demo",
                    "cedula": "123456",
                    "cargo_vinculado": "Auxiliar",
                    "discapacidad": "Fisica",
                    "seguimiento_fechas_1_3": ["2026-04-01", "", ""],
                    "seguimiento_fechas_4_6": ["", "", ""],
                },
            )

            payload = seguimientos.get_followup_payload(str(case_path), 1)

            self.assertEqual(payload["tipo_apoyo"], "")
            self.assertEqual(payload["modalidad"], "")
            self.assertTrue(all(not str(item or "").strip() for item in payload["item_autoevaluacion"]))
            self.assertTrue(all(not str(item or "").strip() for item in payload["item_eval_empresa"]))


if __name__ == "__main__":
    unittest.main()
