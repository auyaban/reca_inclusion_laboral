from __future__ import annotations

import os
import unittest

from openpyxl import load_workbook

from formularios.seleccion_incluyente import seleccion_incluyente as si


class SeleccionIncluyenteTemplateTests(unittest.TestCase):
    @staticmethod
    def _variant_map(variant: str):
        if variant == si.TEMPLATE_VARIANT_GROUP_2_PLUS:
            return si.SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP
        return si.SECTION_2_INDIVIDUAL_CELL_MAP

    def _expected_canonical_options(self, field_id: str, variant: str):
        return tuple(si._get_excel_canonical_options(field_id, template_variant=variant))

    def test_resolve_template_variant_uses_group_layout_for_one_or_more_offerentes(self) -> None:
        self.assertEqual(si._resolve_template_variant([]), si.TEMPLATE_VARIANT_INDIVIDUAL)
        self.assertEqual(
            si._resolve_template_variant([{"numero": "1"}]),
            si.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(
            si._resolve_template_variant([{"numero": "1"}, {"numero": "2"}]),
            si.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(
            si._resolve_template_variant([{"numero": "1"}, {"numero": "2"}, {"numero": "3"}, {"numero": "4"}]),
            si.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(
            si._resolve_template_variant([{"numero": "1"}, {"numero": "2"}, {"numero": "3"}, {"numero": "4"}, {"numero": "5"}]),
            si.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )

    def test_individual_template_mapping_keeps_expected_rows(self) -> None:
        mapping = si.SECTION_2_INDIVIDUAL_CELL_MAP
        self.assertEqual(mapping["numero"], ("A", 19))
        self.assertEqual(mapping["cargo_oferente"], ("A", 21))
        self.assertEqual(mapping["fecha_nacimiento"], ("N", 21))
        self.assertEqual(mapping["desplazamiento_nivel_apoyo"], ("I", 39))
        self.assertEqual(mapping["aseo_nivel_apoyo"], ("I", 57))
        self.assertEqual(mapping["discriminacion_nivel_apoyo"], ("I", 73))
        self.assertNotIn("desarrollo_actividad", mapping)

    def test_group_template_mapping_matches_corrected_rows(self) -> None:
        mapping = si.SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP
        self.assertEqual(mapping["numero"], ("A", 19))
        self.assertEqual(mapping["nombre_oferente"], ("C", 19))
        self.assertEqual(mapping["cargo_oferente"], ("A", 21))
        self.assertEqual(mapping["fecha_nacimiento"], ("N", 21))
        self.assertEqual(mapping["pendiente_otros_oferentes"], ("G", 22))
        self.assertEqual(mapping["cuenta_pension"], ("I", 23))
        self.assertEqual(mapping["desplazamiento_nivel_apoyo"], ("I", 39))
        self.assertEqual(mapping["ubicacion_nivel_apoyo"], ("I", 42))
        self.assertEqual(mapping["dinero_nivel_apoyo"], ("I", 45))
        self.assertEqual(mapping["presentacion_nivel_apoyo"], ("I", 49))
        self.assertEqual(mapping["comunicacion_escrita_nivel_apoyo"], ("I", 51))
        self.assertEqual(mapping["comunicacion_verbal_nivel_apoyo"], ("I", 53))
        self.assertEqual(mapping["decisiones_nivel_apoyo"], ("I", 55))
        self.assertEqual(mapping["aseo_nivel_apoyo"], ("I", 57))
        self.assertEqual(mapping["instrumentales_nivel_apoyo"], ("I", 62))
        self.assertEqual(mapping["actividades_nivel_apoyo"], ("I", 68))
        self.assertEqual(mapping["discriminacion_nivel_apoyo"], ("I", 73))
        self.assertNotIn("desarrollo_actividad", mapping)

    def test_individual_and_group_use_same_section_2_template_layout(self) -> None:
        self.assertEqual(si.SECTION_2_INDIVIDUAL_CELL_MAP, si.SECTION_2_GROUP_FIRST_BLOCK_CELL_MAP)
        self.assertEqual(si.SECTION_2_GROUP_SHARED_ACTIVITY_CELL, "A14")

    def test_group_template_file_exists(self) -> None:
        path = si._find_template_path(si.TEMPLATE_VARIANT_GROUP_2_PLUS)
        self.assertTrue(os.path.exists(path))
        self.assertTrue(path.lower().endswith("seleccion_incluyente.xlsx"))

    def test_certificado_porcentaje_cell_is_percent_formatted_in_template(self) -> None:
        path = si._find_template_path(si.TEMPLATE_VARIANT_GROUP_2_PLUS)
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            self.assertIn("%", ws["K19"].number_format)
        finally:
            wb.close()

    def test_group_block_rows_expand_from_single_template_block(self) -> None:
        self.assertEqual(si._section_2_group_block_start_row(0), 16)
        self.assertEqual(si._section_2_group_block_start_row(1), 77)
        self.assertEqual(si._section_2_group_block_start_row(2), 138)
        self.assertEqual(si._section_2_group_insert_row(1), 77)
        self.assertEqual(si._section_2_group_insert_row(2), 138)

    def test_group_export_title_for_offerentes_uses_expected_ranges(self) -> None:
        self.assertEqual(
            si._group_export_title_for_offerentes(1),
            "PROCESO DE SELECCION INCLUYENTE INDIVIDUAL",
        )
        self.assertEqual(
            si._group_export_title_for_offerentes(2),
            "PROCESO DE SELECCION INCLUYENTE GRUPAL - 2 A 4 OFERENTES",
        )
        self.assertEqual(
            si._group_export_title_for_offerentes(4),
            "PROCESO DE SELECCION INCLUYENTE GRUPAL - 2 A 4 OFERENTES",
        )
        self.assertEqual(
            si._group_export_title_for_offerentes(5),
            "PROCESO DE SELECCION INCLUYENTE GRUPAL - 5 A 7 OFERENTES",
        )
        self.assertEqual(
            si._group_export_title_for_offerentes(7),
            "PROCESO DE SELECCION INCLUYENTE GRUPAL - 5 A 7 OFERENTES",
        )
        self.assertEqual(
            si._group_export_title_for_offerentes(8),
            "PROCESO DE SELECCION INCLUYENTE GRUPAL - 8 A 10 OFERENTES",
        )
        self.assertEqual(
            si._group_export_title_for_offerentes(10),
            "PROCESO DE SELECCION INCLUYENTE GRUPAL - 8 A 10 OFERENTES",
        )
        self.assertEqual(
            si._group_export_title_for_offerentes(11),
            "PROCESO DE SELECCION INCLUYENTE GRUPAL - MAS DE 10 OFERENTES",
        )

    def test_normalize_excel_dropdown_value_maps_tipo_pension(self) -> None:
        value = si.normalize_excel_dropdown_value(
            "tipo_pension",
            "Regimen especial",
            template_variant=si.TEMPLATE_VARIANT_INDIVIDUAL,
        )
        self.assertEqual(value, "Pensión régimen especial (fuerzas militares)")

    def test_normalize_excel_dropdown_value_maps_transport_using_canonical_template_text(self) -> None:
        value = si.normalize_excel_dropdown_value(
            "desplazamiento_transporte",
            "Vehiculo especial.",
            template_variant=si.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(value, "Vehículo especial.")

    def test_all_section_2_list_fields_normalize_against_declared_canonical_options_for_each_variant(self) -> None:
        variants = (
            si.TEMPLATE_VARIANT_INDIVIDUAL,
            si.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        list_field_ids = [
            field["id"]
            for field in si.SECTION_2["fields"]
            if field.get("type") == "lista"
        ]

        for variant in variants:
            with self.subTest(variant=variant):
                for field_id in list_field_ids:
                    expected_options = tuple(si.LIST_FIELD_OPTIONS_BY_ID.get(field_id, []))
                    if not expected_options:
                        continue
                    canonical_options = self._expected_canonical_options(field_id, variant)
                    self.assertEqual(
                        len(canonical_options),
                        len(expected_options),
                        msg=f"field={field_id} variant={variant}",
                    )
                    for index, raw_value in enumerate(expected_options):
                        expected_value = canonical_options[index]
                        actual_value = si.normalize_excel_dropdown_value(
                            field_id,
                            raw_value,
                            template_variant=variant,
                        )
                        self.assertEqual(
                            actual_value,
                            expected_value,
                            msg=(
                                f"field={field_id} variant={variant} index={index} "
                                f"raw={raw_value!r}"
                            ),
                        )

    def test_section_5_and_section_6_rows_shift_only_by_inserted_group_blocks(self) -> None:
        self.assertEqual(
            si._section_row_after_section_2(77, 1, si.TEMPLATE_VARIANT_GROUP_2_PLUS),
            77,
        )
        self.assertEqual(
            si._section_row_after_section_2(77, 2, si.TEMPLATE_VARIANT_GROUP_2_PLUS),
            138,
        )
        self.assertEqual(
            si._section_row_after_section_2(83, 2, si.TEMPLATE_VARIANT_GROUP_2_PLUS),
            144,
        )
        self.assertEqual(
            si._section_row_after_section_2(77, 1, si.TEMPLATE_VARIANT_INDIVIDUAL),
            77,
        )

    def test_normalize_excel_dropdown_value_preserves_raw_when_no_match(self) -> None:
        value = si.normalize_excel_dropdown_value(
            "tipo_pension",
            "Valor imposible",
            template_variant=si.TEMPLATE_VARIANT_GROUP_2_PLUS,
        )
        self.assertEqual(value, "Valor imposible")


if __name__ == "__main__":
    unittest.main()
